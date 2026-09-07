#!/usr/bin/env python3
"""
Generate lifecycle performance reports (monthly or weekly) for BUR and CZ brands.

Pulls data from:
  - Braze Raw Events Datashare (Snowflake) for sends/deliveries/opens/clicks
  - GA4 Snowflake table for revenue and orders

Outputs an Excel file with tabs:
  Summary {label}   — 5-row rollup (B&B / SMS / Triggers / Grand Total)
  B&B {label}       — Long Tail + Batch & Blast sections with subtotals
  SMS {label}       — Per-SMS-campaign rows with total
  Triggers {label}  — Per-canvas-step rows grouped by canvas name
  (Same 4 tabs for LY = prior year same period)
  {label} WBR       — YoY comparison (Email / SMS / Triggered / Total)

Monthly usage:
    uv run python scripts/generate_lifecycle_report.py --month February --year 2026
    uv run python scripts/generate_lifecycle_report.py --brand CZ --month February --year 2026

Weekly usage (--week-ending must be a Sunday):
    uv run python scripts/generate_lifecycle_report.py --brand BUR --week-ending 2026-03-01
    uv run python scripts/generate_lifecycle_report.py --brand CZ --week-ending 2026-03-01
"""

import argparse
import calendar
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

scripts_dir = Path(__file__).parent
if str(scripts_dir) not in sys.path:
    sys.path.insert(0, str(scripts_dir))

from snowflake_client import get_snowflake_client

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BRAZE_DB = "BRAZE_BRAZEWEST_BRAZE_RAW_EVENTS_AWS_US_EAST_1_XJ24206"
BRAZE_SCHEMA = "DATALAKE_SHARING"
GA4_DB = os.environ.get("SNOWFLAKE_DATABASE", "AIRBYTE_DATABASE")

BRAND_CONFIG: Dict[str, Dict[str, Any]] = {
    "BUR": {
        "app_group_id": "67093a1f24ebbe0065cb9c77",
        "ga4_schema":   "LANDING_BURROW_GA4",
        "file_prefix":  "BW",    # BW_Lifecycle_Report_... for backward compat
        "has_sms":      True,
    },
    "CZ": {
        "app_group_id": "666672a4d8965b005ac6c1bd",
        "ga4_schema":   "LANDING_CITIZENRY_GA4",
        "file_prefix":  "CZ",
        "has_sms":      True,
    },
}

CAMPAIGNS_DIR = Path(__file__).parent.parent / "campaigns"


def _bv(view: str) -> str:
    return f"{BRAZE_DB}.{BRAZE_SCHEMA}.{view}"


EMAIL_SEND = _bv("USERS_MESSAGES_EMAIL_SEND_SHARED")
EMAIL_DELIVERY = _bv("USERS_MESSAGES_EMAIL_DELIVERY_SHARED")
EMAIL_OPEN = _bv("USERS_MESSAGES_EMAIL_OPEN_SHARED")
EMAIL_CLICK = _bv("USERS_MESSAGES_EMAIL_CLICK_SHARED")
SMS_SEND = _bv("USERS_MESSAGES_SMS_SEND_SHARED")
SMS_DELIVERY = _bv("USERS_MESSAGES_SMS_DELIVERY_SHARED")
CHANGELOGS_CAMPAIGN = _bv("CHANGELOGS_CAMPAIGN_SHARED")
CHANGELOGS_CANVAS = _bv("CHANGELOGS_CANVAS_SHARED")

# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def month_name_to_num(month_name: str) -> int:
    months = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
    num = months.get(month_name.lower())
    if num is None:
        raise ValueError(f"Unknown month: {month_name!r}")
    return num


def month_date_range(year: int, month: int) -> Tuple[str, str, str, str]:
    """Return (start_ts, end_ts, start_yyyymmdd, end_yyyymmdd) for Braze and GA4 queries."""
    last_day = calendar.monthrange(year, month)[1]
    start_dt = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    end_dt = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
    start_ts = start_dt.strftime("%Y-%m-%dT%H:%M:%S")
    end_ts = end_dt.strftime("%Y-%m-%dT%H:%M:%S")
    start_yyyymmdd = f"{year}{month:02d}01"
    end_yyyymmdd = f"{year}{month:02d}{last_day:02d}"
    return start_ts, end_ts, start_yyyymmdd, end_yyyymmdd


def _ts_to_unix(ts_str: str) -> int:
    """Convert 'YYYY-MM-DDTHH:MM:SS' UTC string to a unix timestamp integer.

    Braze event views store TIME as a plain integer (unix seconds).  Using raw
    integer comparisons — ``TIME >= {unix}`` — lets Snowflake exploit its
    micro-partition pruning on that column.  Wrapping with ``TO_TIMESTAMP(TIME)``
    is a function-on-column expression that defeats pruning and causes a full-table
    scan across years of event history.
    """
    dt = datetime.strptime(ts_str, "%Y-%m-%dT%H:%M:%S")
    return int(calendar.timegm(dt.timetuple()))


def week_date_range(week_ending: date) -> Tuple[str, str, str, str]:
    """Return (start_ts, end_ts, start_yyyymmdd, end_yyyymmdd) for a Mon–Sun week.

    week_ending must be a Sunday (weekday() == 6).
    """
    monday = week_ending - timedelta(days=6)
    start_dt = datetime(monday.year, monday.month, monday.day, 0, 0, 0, tzinfo=timezone.utc)
    end_dt   = datetime(week_ending.year, week_ending.month, week_ending.day, 23, 59, 59, tzinfo=timezone.utc)
    start_ts       = start_dt.strftime("%Y-%m-%dT%H:%M:%S")
    end_ts         = end_dt.strftime("%Y-%m-%dT%H:%M:%S")
    start_yyyymmdd = monday.strftime("%Y%m%d")
    end_yyyymmdd   = week_ending.strftime("%Y%m%d")
    return start_ts, end_ts, start_yyyymmdd, end_yyyymmdd


def ly_week_ending(week_ending: date) -> date:
    """Return the Sunday of the same ISO week number in the prior year."""
    iso_year, iso_week, _ = week_ending.isocalendar()
    try:
        return date.fromisocalendar(iso_year - 1, iso_week, 7)
    except ValueError:
        # Prior year has fewer ISO weeks (no week 53) — fall back to week 52
        return date.fromisocalendar(iso_year - 1, 52, 7)


def build_ym_patterns(start_ts: str, end_ts: str) -> List[str]:
    """Return list of 'YYYY_MM' strings covering the date range.

    Usually 1 element; 2 for a week that spans a month boundary
    (e.g. Feb 23–Mar 1 -> ['2026_02', '2026_03']).
    """
    start_ym = start_ts[:7].replace("-", "_")
    end_ym   = end_ts[:7].replace("-", "_")
    if start_ym == end_ym:
        return [start_ym]
    return [start_ym, end_ym]


def build_ym_filter_clause(patterns: List[str]) -> str:
    """Return SQL fragment matching campaign names against year-month patterns.

    Single pattern:  (NAME ILIKE '%2026_02%')
    Two patterns:    (NAME ILIKE '%2026_02%' OR NAME ILIKE '%2026_03%')
    """
    parts = [f"NAME ILIKE '%{p}%'" for p in patterns]
    return "(" + " OR ".join(parts) + ")"


def format_week_tab_label(week_ending: date) -> str:
    """Return e.g. '03_01' for a week ending March 1."""
    return week_ending.strftime("%m_%d")


# ---------------------------------------------------------------------------
# Period spec
# ---------------------------------------------------------------------------

@dataclass
class PeriodSpec:
    """Encapsulates all date/brand metadata for one reporting period."""
    brand: str           # "BUR" or "CZ"
    mode: str            # "monthly" or "weekly"
    start_ts: str        # ISO: "2026-02-01T00:00:00"
    end_ts: str          # ISO: "2026-02-28T23:59:59"
    start_yyyymmdd: str  # GA4: "20260201"
    end_yyyymmdd: str    # GA4: "20260228"
    tab_label: str       # e.g. "February 2026" or "03_01"
    year: int            # used for WBR column headers ("TY (2026)")


# ---------------------------------------------------------------------------
# Canvas step name lookup from local YAML files
# ---------------------------------------------------------------------------

def load_canvas_step_name_lookup() -> Dict[Tuple[str, str], str]:
    """Build {(canvas_braze_id, step_api_id): step_name} from parent canvas YAMLs."""
    lookup: Dict[Tuple[str, str], str] = {}
    for f in CAMPAIGNS_DIR.glob("*.yaml"):
        if f.name.startswith("_"):
            continue
        try:
            data = yaml.safe_load(open(f))
        except Exception:
            continue
        if not data or data.get("braze_type") != "canvas":
            continue
        canvas_id = data.get("braze_id") or data.get("canvas_id")
        if not canvas_id:
            continue
        for send in data.get("sends") or []:
            step_id = send.get("step_id")
            name = send.get("name") or send.get("step_name")
            if step_id and name:
                lookup[(canvas_id, step_id)] = name
    return lookup


def load_canvas_api_id_lookup() -> Dict[str, str]:
    """Build {canvas_name: canvas_braze_id} from parent canvas YAMLs."""
    lookup: Dict[str, str] = {}
    for f in CAMPAIGNS_DIR.glob("*.yaml"):
        if f.name.startswith("_"):
            continue
        try:
            data = yaml.safe_load(open(f))
        except Exception:
            continue
        if not data or data.get("braze_type") != "canvas":
            continue
        name = data.get("name") or data.get("canvas_name")
        braze_id = data.get("braze_id") or data.get("canvas_id")
        if name and braze_id:
            lookup[name] = braze_id
    return lookup



# ---------------------------------------------------------------------------
# Braze queries
# ---------------------------------------------------------------------------

def _get_braze_client():
    return get_snowflake_client(schema=BRAZE_SCHEMA, database=BRAZE_DB)


def query_braze_email_batch(
    client,
    start_ts: str,
    end_ts: str,
    app_group_id: str,
) -> List[Dict]:
    """Batch email campaigns (CANVAS_ID IS NULL) — delivered, opens, clicks.

    Warehouse-efficient design:
    - Raw integer TIME comparisons enable Snowflake micro-partition pruning.
    - Each event table is joined directly on CAMPAIGN_API_ID (from the tiny
      campaign_names CTE) plus a tight time window — no two-hop DISPATCH_ID lookups.
    - Name filter extended ±7 days beyond the reporting window to catch campaigns
      shifted by a few days. TIME filter is the authoritative gate.
    - Delivery buffer: 3 days past period end.
    - Opens / clicks buffer: 90 days past period end.

    Still-sending deferral (automatic, no pre-identification needed):
    - Any campaign whose last send falls after the period end is excluded.
      This handles Intelligent Timing (STO) and any other campaign still in-flight
      at the period boundary without requiring a pre-built UUID list or YAML flags.
    - campaign_windows scans EMAIL_SEND in [period_start - 7d, period_end + 3d] to
      find actual first/last send times for all candidate campaigns.
      - 7-day lookback: catches deferred campaigns that started up to a week before
        the period (matches the ±7-day name filter window).
      - 3-day lookahead: detects campaigns still sending after period end.
    - still_sending: campaigns with last_send > end_unix — excluded entirely.
    - deferred_complete: campaigns with first_send before the period but last_send
      within it (deferred from the previous period). Included with all sends counted
      from first_send_unix so no recipients are missed.
    - Regular campaigns (started and finished within the period) are unaffected.
    - opens/clicks lower bound extended to windows_start_unix so pre-period opens
      from deferred campaigns are captured via the in_window_senders semi-join.
    """
    # Widen name filter ±7 days to catch campaigns shifted by up to a week.
    _name_start = (datetime.strptime(start_ts, "%Y-%m-%dT%H:%M:%S") - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    _name_end   = (datetime.strptime(end_ts,   "%Y-%m-%dT%H:%M:%S") + timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    ym_clause = build_ym_filter_clause(build_ym_patterns(_name_start, _name_end))
    start_unix         = _ts_to_unix(start_ts)
    end_unix           = _ts_to_unix(end_ts)
    windows_start_unix = start_unix - 7 * 86400   # 7-day lookback for deferred campaigns
    windows_end_unix   = end_unix   + 3 * 86400   # 3-day lookahead to detect still-sending
    delivered_end_unix = end_unix   + 3 * 86400
    engage_end_unix    = end_unix   + 90 * 86400

    sql = f"""
    WITH campaign_names AS (
        -- Scoped to reporting period months ±7 days. Computed once; shared by all CTEs.
        SELECT API_ID, NAME
        FROM {CHANGELOGS_CAMPAIGN}
        WHERE {ym_clause}
          AND NAME NOT ILIKE 'Copy of %'
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    -- campaign_windows: actual first/last send time for each candidate campaign.
    -- Wider window (period ±7d back, +3d forward) detects still-sending and deferred campaigns.
    -- JOIN on campaign_names limits the scan to this period's candidates.
    campaign_windows AS (
        SELECT s.CAMPAIGN_API_ID,
               MIN(s.TIME) AS first_send_unix,
               MAX(s.TIME) AS last_send_unix
        FROM {EMAIL_SEND} s
        JOIN campaign_names cn ON s.CAMPAIGN_API_ID = cn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= {windows_start_unix}
          AND s.TIME <= {windows_end_unix}
        GROUP BY s.CAMPAIGN_API_ID
    ),
    -- still_sending: last send is after the period end — campaign is still in-flight.
    -- Excluded from this period; will appear in the next period once complete.
    still_sending AS (
        SELECT CAMPAIGN_API_ID
        FROM campaign_windows
        WHERE last_send_unix > {end_unix}
    ),
    -- deferred_complete: started before the period but finished within it.
    -- These were excluded from the previous period (still sending at that boundary).
    -- Count all sends from first_send_unix so pre-period recipients are included.
    deferred_complete AS (
        SELECT cw.CAMPAIGN_API_ID, cn.NAME, cw.first_send_unix
        FROM campaign_windows cw
        JOIN campaign_names cn ON cw.CAMPAIGN_API_ID = cn.API_ID
        WHERE cw.first_send_unix <  {start_unix}
          AND cw.last_send_unix  >= {start_unix}
          AND cw.last_send_unix  <= {end_unix}
    ),
    sends AS (
        -- Regular: started and finished within this period.
        SELECT cn.NAME AS campaign_name, COUNT(DISTINCT s.USER_ID) AS sends_count
        FROM {EMAIL_SEND} s
        JOIN campaign_names cn ON s.CAMPAIGN_API_ID = cn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
          AND s.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM still_sending)
          AND s.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM deferred_complete)
        GROUP BY cn.NAME
        UNION ALL
        -- Deferred-complete: all sends from campaign start through period end.
        SELECT dc.NAME AS campaign_name, COUNT(DISTINCT s.USER_ID) AS sends_count
        FROM {EMAIL_SEND} s
        JOIN deferred_complete dc ON s.CAMPAIGN_API_ID = dc.CAMPAIGN_API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= dc.first_send_unix
          AND s.TIME <= {end_unix}
        GROUP BY dc.NAME
    ),
    -- in_window_senders: (CAMPAIGN_API_ID, USER_ID) pairs for this period.
    -- Opens/clicks CTEs join here to restrict engagement to actual recipients.
    -- Batch DISPATCH_IDs are shared across all users in a send, so USER_ID join
    -- is correct; DISPATCH_ID join would fan out across the entire batch.
    in_window_senders AS (
        SELECT s.CAMPAIGN_API_ID, s.USER_ID
        FROM {EMAIL_SEND} s
        JOIN campaign_names cn ON s.CAMPAIGN_API_ID = cn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
          AND s.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM still_sending)
          AND s.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM deferred_complete)
        UNION ALL
        -- Deferred-complete: include pre-period senders so their opens/clicks are counted.
        SELECT s.CAMPAIGN_API_ID, s.USER_ID
        FROM {EMAIL_SEND} s
        JOIN deferred_complete dc ON s.CAMPAIGN_API_ID = dc.CAMPAIGN_API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= dc.first_send_unix
          AND s.TIME <= {end_unix}
    ),
    delivered AS (
        SELECT cn.NAME AS campaign_name, COUNT(DISTINCT d.USER_ID) AS delivered
        FROM {EMAIL_DELIVERY} d
        JOIN campaign_names cn ON d.CAMPAIGN_API_ID = cn.API_ID
        WHERE d.APP_GROUP_ID = '{app_group_id}'
          AND d.CANVAS_ID IS NULL
          AND d.TIME >= {start_unix} AND d.TIME <= {delivered_end_unix}
          AND d.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM still_sending)
          AND d.CAMPAIGN_API_ID NOT IN (SELECT CAMPAIGN_API_ID FROM deferred_complete)
        GROUP BY cn.NAME
        UNION ALL
        -- Deferred-complete: extend delivery window back to first_send_unix.
        SELECT dc.NAME AS campaign_name, COUNT(DISTINCT d.USER_ID) AS delivered
        FROM {EMAIL_DELIVERY} d
        JOIN deferred_complete dc ON d.CAMPAIGN_API_ID = dc.CAMPAIGN_API_ID
        WHERE d.APP_GROUP_ID = '{app_group_id}'
          AND d.CANVAS_ID IS NULL
          AND d.TIME >= dc.first_send_unix
          AND d.TIME <= {delivered_end_unix}
        GROUP BY dc.NAME
    ),
    opens AS (
        -- MACHINE_OPEN is 'true' for machine opens, NULL for human (no 'false' values).
        -- Lower bound uses windows_start_unix to capture pre-period opens for deferred
        -- campaigns; in_window_senders semi-join scopes them to the right recipients.
        SELECT cn.NAME AS campaign_name,
               COUNT(CASE WHEN o.MACHINE_OPEN = 'true' THEN 1 END)
                   + COUNT(CASE WHEN o.MACHINE_OPEN IS NULL THEN 1 END) AS total_opens,
               COUNT(DISTINCT o.USER_ID) AS unique_opens
        FROM {EMAIL_OPEN} o
        JOIN campaign_names cn ON o.CAMPAIGN_API_ID = cn.API_ID
        JOIN in_window_senders iws
          ON iws.CAMPAIGN_API_ID = o.CAMPAIGN_API_ID
         AND iws.USER_ID = o.USER_ID
        WHERE o.APP_GROUP_ID = '{app_group_id}'
          AND o.CANVAS_ID IS NULL
          AND o.TIME >= {windows_start_unix} AND o.TIME <= {engage_end_unix}
        GROUP BY cn.NAME
    ),
    clicks AS (
        SELECT cn.NAME AS campaign_name,
               COUNT(DISTINCT ck.ID) AS total_clicks,
               COUNT(DISTINCT CASE WHEN (ck.IS_SUSPECTED_BOT_CLICK IS NULL
                                         OR ck.IS_SUSPECTED_BOT_CLICK = 'false')
                               THEN ck.USER_ID END) AS unique_clicks
        FROM {EMAIL_CLICK} ck
        JOIN campaign_names cn ON ck.CAMPAIGN_API_ID = cn.API_ID
        JOIN in_window_senders iws
          ON iws.CAMPAIGN_API_ID = ck.CAMPAIGN_API_ID
         AND iws.USER_ID = ck.USER_ID
        WHERE ck.APP_GROUP_ID = '{app_group_id}'
          AND ck.CANVAS_ID IS NULL
          AND ck.TIME >= {windows_start_unix} AND ck.TIME <= {engage_end_unix}
        GROUP BY cn.NAME
    )
    SELECT
        s.campaign_name,
        COALESCE(d.delivered, 0)      AS delivered,
        COALESCE(o.total_opens, 0)    AS total_opens,
        COALESCE(o.unique_opens, 0)   AS unique_opens,
        COALESCE(c.total_clicks, 0)   AS total_clicks,
        COALESCE(c.unique_clicks, 0)  AS unique_clicks
    FROM sends s
    LEFT JOIN delivered d ON s.campaign_name = d.campaign_name
    LEFT JOIN opens o     ON s.campaign_name = o.campaign_name
    LEFT JOIN clicks c    ON s.campaign_name = c.campaign_name
    ORDER BY delivered DESC
    """
    rows = client.execute_query(sql)
    return [
        {
            "campaign_name": r.get("CAMPAIGN_NAME") or "",
            "delivered": int(r.get("DELIVERED") or 0),
            "total_opens": int(r.get("TOTAL_OPENS") or 0),
            "unique_opens": int(r.get("UNIQUE_OPENS") or 0),
            "total_clicks": int(r.get("TOTAL_CLICKS") or 0),
            "unique_clicks": int(r.get("UNIQUE_CLICKS") or 0),
        }
        for r in rows
    ]


CANVAS_STEP_NAMES = _bv("SNAPSHOTS_CANVAS_STEP_SHARED")


def query_braze_email_canvas(client, start_ts: str, end_ts: str, app_group_id: str) -> List[Dict]:
    """Canvas email steps — delivered, opens, clicks grouped by canvas + step.

    Joins SNAPSHOTS_CANVAS_STEP_SHARED to get step names (matching GA4
    SESSIONCAMPAIGNNAME UTM values) so revenue can be matched directly.

    Warehouse-efficient design:
    - Raw integer TIME comparisons for micro-partition pruning on all event tables.
    - step_dispatches CTE collects February DISPATCH_IDs from EMAIL_SEND (already
      time-filtered); delivery/opens/clicks join on DISPATCH_ID (exact per-user
      matching) + time windows (enables partition pruning on large event tables).
      Canvas DISPATCH_IDs are per-user (not shared across a batch), so there is
      no fan-out risk. CANVAS_API_ID-only joins were removed because always-on
      triggered canvases accumulate events from prior months on the same step,
      causing over-counting when filtered purely by delivery time.
    """
    start_unix         = _ts_to_unix(start_ts)
    end_unix           = _ts_to_unix(end_ts)
    delivered_end_unix = end_unix + 3 * 86400    # delivery lag < 3 days
    engage_end_unix    = end_unix + 90 * 86400   # opens/clicks within 90 days

    sql = f"""
    WITH canvas_names AS (
        SELECT API_ID, NAME
        FROM {CHANGELOGS_CANVAS}
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    step_names AS (
        SELECT API_ID, NAME
        FROM {CANVAS_STEP_NAMES}
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    -- sends: group by canvas+step for the summary row counts
    sends AS (
        SELECT cv.NAME AS canvas_name,
               s.CANVAS_STEP_API_ID AS step_api_id,
               COALESCE(sn.NAME, s.CANVAS_STEP_API_ID) AS step_name,
               COUNT(DISTINCT s.USER_ID) AS sends_count
        FROM {EMAIL_SEND} s
        JOIN canvas_names cv ON s.CANVAS_API_ID = cv.API_ID
        LEFT JOIN step_names sn ON s.CANVAS_STEP_API_ID = sn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NOT NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
        GROUP BY cv.NAME, s.CANVAS_STEP_API_ID, step_name
    ),
    -- step_dispatches: compact set of period DISPATCH_IDs per canvas step.
    -- Canvas DISPATCH_IDs are per-user (one per trigger), so no fan-out risk.
    -- Joining downstream event tables on DISPATCH_ID pins them to this period's
    -- sends only, avoiding bleed from prior periods on always-on canvases.
    step_dispatches AS (
        SELECT DISTINCT cv.NAME AS canvas_name,
                        s.CANVAS_STEP_API_ID AS step_api_id,
                        s.DISPATCH_ID
        FROM {EMAIL_SEND} s
        JOIN canvas_names cv ON s.CANVAS_API_ID = cv.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NOT NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
    ),
    delivered AS (
        -- DISPATCH_ID join naturally scopes to BUR Feb sends; time filter enables pruning.
        -- No APP_GROUP_ID or CANVAS_ID IS NOT NULL filter — original code omitted both,
        -- and CANVAS_ID can sometimes be NULL in delivery events for canvas sends.
        -- COUNT(DISTINCT d.ID) counts total delivery events, not unique users.
        -- Most canvases allow re-enrollment (Abandon Browse, Cart Abandon, Post Purchase,
        -- etc.) — only Welcome Series typically does not. A user who triggers the same
        -- canvas multiple times in a month receives each step again. Braze's own
        -- "Deliveries (Email)" metric counts each delivery event separately;
        -- COUNT(DISTINCT USER_ID) would undercount for any re-enrollable canvas.
        SELECT sd.canvas_name, sd.step_api_id,
               COUNT(DISTINCT d.ID) AS delivered
        FROM {EMAIL_DELIVERY} d
        JOIN step_dispatches sd ON d.DISPATCH_ID = sd.DISPATCH_ID
        WHERE d.TIME >= {start_unix} AND d.TIME <= {delivered_end_unix}
        GROUP BY sd.canvas_name, sd.step_api_id
    ),
    opens AS (
        -- Same methodology as batch: total_opens = machine events + human events (additive);
        -- unique_opens = distinct users with any open (machine OR human)
        SELECT sd.canvas_name, sd.step_api_id,
               COUNT(CASE WHEN o.MACHINE_OPEN = 'true' THEN 1 END)
                   + COUNT(CASE WHEN o.MACHINE_OPEN IS NULL THEN 1 END) AS total_opens,
               COUNT(DISTINCT o.USER_ID) AS unique_opens
        FROM {EMAIL_OPEN} o
        JOIN step_dispatches sd ON o.DISPATCH_ID = sd.DISPATCH_ID
        WHERE o.TIME >= {start_unix} AND o.TIME <= {engage_end_unix}
        GROUP BY sd.canvas_name, sd.step_api_id
    ),
    clicks AS (
        SELECT sd.canvas_name, sd.step_api_id,
               COUNT(DISTINCT ck.ID) AS total_clicks,
               COUNT(DISTINCT CASE WHEN (ck.IS_SUSPECTED_BOT_CLICK IS NULL
                                         OR ck.IS_SUSPECTED_BOT_CLICK = 'false')
                               THEN ck.USER_ID END) AS unique_clicks
        FROM {EMAIL_CLICK} ck
        JOIN step_dispatches sd ON ck.DISPATCH_ID = sd.DISPATCH_ID
        WHERE ck.TIME >= {start_unix} AND ck.TIME <= {engage_end_unix}
        GROUP BY sd.canvas_name, sd.step_api_id
    )
    SELECT
        s.canvas_name,
        s.step_api_id,
        s.step_name,
        COALESCE(d.delivered, 0)     AS delivered,
        COALESCE(o.total_opens, 0)   AS total_opens,
        COALESCE(o.unique_opens, 0)  AS unique_opens,
        COALESCE(c.total_clicks, 0)  AS total_clicks,
        COALESCE(c.unique_clicks, 0) AS unique_clicks
    FROM sends s
    LEFT JOIN delivered d ON s.canvas_name = d.canvas_name AND s.step_api_id = d.step_api_id
    LEFT JOIN opens o     ON s.canvas_name = o.canvas_name AND s.step_api_id = o.step_api_id
    LEFT JOIN clicks c    ON s.canvas_name = c.canvas_name AND s.step_api_id = c.step_api_id
    ORDER BY s.canvas_name, s.step_name
    """
    rows = client.execute_query(sql)
    return [
        {
            "canvas_name": r.get("CANVAS_NAME") or "",
            "step_api_id": r.get("STEP_API_ID") or "",
            "step_name": r.get("STEP_NAME") or r.get("STEP_API_ID") or "",
            "delivered": int(r.get("DELIVERED") or 0),
            "total_opens": int(r.get("TOTAL_OPENS") or 0),
            "unique_opens": int(r.get("UNIQUE_OPENS") or 0),
            "total_clicks": int(r.get("TOTAL_CLICKS") or 0),
            "unique_clicks": int(r.get("UNIQUE_CLICKS") or 0),
        }
        for r in rows
    ]


# Per-brand cache: None = unchecked, True/False = result of check
def query_braze_sms_batch(client, start_ts: str, end_ts: str, app_group_id: str) -> List[Dict]:
    """Batch SMS campaigns — sends and delivered counts.

    Clicks come from GA4 sessions (query_ga4_sms_sessions), not from the Braze
    shortlink view, which has sparse/unreliable data and high bot rates.

    Warehouse-efficiency: raw integer TIME comparisons for partition pruning,
    direct CAMPAIGN_API_ID joins, no DISPATCH_ID two-hop lookup.
    Name filter extended ±7 days (same rationale as query_braze_email_batch).
    """
    # Widen name filter ±7 days to catch campaigns shifted by up to a week.
    _name_start = (datetime.strptime(start_ts, "%Y-%m-%dT%H:%M:%S") - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    _name_end   = (datetime.strptime(end_ts,   "%Y-%m-%dT%H:%M:%S") + timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    ym_clause = build_ym_filter_clause(build_ym_patterns(_name_start, _name_end))
    start_unix         = _ts_to_unix(start_ts)
    end_unix           = _ts_to_unix(end_ts)
    delivered_end_unix = end_unix + 3 * 86400

    sql = f"""
    WITH campaign_names AS (
        SELECT API_ID, NAME
        FROM {CHANGELOGS_CAMPAIGN}
        WHERE {ym_clause}
          AND NAME NOT ILIKE 'Copy of %'
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    sends AS (
        SELECT cn.NAME AS campaign_name, COUNT(DISTINCT s.USER_ID) AS sends_count
        FROM {SMS_SEND} s
        JOIN campaign_names cn ON s.CAMPAIGN_API_ID = cn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
        GROUP BY cn.NAME
    ),
    delivered AS (
        SELECT cn.NAME AS campaign_name, COUNT(DISTINCT d.USER_ID) AS delivered
        FROM {SMS_DELIVERY} d
        JOIN campaign_names cn ON d.CAMPAIGN_API_ID = cn.API_ID
        WHERE d.APP_GROUP_ID = '{app_group_id}'
          AND d.CANVAS_ID IS NULL
          AND d.TIME >= {start_unix} AND d.TIME <= {delivered_end_unix}
        GROUP BY cn.NAME
    )
    SELECT
        s.campaign_name,
        COALESCE(d.delivered, 0) AS delivered,
        0                        AS total_clicks
    FROM sends s
    LEFT JOIN delivered d ON s.campaign_name = d.campaign_name
    ORDER BY delivered DESC
    """
    rows = client.execute_query(sql)
    return [
        {
            "campaign_name": r.get("CAMPAIGN_NAME") or "",
            "delivered": int(r.get("DELIVERED") or 0),
            "total_clicks": int(r.get("TOTAL_CLICKS") or 0),
        }
        for r in rows
    ]


def _sms_batch_from_api(start_ts: str, end_ts: str, brand: str) -> List[Dict]:
    """Fallback: fetch batch SMS campaign delivered counts from the Braze REST API.

    Used when the Braze datashare returns zero SMS rows because CAMPAIGN_API_ID is
    NULL in USERS_MESSAGES_SMS_SEND_SHARED (a known Braze datashare gap for some
    workspaces).

    Fast path: fetches the campaign list, filters to P_SMS_ names only, then pulls
    analytics for just those campaigns — avoids iterating all 700+ BW campaigns.

    Returns the same list-of-dicts format as query_braze_sms_batch:
        [{"campaign_name": str, "delivered": int, "total_clicks": int}, ...]
    total_clicks is always 0 here — build_sms_data overwrites it with GA4 sessions.
    """
    from braze_api_client import (
        get_all_campaigns, get_campaign_analytics, aggregate_campaign_analytics
    )
    api_key  = os.environ.get(f"BRAZE_API_KEY_{brand.upper()}") or os.environ.get("BRAZE_API_KEY")
    api_url  = os.environ.get("BRAZE_BASE_URL") or os.environ.get("BRAZE_API_URL")
    file_prefix = BRAND_CONFIG.get(brand.upper(), {}).get("file_prefix", brand)
    start_dt = datetime.strptime(start_ts, "%Y-%m-%dT%H:%M:%S")
    end_dt   = datetime.strptime(end_ts,   "%Y-%m-%dT%H:%M:%S")

    # Step 1: get full campaign list (including archived for LY historical data).
    # Two naming conventions exist:
    #   Modern  (Oct 2025+): P_SMS_YYYY_MM_DD_BW_...
    #   Legacy  (pre ~Oct 2025): YYYYMMDD_BW_..._SMS  (gradual switchover, no clean cutoff)
    all_campaigns = get_all_campaigns(api_key, api_url, brand=None, include_archived=True)
    def _is_sms_campaign(name: str, prefix: str) -> bool:
        n = name.upper()
        p = prefix.upper()
        if p not in n:
            return False
        # Modern convention: P_SMS_ prefix
        if "P_SMS_" in n:
            return True
        # Legacy convention: date-prefixed name containing SMS
        # e.g. 20250513_BW_Memorial_Day_Sale_Launch_SMS
        import re as _re
        if "SMS" in n and _re.match(r'^\d{6,8}', n):
            return True
        return False
    sms_campaigns = [c for c in all_campaigns if _is_sms_campaign(c.get("name", ""), file_prefix)]
    if not sms_campaigns:
        return []

    print(f"  SMS API: fetching analytics for {len(sms_campaigns)} P_SMS_ campaigns...", flush=True)

    # Step 2: fetch analytics only for the SMS campaigns
    result = []
    for c in sms_campaigns:
        try:
            analytics = get_campaign_analytics(c["id"], start_dt, end_dt, api_key, api_url)
            agg = aggregate_campaign_analytics(analytics, c["name"])
            delivered = 0
            for col in ("Confirmed Deliveries (SMS)", "Deliveries (Email)"):
                try:
                    v = int(float(agg.get(col, 0) or 0))
                    if v > 0:
                        delivered = v
                        break
                except (ValueError, TypeError):
                    pass
            if delivered > 0:
                result.append({
                    "campaign_name": c["name"].strip(),
                    "delivered":     delivered,
                    "total_clicks":  0,
                })
        except Exception:
            pass
    return result


def query_braze_sms_canvas(client, start_ts: str, end_ts: str, app_group_id: str) -> List[Dict]:
    """Canvas SMS steps — delivered counts.

    Clicks come from GA4 sessions (not tracked here — canvas SMS step names
    follow TRG_ naming, handled downstream in build_triggers_data).

    Warehouse-efficient design (mirrors query_braze_email_canvas):
    - Raw integer TIME comparisons for micro-partition pruning.
    - step_dispatches CTE collects period DISPATCH_IDs from SMS_SEND; delivery
      joins on DISPATCH_ID for exact per-send scoping and partition pruning.
    """
    start_unix         = _ts_to_unix(start_ts)
    end_unix           = _ts_to_unix(end_ts)
    delivered_end_unix = end_unix + 3 * 86400

    sql = f"""
    WITH canvas_names AS (
        SELECT API_ID, NAME
        FROM {CHANGELOGS_CANVAS}
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    step_names AS (
        SELECT API_ID, NAME
        FROM {CANVAS_STEP_NAMES}
        QUALIFY ROW_NUMBER() OVER (PARTITION BY API_ID ORDER BY TIME DESC) = 1
    ),
    -- sends: group by canvas+step for the summary row counts
    sends AS (
        SELECT cv.NAME AS canvas_name,
               s.CANVAS_STEP_API_ID AS step_api_id,
               COALESCE(sn.NAME, s.CANVAS_STEP_API_ID) AS step_name,
               COUNT(DISTINCT s.USER_ID) AS sends_count
        FROM {SMS_SEND} s
        JOIN canvas_names cv ON s.CANVAS_API_ID = cv.API_ID
        LEFT JOIN step_names sn ON s.CANVAS_STEP_API_ID = sn.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NOT NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
        GROUP BY cv.NAME, s.CANVAS_STEP_API_ID, step_name
    ),
    -- step_dispatches: compact set of period DISPATCH_IDs + canvas API IDs per step.
    -- Canvas DISPATCH_IDs are per-user (one per trigger), so no fan-out risk.
    step_dispatches AS (
        SELECT DISTINCT cv.NAME AS canvas_name,
                        cv.API_ID AS canvas_api_id,
                        s.CANVAS_STEP_API_ID AS step_api_id,
                        s.DISPATCH_ID
        FROM {SMS_SEND} s
        JOIN canvas_names cv ON s.CANVAS_API_ID = cv.API_ID
        WHERE s.APP_GROUP_ID = '{app_group_id}'
          AND s.CANVAS_ID IS NOT NULL
          AND s.TIME >= {start_unix} AND s.TIME <= {end_unix}
    ),
    delivered AS (
        -- DISPATCH_ID join naturally scopes to BUR Feb sends; time filter enables pruning.
        -- No APP_GROUP_ID or CANVAS_ID IS NOT NULL filter — CANVAS_ID can be NULL in
        -- delivery events even for canvas sends.
        -- COUNT(DISTINCT d.ID): same re-enrollment rationale as email canvas.
        -- Most canvases allow re-enrollment; Braze "Confirmed Deliveries (SMS)" counts
        -- total delivery events, not unique users. COUNT(DISTINCT USER_ID) would undercount.
        SELECT sd.canvas_name, sd.step_api_id,
               COUNT(DISTINCT d.ID) AS delivered
        FROM {SMS_DELIVERY} d
        JOIN step_dispatches sd ON d.DISPATCH_ID = sd.DISPATCH_ID
        WHERE d.TIME >= {start_unix} AND d.TIME <= {delivered_end_unix}
        GROUP BY sd.canvas_name, sd.step_api_id
    )
    SELECT
        s.canvas_name,
        s.step_api_id,
        s.step_name,
        COALESCE(d.delivered, 0) AS delivered,
        0                        AS total_clicks
    FROM sends s
    LEFT JOIN delivered d ON s.canvas_name = d.canvas_name AND s.step_api_id = d.step_api_id
    ORDER BY s.canvas_name, s.step_name
    """
    rows = client.execute_query(sql)
    return [
        {
            "canvas_name": r.get("CANVAS_NAME") or "",
            "step_api_id": r.get("STEP_API_ID") or "",
            "step_name": r.get("STEP_NAME") or r.get("STEP_API_ID") or "",
            "delivered": int(r.get("DELIVERED") or 0),
            "total_clicks": int(r.get("TOTAL_CLICKS") or 0),
        }
        for r in rows
    ]


def query_ga4_revenue(
    start_yyyymmdd: str, end_yyyymmdd: str, brand: str, channel: Optional[str] = None
) -> Dict[str, Dict]:
    """GA4: campaigns with revenue in the period. Returns {lower_name: {orders, revenue}}.

    Args:
        brand:   Brand code (e.g. 'BUR', 'CZ') — determines which GA4 schema to query.
        channel: If given, filter SESSIONPRIMARYCHANNELGROUP (case-insensitive).
                 E.g. 'Email' for email attribution, 'SMS' for SMS attribution.
    """
    ga4_schema = BRAND_CONFIG[brand]["ga4_schema"]
    ga4_table  = f"{GA4_DB}.{ga4_schema}.TRAFFIC_SESSION_PERFORMANCE_DAILY"
    client = get_snowflake_client(schema=ga4_schema)
    channel_clause = (
        f"AND UPPER(TRIM(SESSIONPRIMARYCHANNELGROUP)) = '{channel.upper()}'"
        if channel
        else ""
    )
    sql = f"""
    SELECT
        LOWER(TRIM(SESSIONCAMPAIGNNAME))  AS campaign_name_lower,
        SUM(ECOMMERCEPURCHASES)           AS orders,
        SUM(TOTALREVENUE)                 AS revenue
    FROM {ga4_table}
    WHERE DATE >= '{start_yyyymmdd}'
      AND DATE <= '{end_yyyymmdd}'
      AND SESSIONCAMPAIGNNAME IS NOT NULL
      AND TRIM(SESSIONCAMPAIGNNAME) != ''
      {channel_clause}
    GROUP BY LOWER(TRIM(SESSIONCAMPAIGNNAME))
    HAVING SUM(ECOMMERCEPURCHASES) > 0 OR SUM(TOTALREVENUE) > 0
    """
    rows = client.execute_query(sql)
    client.close()
    return {
        (r.get("CAMPAIGN_NAME_LOWER") or ""): {
            "orders": int(r.get("ORDERS") or 0),
            "revenue": float(r.get("REVENUE") or 0.0),
        }
        for r in rows
    }


def query_ga4_sms_sessions(
    start_yyyymmdd: str, end_yyyymmdd: str, brand: str
) -> Dict[str, int]:
    """GA4: session counts for P_SMS_ campaigns, used as a click proxy.

    Uses LEFT(..., 6) = 'P_SMS_' for exact prefix matching (LIKE would treat
    the underscore as a single-char wildcard). No channel filter — SMS sessions
    can land under 'Direct' or other GA4 channel groups depending on UTM parsing.
    No HAVING filter — we want counts even for campaigns with no revenue.
    """
    ga4_schema = BRAND_CONFIG[brand]["ga4_schema"]
    ga4_table  = f"{GA4_DB}.{ga4_schema}.TRAFFIC_SESSION_PERFORMANCE_DAILY"
    client = get_snowflake_client(schema=ga4_schema)
    sql = f"""
    SELECT
        LOWER(TRIM(SESSIONCAMPAIGNNAME)) AS campaign_name_lower,
        SUM(SESSIONS)                    AS sessions
    FROM {ga4_table}
    WHERE DATE >= '{start_yyyymmdd}'
      AND DATE <= '{end_yyyymmdd}'
      AND SESSIONCAMPAIGNNAME IS NOT NULL
      AND TRIM(SESSIONCAMPAIGNNAME) != ''
      AND LEFT(UPPER(TRIM(SESSIONCAMPAIGNNAME)), 6) = 'P_SMS_'
    GROUP BY LOWER(TRIM(SESSIONCAMPAIGNNAME))
    """
    rows = client.execute_query(sql)
    client.close()
    return {
        (r.get("CAMPAIGN_NAME_LOWER") or ""): int(r.get("SESSIONS") or 0)
        for r in rows
    }


# ---------------------------------------------------------------------------
# Metric helpers
# ---------------------------------------------------------------------------

def safe_div(numerator, denominator) -> Optional[float]:
    if denominator and float(denominator) > 0:
        return float(numerator) / float(denominator)
    return None


def _merge_ga4(name: str, ga4: Dict[str, Dict]) -> Tuple[int, float]:
    ga4_data = ga4.get(name.lower().strip(), {})
    return ga4_data.get("orders", 0), ga4_data.get("revenue", 0.0)


def add_email_computed(row: Dict) -> None:
    d = row.get("delivered", 0)
    uo = row.get("unique_opens", 0)
    to_ = row.get("total_opens", 0)
    uc = row.get("unique_clicks", 0)
    tc = row.get("total_clicks", 0)
    orders = row.get("orders", 0)
    rev = row.get("revenue", 0.0)
    row["total_open_rate"] = safe_div(to_, d)
    row["unique_open_rate"] = safe_div(uo, d)
    row["total_ctr"] = safe_div(tc, d)
    row["unique_ctr"] = safe_div(uc, d)
    row["unique_cto"] = safe_div(uc, uo)
    row["conv_rate"] = safe_div(orders, uc)
    row["aov"] = safe_div(rev, orders)
    row["dpm"] = safe_div(rev, d) * 1000 if d else None


def add_sms_computed(row: Dict) -> None:
    d = row.get("delivered", 0)
    tc = row.get("total_clicks", 0)
    orders = row.get("orders", 0)
    rev = row.get("revenue", 0.0)
    row["conv_rate"] = safe_div(orders, tc)
    row["aov"] = safe_div(rev, orders)
    row["dpm"] = safe_div(rev, d) * 1000 if d else None


# ---------------------------------------------------------------------------
# Data building
# ---------------------------------------------------------------------------

def _is_canvas_step_name(name: str) -> bool:
    """Return True if the GA4 campaign name looks like a canvas step (TRG_/OT_EM_)."""
    n = name.upper()
    return n.startswith("TRG_") or n.startswith("OT_EM_")


def build_bb_data(
    braze_email: List[Dict],
    ga4: Dict[str, Dict],
) -> Tuple[List[Dict], List[Dict]]:
    """Split into (long_tail, batch_blast).
    Long Tail  = appears in GA4 with revenue but no Braze sends this period,
                 and is NOT a canvas step name (TRG_/OT_).
    Batch Blast = has Braze sends this period (with any GA4 revenue merged).
    """
    braze_names_lower = {r["campaign_name"].lower().strip() for r in braze_email}

    batch_blast = []
    for row in braze_email:
        row = dict(row)
        row["orders"], row["revenue"] = _merge_ga4(row["campaign_name"], ga4)
        add_email_computed(row)
        batch_blast.append(row)

    long_tail = []
    for name_lower, ga4_data in ga4.items():
        if name_lower in braze_names_lower:
            continue
        # Skip canvas step names — those belong in Triggers tab, not Long Tail
        if _is_canvas_step_name(name_lower):
            continue
        row = {
            "campaign_name": name_lower,
            "delivered": 0,
            "total_opens": 0,
            "unique_opens": 0,
            "total_clicks": 0,
            "unique_clicks": 0,
            "orders": ga4_data.get("orders", 0),
            "revenue": ga4_data.get("revenue", 0.0),
        }
        add_email_computed(row)
        long_tail.append(row)

    long_tail.sort(key=lambda x: x.get("revenue", 0), reverse=True)
    batch_blast.sort(key=lambda x: x.get("delivered", 0), reverse=True)
    return long_tail, batch_blast


def build_sms_data(
    braze_sms: List[Dict],
    ga4: Dict[str, Dict],
    ga4_sms_sessions: Dict[str, int],
) -> List[Dict]:
    result = []
    for row in braze_sms:
        row = dict(row)
        row["orders"], row["revenue"] = _merge_ga4(row["campaign_name"], ga4)
        # Derive clicks from GA4 sessions (UTM-attributed) rather than Braze
        # shortlink clicks, which are unreliable (sparse data, high bot rate).
        row["total_clicks"] = ga4_sms_sessions.get(row["campaign_name"].lower().strip(), 0)
        add_sms_computed(row)
        result.append(row)
    return result


def build_triggers_data(
    email_canvas: List[Dict],
    sms_canvas: List[Dict],
    ga4: Dict[str, Dict],
    step_name_lookup: Dict[Tuple[str, str], str],  # kept for signature compat, unused
    canvas_api_id_lookup: Dict[str, str],            # kept for signature compat, unused
) -> Dict[str, List[Dict]]:
    """Build {canvas_name: [step_rows]} for Triggers tab.

    Step names come directly from SNAPSHOTS_CANVAS_STEP_SHARED (via the Braze query),
    which matches the GA4 SESSIONCAMPAIGNNAME UTM value.
    """
    canvases: Dict[str, List[Dict]] = {}
    # Track step names that have already received GA4 attribution to prevent
    # double-counting when two steps share the same name (e.g. T3_V3 named twice).
    _ga4_attributed: set = set()

    for row in email_canvas:
        canvas_name = row["canvas_name"]
        step_api_id = row["step_api_id"]
        step_name = row.get("step_name") or step_api_id
        _key = step_name.lower().strip()
        if _key in _ga4_attributed:
            orders, revenue = 0, 0.0
        else:
            orders, revenue = _merge_ga4(step_name, ga4)
            _ga4_attributed.add(_key)
        step_row = {
            "step_name": step_name,
            "step_api_id": step_api_id,
            "channel": "email",
            "delivered": row["delivered"],
            "sms_delivered": 0,
            "total_opens": row["total_opens"],
            "unique_opens": row["unique_opens"],
            "total_clicks": row["total_clicks"],
            "unique_clicks": row["unique_clicks"],
            "sms_clicks": 0,
            "orders": orders,
            "revenue": revenue,
        }
        add_email_computed(step_row)
        canvases.setdefault(canvas_name, []).append(step_row)

    for row in sms_canvas:
        canvas_name = row["canvas_name"]
        step_api_id = row["step_api_id"]
        step_name = row.get("step_name") or step_api_id
        existing = None
        if canvas_name in canvases:
            for s in canvases[canvas_name]:
                if s["step_api_id"] == step_api_id:
                    existing = s
                    break
        if existing:
            existing["sms_delivered"] = row["delivered"]
            existing["sms_clicks"] = row["total_clicks"]
        else:
            _key = step_name.lower().strip()
            if _key in _ga4_attributed:
                orders, revenue = 0, 0.0
            else:
                orders, revenue = _merge_ga4(step_name, ga4)
                _ga4_attributed.add(_key)
            step_row = {
                "step_name": step_name,
                "step_api_id": step_api_id,
                "channel": "sms",
                "delivered": 0,
                "sms_delivered": row["delivered"],
                "total_opens": 0,
                "unique_opens": 0,
                "total_clicks": 0,
                "unique_clicks": 0,
                "sms_clicks": row["total_clicks"],
                "orders": orders,
                "revenue": revenue,
            }
            add_email_computed(step_row)
            canvases.setdefault(canvas_name, []).append(step_row)

    return canvases


# ---------------------------------------------------------------------------
# Subtotals
# ---------------------------------------------------------------------------

def subtotal_email(rows: List[Dict]) -> Dict:
    st = {
        "delivered": sum(r.get("delivered", 0) for r in rows),
        "total_opens": sum(r.get("total_opens", 0) for r in rows),
        "unique_opens": sum(r.get("unique_opens", 0) for r in rows),
        "total_clicks": sum(r.get("total_clicks", 0) for r in rows),
        "unique_clicks": sum(r.get("unique_clicks", 0) for r in rows),
        "orders": sum(r.get("orders", 0) for r in rows),
        "revenue": sum(r.get("revenue", 0.0) for r in rows),
    }
    add_email_computed(st)
    return st


def subtotal_sms(rows: List[Dict]) -> Dict:
    st = {
        "delivered": sum(r.get("delivered", 0) for r in rows),
        "total_clicks": sum(r.get("total_clicks", 0) for r in rows),
        "orders": sum(r.get("orders", 0) for r in rows),
        "revenue": sum(r.get("revenue", 0.0) for r in rows),
    }
    add_sms_computed(st)
    return st


def subtotal_canvas(steps: List[Dict]) -> Dict:
    st = {
        "delivered": sum(s.get("delivered", 0) for s in steps),
        "sms_delivered": sum(s.get("sms_delivered", 0) for s in steps),
        "total_opens": sum(s.get("total_opens", 0) for s in steps),
        "unique_opens": sum(s.get("unique_opens", 0) for s in steps),
        "total_clicks": sum(s.get("total_clicks", 0) for s in steps),
        "unique_clicks": sum(s.get("unique_clicks", 0) for s in steps),
        "sms_clicks": sum(s.get("sms_clicks", 0) for s in steps),
        "orders": sum(s.get("orders", 0) for s in steps),
        "revenue": sum(s.get("revenue", 0.0) for s in steps),
    }
    add_email_computed(st)
    return st


def build_summary(
    long_tail: List[Dict],
    batch_blast: List[Dict],
    sms_rows: List[Dict],
    canvases: Dict[str, List[Dict]],
) -> Dict[str, Any]:
    all_bb = long_tail + batch_blast
    bb_total = subtotal_email(all_bb)
    sms_total = subtotal_sms(sms_rows)
    all_steps = [s for steps in canvases.values() for s in steps]
    trig_total = subtotal_canvas(all_steps)

    grand_delivered = bb_total["delivered"] + sms_total["delivered"] + trig_total["delivered"]
    grand_revenue = bb_total["revenue"] + sms_total["revenue"] + trig_total["revenue"]
    grand_orders = bb_total["orders"] + sms_total["orders"] + trig_total["orders"]
    grand_unique_clicks = (
        bb_total["unique_clicks"]
        + sms_total.get("total_clicks", 0)
        + trig_total["unique_clicks"]
    )
    grand_total: Dict[str, Any] = {
        "delivered": grand_delivered,
        "sms_delivered": trig_total.get("sms_delivered", 0),
        "total_opens": bb_total["total_opens"] + trig_total["total_opens"],
        "unique_opens": bb_total["unique_opens"] + trig_total["unique_opens"],
        "total_clicks": (
            bb_total["total_clicks"]
            + sms_total.get("total_clicks", 0)
            + trig_total["total_clicks"]
        ),
        "unique_clicks": grand_unique_clicks,
        "sms_clicks": trig_total.get("sms_clicks", 0),
        "orders": grand_orders,
        "revenue": grand_revenue,
        "aov": safe_div(grand_revenue, grand_orders),
        "dpm": safe_div(grand_revenue, grand_delivered) * 1000 if grand_delivered else None,
        "conv_rate": safe_div(grand_orders, grand_unique_clicks),
    }

    return {
        "bb": bb_total,
        "sms": sms_total,
        "triggers": trig_total,
        "grand_total": grand_total,
    }


# ---------------------------------------------------------------------------
# Run for one period (monthly or weekly)
# ---------------------------------------------------------------------------

def run_for_period(
    spec: PeriodSpec,
    step_name_lookup: Dict[Tuple[str, str], str],
    canvas_api_id_lookup: Dict[str, str],
) -> Dict[str, Any]:
    cfg          = BRAND_CONFIG[spec.brand]
    app_group_id = cfg["app_group_id"]
    has_sms      = cfg["has_sms"]

    print(f"  Braze range : {spec.start_ts} -> {spec.end_ts}")
    print(f"  GA4 range   : {spec.start_yyyymmdd} -> {spec.end_yyyymmdd}")

    braze_client = _get_braze_client()
    try:
        print("  Querying Braze email batch ...", end="", flush=True)
        braze_email = query_braze_email_batch(braze_client, spec.start_ts, spec.end_ts, app_group_id)
        print(f" {len(braze_email)} campaigns")

        print("  Querying Braze email canvas...", end="", flush=True)
        email_canvas = query_braze_email_canvas(braze_client, spec.start_ts, spec.end_ts, app_group_id)
        print(f" {len(email_canvas)} steps")

        if has_sms:
            print("  Querying Braze SMS batch    ...", end="", flush=True)
            braze_sms = query_braze_sms_batch(braze_client, spec.start_ts, spec.end_ts, app_group_id)
            print(f" {len(braze_sms)} campaigns")
            if len(braze_sms) == 0:
                print("  SMS datashare returned 0 (CAMPAIGN_API_ID likely NULL) -- falling back to Braze API...", flush=True)
                braze_sms = _sms_batch_from_api(spec.start_ts, spec.end_ts, spec.brand)
                print(f"  SMS API fallback: {len(braze_sms)} campaigns")

            print("  Querying Braze SMS canvas   ...", end="", flush=True)
            sms_canvas = query_braze_sms_canvas(braze_client, spec.start_ts, spec.end_ts, app_group_id)
            print(f" {len(sms_canvas)} steps")
        else:
            braze_sms, sms_canvas = [], []
    finally:
        braze_client.close()

    # GA4 revenue strategy:
    # - B&B email (batch_blast + long_tail): use channel="Email" so Long Tail only includes
    #   email-attributed GA4 campaigns (prevents paid/organic campaigns from inflating totals).
    # - SMS batch + all triggers: use no channel filter because Braze SMS link clicks are
    #   often attributed to non-"SMS" channel groups in GA4 (GA4 may classify the resulting
    #   session as Direct/Email/Other depending on UTM parameters). Matching by campaign name
    #   alone is sufficient for P_SMS_ and TRG_ campaign names, which are unique to
    #   email/SMS traffic.
    print("  Querying GA4 revenue (email)...", end="", flush=True)
    ga4_email = query_ga4_revenue(spec.start_yyyymmdd, spec.end_yyyymmdd, spec.brand, channel="Email")
    print(f" {len(ga4_email)} records")

    print("  Querying GA4 revenue (all)  ...", end="", flush=True)
    ga4_all = query_ga4_revenue(spec.start_yyyymmdd, spec.end_yyyymmdd, spec.brand)
    print(f" {len(ga4_all)} records")

    ga4_sms_sessions: Dict[str, int] = {}
    if cfg["has_sms"]:
        print("  Querying GA4 SMS sessions   ...", end="", flush=True)
        ga4_sms_sessions = query_ga4_sms_sessions(spec.start_yyyymmdd, spec.end_yyyymmdd, spec.brand)
        print(f" {len(ga4_sms_sessions)} campaigns")

    long_tail, batch_blast = build_bb_data(braze_email, ga4_email)
    sms_rows = build_sms_data(braze_sms, ga4_all, ga4_sms_sessions)
    canvases = build_triggers_data(
        email_canvas, sms_canvas, ga4_all, step_name_lookup, canvas_api_id_lookup
    )
    summary = build_summary(long_tail, batch_blast, sms_rows, canvases)

    return {
        "year": spec.year,
        "tab_label": spec.tab_label,
        "long_tail": long_tail,
        "batch_blast": batch_blast,
        "sms_rows": sms_rows,
        "canvases": canvases,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

BB_COLS = [
    ("Campaign Name", "campaign_name", None),
    ("Delivered", "delivered", "#,##0"),
    ("Total Opens", "total_opens", "#,##0"),
    ("Unique Opens", "unique_opens", "#,##0"),
    ("Total Open Rate", "total_open_rate", "0.00%"),
    ("Unique Open Rate", "unique_open_rate", "0.00%"),
    ("Total Clicks", "total_clicks", "#,##0"),
    ("Unique Clicks", "unique_clicks", "#,##0"),
    ("Total CTR", "total_ctr", "0.00%"),
    ("Unique CTR", "unique_ctr", "0.00%"),
    ("Unique CTO", "unique_cto", "0.00%"),
    ("Orders", "orders", "#,##0"),
    ("Revenue", "revenue", "$#,##0.00"),
    ("AOV", "aov", "$#,##0.00"),
    ("Conv Rate", "conv_rate", "0.00%"),
    ("$/M", "dpm", "$#,##0.00"),
]

SMS_COLS = [
    ("Campaign Name", "campaign_name", None),
    ("Delivered", "delivered", "#,##0"),
    ("Total Clicks", "total_clicks", "#,##0"),
    ("Orders", "orders", "#,##0"),
    ("Revenue", "revenue", "$#,##0.00"),
    ("AOV", "aov", "$#,##0.00"),
    ("Conv Rate", "conv_rate", "0.00%"),
    ("$/M", "dpm", "$#,##0.00"),
]

TRIGGERS_COLS = [
    ("Step Name", "step_name", None),
    ("Delivered", "delivered", "#,##0"),
    ("Confirmed Deliveries (SMS)", "sms_delivered", "#,##0"),
    ("Total Opens", "total_opens", "#,##0"),
    ("Unique Opens", "unique_opens", "#,##0"),
    ("Total Open Rate", "total_open_rate", "0.00%"),
    ("Unique Open Rate", "unique_open_rate", "0.00%"),
    ("Total Clicks", "total_clicks", "#,##0"),
    ("Unique Clicks", "unique_clicks", "#,##0"),
    ("Total CTR", "total_ctr", "0.00%"),
    ("Unique CTR", "unique_ctr", "0.00%"),
    ("Unique CTO", "unique_cto", "0.00%"),
    ("SMS Clicks", "sms_clicks", "#,##0"),
    ("Orders", "orders", "#,##0"),
    ("Revenue", "revenue", "$#,##0.00"),
    ("AOV", "aov", "$#,##0.00"),
    ("Conv Rate", "conv_rate", "0.00%"),
    ("$/M", "dpm", "$#,##0.00"),
]

SUMMARY_COLS = [
    ("Section", None, None),
    ("Delivered", "delivered", "#,##0"),
    ("Confirmed Deliveries (SMS)", "sms_delivered", "#,##0"),
    ("Total Opens", "total_opens", "#,##0"),
    ("Unique Opens", "unique_opens", "#,##0"),
    ("Total Open Rate", "total_open_rate", "0.00%"),
    ("Unique Open Rate", "unique_open_rate", "0.00%"),
    ("Total Clicks", "total_clicks", "#,##0"),
    ("Unique Clicks", "unique_clicks", "#,##0"),
    ("Total CTR", "total_ctr", "0.00%"),
    ("Unique CTR", "unique_ctr", "0.00%"),
    ("Unique CTO", "unique_cto", "0.00%"),
    ("SMS Clicks", "sms_clicks", "#,##0"),
    ("Orders", "orders", "#,##0"),
    ("Revenue", "revenue", "$#,##0.00"),
    ("AOV", "aov", "$#,##0.00"),
    ("Conv Rate", "conv_rate", "0.00%"),
    ("$/M", "dpm", "$#,##0.00"),
]

SUMMARY_ROWS_DEF = [
    ("B&B", "bb"),
    ("SMS", "sms"),
    ("Triggers", "triggers"),
    ("Grand Total", "grand_total"),
]


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

def _write_header(ws, cols, row_num: int = 1) -> None:
    from openpyxl.styles import Font
    for col_idx, (header, _key, _fmt) in enumerate(cols, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True)


def _write_row(
    ws, row_num: int, cols, data: Dict, label: str = None, bold: bool = False
) -> None:
    from openpyxl.styles import Font
    for col_idx, (header, key, fmt) in enumerate(cols, start=1):
        if col_idx == 1 and label is not None:
            val = label
        elif key is None:
            val = None
        else:
            val = data.get(key)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        if fmt and val is not None:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)


# ---------------------------------------------------------------------------
# Tab writers
# ---------------------------------------------------------------------------

def write_summary_tab(wb, data: Dict, tab_label: str) -> None:
    ws = wb.create_sheet(title=f"Summary {tab_label}")
    _write_header(ws, SUMMARY_COLS)
    summary = data["summary"]
    for row_idx, (label, key) in enumerate(SUMMARY_ROWS_DEF, start=2):
        row_data = summary.get(key, {})
        _write_row(ws, row_idx, SUMMARY_COLS, row_data, label=label, bold=(label == "Grand Total"))


def write_bb_tab(wb, data: Dict, tab_label: str) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet(title=f"B&B {tab_label}")
    _write_header(ws, BB_COLS)
    row_num = 2
    long_tail = data["long_tail"]
    batch_blast = data["batch_blast"]

    cell = ws.cell(row=row_num, column=1, value="— Long Tail —")
    cell.font = Font(bold=True, italic=True)
    row_num += 1
    for c in long_tail:
        _write_row(ws, row_num, BB_COLS, c)
        row_num += 1
    _write_row(ws, row_num, BB_COLS, subtotal_email(long_tail), label="Subtotal — Long Tail", bold=True)
    row_num += 2

    cell = ws.cell(row=row_num, column=1, value="— Batch & Blast —")
    cell.font = Font(bold=True, italic=True)
    row_num += 1
    for c in batch_blast:
        _write_row(ws, row_num, BB_COLS, c)
        row_num += 1
    _write_row(ws, row_num, BB_COLS, subtotal_email(batch_blast), label="Subtotal — Batch & Blast", bold=True)
    row_num += 2

    _write_row(ws, row_num, BB_COLS, subtotal_email(long_tail + batch_blast), label="Total", bold=True)


def write_sms_tab(wb, data: Dict, tab_label: str) -> None:
    ws = wb.create_sheet(title=f"SMS {tab_label}")
    _write_header(ws, SMS_COLS)
    sms_rows = data["sms_rows"]
    for row_idx, c in enumerate(sms_rows, start=2):
        _write_row(ws, row_idx, SMS_COLS, c)
    _write_row(ws, len(sms_rows) + 2, SMS_COLS, subtotal_sms(sms_rows), label="Total", bold=True)


def write_triggers_tab(wb, data: Dict, tab_label: str) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet(title=f"Triggers {tab_label}")
    _write_header(ws, TRIGGERS_COLS)
    row_num = 2
    all_steps_flat: List[Dict] = []

    for canvas_name, steps in sorted(data["canvases"].items()):
        cell = ws.cell(row=row_num, column=1, value=canvas_name)
        cell.font = Font(bold=True, italic=True)
        row_num += 1
        for step in steps:
            _write_row(ws, row_num, TRIGGERS_COLS, step)
            row_num += 1
            all_steps_flat.append(step)
        _write_row(
            ws, row_num, TRIGGERS_COLS,
            subtotal_canvas(steps),
            label=f"Subtotal — {canvas_name}",
            bold=True,
        )
        row_num += 2  # blank row between canvases

    _write_row(ws, row_num, TRIGGERS_COLS, subtotal_canvas(all_steps_flat), label="Total", bold=True)


def write_wbr_tab(wb, ty_data: Dict, ly_data: Dict, tab_label: str) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet(title=f"{tab_label} WBR")

    def var(ty_val, ly_val):
        if ly_val and float(ly_val) > 0 and ty_val is not None:
            return float(ty_val) / float(ly_val) - 1
        return None

    ty_s = ty_data["summary"]
    ly_s = ly_data["summary"]
    ty_bb = ty_s.get("bb", {})
    ly_bb = ly_s.get("bb", {})
    ty_sms = ty_s.get("sms", {})
    ly_sms = ly_s.get("sms", {})
    ty_trig = ty_s.get("triggers", {})
    ly_trig = ly_s.get("triggers", {})
    ty_gt = ty_s.get("grand_total", {})
    ly_gt = ly_s.get("grand_total", {})

    # Email revenue = Grand Total - Triggered - SMS (residual, per reference)
    ty_email_rev = ty_gt.get("revenue", 0) - ty_trig.get("revenue", 0) - ty_sms.get("revenue", 0)
    ly_email_rev = ly_gt.get("revenue", 0) - ly_trig.get("revenue", 0) - ly_sms.get("revenue", 0)
    ty_email_dpm = safe_div(ty_email_rev, ty_bb.get("delivered", 0)) * 1000 if ty_bb.get("delivered") else None
    ly_email_dpm = safe_div(ly_email_rev, ly_bb.get("delivered", 0)) * 1000 if ly_bb.get("delivered") else None

    ty_sms_dpm = safe_div(ty_sms.get("revenue", 0), ty_sms.get("delivered", 0)) * 1000 if ty_sms.get("delivered") else None
    ly_sms_dpm = safe_div(ly_sms.get("revenue", 0), ly_sms.get("delivered", 0)) * 1000 if ly_sms.get("delivered") else None

    ty_trig_dpm = safe_div(ty_trig.get("revenue", 0), ty_trig.get("delivered", 0)) * 1000 if ty_trig.get("delivered") else None
    ly_trig_dpm = safe_div(ly_trig.get("revenue", 0), ly_trig.get("delivered", 0)) * 1000 if ly_trig.get("delivered") else None

    ty_tot_del = ty_bb.get("delivered", 0) + ty_sms.get("delivered", 0) + ty_trig.get("delivered", 0)
    ly_tot_del = ly_bb.get("delivered", 0) + ly_sms.get("delivered", 0) + ly_trig.get("delivered", 0)
    ty_tot_clicks = ty_bb.get("unique_clicks", 0) + ty_sms.get("total_clicks", 0) + ty_trig.get("unique_clicks", 0)
    ly_tot_clicks = ly_bb.get("unique_clicks", 0) + ly_sms.get("total_clicks", 0) + ly_trig.get("unique_clicks", 0)
    ty_tot_rev = ty_gt.get("revenue", 0)
    ly_tot_rev = ly_gt.get("revenue", 0)
    ty_tot_dpm = safe_div(ty_tot_rev, ty_tot_del) * 1000 if ty_tot_del else None
    ly_tot_dpm = safe_div(ly_tot_rev, ly_tot_del) * 1000 if ly_tot_del else None
    ty_trig_pct = safe_div(ty_trig.get("revenue", 0), ty_tot_rev)
    ly_trig_pct = safe_div(ly_trig.get("revenue", 0), ly_tot_rev)

    headers = ["Channel", "Metric", f"LY ({ly_data['year']})", f"TY ({ty_data['year']})", "Variance"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    wbr_rows = [
        ("Email",     "Delivered",     ly_bb.get("delivered"),       ty_bb.get("delivered"),       "#,##0"),
        ("",          "Unique Clicks", ly_bb.get("unique_clicks"),   ty_bb.get("unique_clicks"),   "#,##0"),
        ("",          "Revenue",       ly_email_rev,                 ty_email_rev,                 "$#,##0.00"),
        ("",          "$/M",           ly_email_dpm,                 ty_email_dpm,                 "$#,##0.00"),
        ("SMS",       "Delivered",     ly_sms.get("delivered"),      ty_sms.get("delivered"),      "#,##0"),
        ("",          "Clicks",        ly_sms.get("total_clicks"),   ty_sms.get("total_clicks"),   "#,##0"),
        ("",          "Revenue",       ly_sms.get("revenue"),        ty_sms.get("revenue"),        "$#,##0.00"),
        ("",          "$/M",           ly_sms_dpm,                   ty_sms_dpm,                   "$#,##0.00"),
        ("Triggered", "Delivered",     ly_trig.get("delivered"),     ty_trig.get("delivered"),     "#,##0"),
        ("",          "Unique Clicks", ly_trig.get("unique_clicks"), ty_trig.get("unique_clicks"), "#,##0"),
        ("",          "Revenue",       ly_trig.get("revenue"),       ty_trig.get("revenue"),       "$#,##0.00"),
        ("",          "$/M",           ly_trig_dpm,                  ty_trig_dpm,                  "$#,##0.00"),
        ("Total",     "Delivered",     ly_tot_del,                   ty_tot_del,                   "#,##0"),
        ("",          "Clicks",        ly_tot_clicks,                ty_tot_clicks,                "#,##0"),
        ("",          "Revenue",       ly_tot_rev,                   ty_tot_rev,                   "$#,##0.00"),
        ("",          "$/M",           ly_tot_dpm,                   ty_tot_dpm,                   "$#,##0.00"),
        ("",          "Trig % Rev",    ly_trig_pct,                  ty_trig_pct,                  "0.00%"),
    ]

    for row_idx, (channel, metric, ly_val, ty_val, fmt) in enumerate(wbr_rows, start=2):
        ws.cell(row=row_idx, column=1, value=channel)
        ws.cell(row=row_idx, column=2, value=metric)
        ly_cell = ws.cell(row=row_idx, column=3, value=ly_val)
        ty_cell = ws.cell(row=row_idx, column=4, value=ty_val)
        if fmt:
            ly_cell.number_format = fmt
            ty_cell.number_format = fmt
        var_cell = ws.cell(row=row_idx, column=5, value=var(ty_val, ly_val))
        var_cell.number_format = "0.00%"
        if channel:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)


# ---------------------------------------------------------------------------
# CSV comparison diagnostic
# ---------------------------------------------------------------------------

def _parse_int(val: str) -> int:
    """Parse a comma-formatted integer string (e.g. '1,234') to int."""
    return int((val or "0").replace(",", "").strip() or "0")


def compare_with_csv(
    month_num: int,
    year: int,
    batch_csv_path: str,
    canvas_csv_path: Optional[str] = None,
    brand: str = "BUR",
) -> None:
    """Load Braze CSV exports and compare delivered/opens/clicks against raw events.

    Prints a side-by-side table per campaign/canvas step showing CSV value,
    raw-events value, and the diff. Highlights large discrepancies and lists
    campaigns present in one source but not the other.
    """
    import csv as csv_mod

    start_ts, end_ts, _, _ = month_date_range(year, month_num)

    # ---- Load batch CSV ----
    csv_email: Dict[str, Dict] = {}
    csv_sms: Dict[str, Dict] = {}
    with open(batch_csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv_mod.DictReader(f)
        for row in reader:
            name = (row.get("Campaign Name") or "").strip()
            if not name:
                continue
            email_del = _parse_int(row.get("Deliveries (Email)", "0"))
            sms_del = _parse_int(row.get("Confirmed Deliveries (SMS)", "0"))
            entry = {
                "delivered":     email_del,
                "total_opens":   _parse_int(row.get("Total Opens (Email)", "0")),
                "unique_opens":  _parse_int(row.get("Unique Opens (Email)", "0")),
                "total_clicks":  _parse_int(row.get("Total Clicks (Email)", "0")),
                "unique_clicks": _parse_int(row.get("Unique Clicks (Email)", "0")),
                "sms_delivered": sms_del,
                "sms_clicks":    _parse_int(row.get("Total Clicks (SMS)", "0")),
            }
            if email_del > 0:
                csv_email[name] = entry
            if sms_del > 0:
                csv_sms[name] = entry

    # ---- Fetch from raw events ----
    app_group_id = BRAND_CONFIG[brand]["app_group_id"]
    client = _get_braze_client()
    try:
        print(f"  Fetching Braze batch email ({month_num}/{year})...", flush=True)
        raw_email = query_braze_email_batch(client, start_ts, end_ts, app_group_id)
        print(f"  Fetching Braze batch SMS ({month_num}/{year})...", flush=True)
        raw_sms = query_braze_sms_batch(client, start_ts, end_ts, app_group_id)
    finally:
        client.close()

    raw_email_dict = {r["campaign_name"]: r for r in raw_email}
    raw_sms_dict = {r["campaign_name"]: r for r in raw_sms}

    W = 55  # campaign name column width

    # ---- Email comparison ----
    print(f"\n{'='*120}")
    print(f"EMAIL CAMPAIGNS — CSV vs Raw Events  ({month_num}/{year})")
    print(f"{'='*120}")
    hdr = f"{'Campaign':<{W}} {'CSV Del':>10} {'Raw Del':>10} {'Diff':>8}  {'CSV UOpens':>10} {'Raw UOpens':>10}  {'CSV UClks':>9} {'Raw UClks':>9}"
    print(hdr)
    print("-" * len(hdr))

    csv_total, raw_total = 0, 0
    all_email_names = sorted(set(list(csv_email.keys()) + list(raw_email_dict.keys())))
    for name in all_email_names:
        cd = csv_email.get(name, {})
        rd = raw_email_dict.get(name, {})
        csv_del = cd.get("delivered", 0)
        raw_del = rd.get("delivered", 0)
        diff = raw_del - csv_del
        csv_total += csv_del
        raw_total += raw_del
        flag = "  <<<" if abs(diff) > 1000 else ("  (csv only)" if csv_del and not raw_del else ("  (raw only)" if raw_del and not csv_del else ""))
        print(f"  {name:<{W-2}} {csv_del:>10,} {raw_del:>10,} {diff:>+8,}  {cd.get('unique_opens',0):>10,} {rd.get('unique_opens',0):>10,}  {cd.get('unique_clicks',0):>9,} {rd.get('unique_clicks',0):>9,}{flag}")
    print("-" * len(hdr))
    diff_total = raw_total - csv_total
    print(f"  {'TOTAL':<{W-2}} {csv_total:>10,} {raw_total:>10,} {diff_total:>+8,}")

    # ---- SMS comparison ----
    print(f"\n{'='*90}")
    print(f"SMS CAMPAIGNS — CSV vs Raw Events  ({month_num}/{year})")
    print(f"{'='*90}")
    hdr2 = f"{'Campaign':<{W}} {'CSV Del':>10} {'Raw Del':>10} {'Diff':>8}  {'CSV Clks':>8} {'Raw Clks':>8}"
    print(hdr2)
    print("-" * len(hdr2))
    csv_sms_total, raw_sms_total = 0, 0
    all_sms_names = sorted(set(list(csv_sms.keys()) + list(raw_sms_dict.keys())))
    for name in all_sms_names:
        cd = csv_sms.get(name, {})
        rd = raw_sms_dict.get(name, {})
        csv_del = cd.get("sms_delivered", 0)
        raw_del = rd.get("delivered", 0)
        diff = raw_del - csv_del
        csv_sms_total += csv_del
        raw_sms_total += raw_del
        flag = "  <<<" if abs(diff) > 200 else ("  (csv only)" if csv_del and not raw_del else ("  (raw only)" if raw_del and not csv_del else ""))
        print(f"  {name:<{W-2}} {csv_del:>10,} {raw_del:>10,} {diff:>+8,}  {cd.get('sms_clicks',0):>8,} {rd.get('total_clicks',0):>8,}{flag}")
    print("-" * len(hdr2))
    print(f"  {'TOTAL':<{W-2}} {csv_sms_total:>10,} {raw_sms_total:>10,} {raw_sms_total-csv_sms_total:>+8,}")

    # ---- Canvas comparison ----
    if canvas_csv_path:
        csv_steps: Dict[str, Dict] = {}
        with open(canvas_csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv_mod.DictReader(f)
            for row in reader:
                step_name = (row.get("Step Name") or "").strip()
                if not step_name:
                    continue
                e_del = _parse_int(row.get("Deliveries (Email)", "0"))
                s_del = _parse_int(row.get("Confirmed Deliveries (SMS)", "0"))
                if step_name in csv_steps:
                    # Same step name appears in multiple canvas variants — accumulate
                    csv_steps[step_name]["delivered"] += e_del
                    csv_steps[step_name]["sms_delivered"] += s_del
                    csv_steps[step_name]["sms_clicks"] += _parse_int(row.get("Total Clicks (SMS)", "0"))
                    csv_steps[step_name]["unique_opens"] += _parse_int(row.get("Unique Opens (Email)", "0"))
                    csv_steps[step_name]["unique_clicks"] += _parse_int(row.get("Unique Clicks (Email)", "0"))
                else:
                    csv_steps[step_name] = {
                        "canvas_name":   (row.get("Canvas Name") or "").strip(),
                        "delivered":     e_del,
                        "unique_opens":  _parse_int(row.get("Unique Opens (Email)", "0")),
                        "unique_clicks": _parse_int(row.get("Unique Clicks (Email)", "0")),
                        "sms_delivered": s_del,
                        "sms_clicks":    _parse_int(row.get("Total Clicks (SMS)", "0")),
                    }

        client2 = _get_braze_client()
        try:
            print(f"\n  Fetching Braze canvas email ({month_num}/{year})...", flush=True)
            raw_ec = query_braze_email_canvas(client2, start_ts, end_ts, app_group_id)
            print(f"  Fetching Braze canvas SMS ({month_num}/{year})...", flush=True)
            raw_sc = query_braze_sms_canvas(client2, start_ts, end_ts, app_group_id)
        finally:
            client2.close()

        raw_ce = {r["step_name"]: r for r in raw_ec}
        raw_cs = {r["step_name"]: r for r in raw_sc}

        print(f"\n{'='*120}")
        print(f"CANVAS STEPS — CSV vs Raw Events  ({month_num}/{year})")
        print(f"{'='*120}")
        hdr3 = (f"{'Step Name':<{W}} {'CSV Del':>10} {'Raw Del':>10} {'Diff':>8}"
                f"  {'CSV SMS Del':>11} {'Raw SMS Del':>11}  {'CSV SMS Clk':>11} {'Raw SMS Clk':>11}")
        print(hdr3)
        print("-" * len(hdr3))

        csv_c_del, raw_c_del = 0, 0
        csv_c_sdel, raw_c_sdel = 0, 0
        all_step_names = sorted(set(list(csv_steps.keys()) + list(raw_ce.keys()) + list(raw_cs.keys())))
        for sname in all_step_names:
            cd = csv_steps.get(sname, {})
            re = raw_ce.get(sname, {})
            rs = raw_cs.get(sname, {})
            csv_del = cd.get("delivered", 0)
            raw_del = re.get("delivered", 0)
            csv_sdel = cd.get("sms_delivered", 0)
            raw_sdel = rs.get("delivered", 0)
            diff = raw_del - csv_del
            csv_c_del += csv_del
            raw_c_del += raw_del
            csv_c_sdel += csv_sdel
            raw_c_sdel += raw_sdel
            note = ""
            if not cd:
                note = "  (raw only)"
            elif not re and not rs:
                note = "  (csv only)"
            elif abs(diff) > 50 or abs(raw_sdel - csv_sdel) > 20:
                note = "  <<<"
            print(f"  {sname:<{W-2}} {csv_del:>10,} {raw_del:>10,} {diff:>+8,}"
                  f"  {csv_sdel:>11,} {raw_sdel:>11,}  {cd.get('sms_clicks',0):>11,} {rs.get('total_clicks',0):>11,}{note}")

        print("-" * len(hdr3))
        print(f"  {'TOTAL':<{W-2}} {csv_c_del:>10,} {raw_c_del:>10,} {raw_c_del-csv_c_del:>+8,}"
              f"  {csv_c_sdel:>11,} {raw_c_sdel:>11,}")

    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate lifecycle performance report (monthly or weekly) for BUR or CZ"
    )
    parser.add_argument(
        "--brand", default="BUR", choices=list(BRAND_CONFIG.keys()),
        help="Brand code (default: BUR)",
    )
    parser.add_argument("--month", default=None, help="Month name for monthly report (e.g. February)")
    parser.add_argument("--year", default=None, type=int, help="Year for monthly report (e.g. 2026)")
    parser.add_argument(
        "--week-ending", default=None, metavar="YYYY-MM-DD",
        help="Generate weekly report for Mon–Sun ending on this Sunday",
    )
    parser.add_argument("--output", default=None, help="Output Excel path")
    parser.add_argument(
        "--compare-csv",
        default=None,
        metavar="PATH",
        help=(
            "Path to Braze 'Batch and Blast' CSV export. "
            "When provided, prints a side-by-side comparison of delivered/opens/clicks "
            "vs raw events and exits without building the Excel file."
        ),
    )
    parser.add_argument(
        "--compare-canvas-csv",
        default=None,
        metavar="PATH",
        help="Path to Braze canvas/triggers CSV export (used together with --compare-csv).",
    )
    args = parser.parse_args()

    brand = args.brand

    # ---- Validate mode: exactly one of --week-ending or --month/--year ----
    if args.week_ending and (args.month or args.year):
        parser.error("--week-ending and --month/--year are mutually exclusive")
    if not args.week_ending and not args.month:
        parser.error("one of --month/--year or --week-ending is required")
    if args.month and not args.year:
        parser.error("--year is required with --month")

    file_prefix = BRAND_CONFIG[brand]["file_prefix"]
    reports_dir = Path(__file__).parent.parent / "reports"

    # ---- Diagnostic comparison mode (monthly only) ----
    if args.compare_csv:
        if not args.month:
            parser.error("--compare-csv requires --month and --year (monthly mode only)")
        month_name = args.month.capitalize()
        year = args.year
        month_num = month_name_to_num(month_name)
        print(f"{brand} Lifecycle Report — CSV comparison mode  |  {month_name} {year}\n")
        compare_with_csv(
            month_num,
            year,
            batch_csv_path=args.compare_csv,
            canvas_csv_path=args.compare_canvas_csv,
            brand=brand,
        )
        return

    # ---- Build TY and LY PeriodSpec objects ----
    if args.week_ending:
        # --- Weekly mode ---
        week_end = date.fromisoformat(args.week_ending)
        if week_end.weekday() != 6:  # 0=Mon … 6=Sun
            parser.error(
                f"--week-ending must be a Sunday; "
                f"{args.week_ending} is a {week_end.strftime('%A')}"
            )
        week_start = week_end - timedelta(days=6)

        start_ts, end_ts, start_yyyymmdd, end_yyyymmdd = week_date_range(week_end)
        ty_label = format_week_tab_label(week_end)
        ty_spec = PeriodSpec(
            brand=brand, mode="weekly",
            start_ts=start_ts, end_ts=end_ts,
            start_yyyymmdd=start_yyyymmdd, end_yyyymmdd=end_yyyymmdd,
            tab_label=ty_label, year=week_end.year,
        )

        ly_end = ly_week_ending(week_end)
        ly_start = ly_end - timedelta(days=6)
        ly_s_ts, ly_e_ts, ly_s_yyyymmdd, ly_e_yyyymmdd = week_date_range(ly_end)
        ly_label = format_week_tab_label(ly_end)
        ly_spec = PeriodSpec(
            brand=brand, mode="weekly",
            start_ts=ly_s_ts, end_ts=ly_e_ts,
            start_yyyymmdd=ly_s_yyyymmdd, end_yyyymmdd=ly_e_yyyymmdd,
            tab_label=ly_label, year=ly_end.year,
        )

        default_filename = (
            f"{file_prefix}_Lifecycle_Report_Week"
            f"_{week_start.strftime('%Y-%m-%d')}"
            f"_{week_end.strftime('%Y-%m-%d')}.xlsx"
        )
        print_header = (
            f"{brand} Lifecycle Report  |  Week {week_start} to {week_end}"
        )

    else:
        # --- Monthly mode ---
        month_name = args.month.capitalize()
        year = args.year
        month_num = month_name_to_num(month_name)

        start_ts, end_ts, start_yyyymmdd, end_yyyymmdd = month_date_range(year, month_num)
        ty_label = f"{month_name} {year}"
        ty_spec = PeriodSpec(
            brand=brand, mode="monthly",
            start_ts=start_ts, end_ts=end_ts,
            start_yyyymmdd=start_yyyymmdd, end_yyyymmdd=end_yyyymmdd,
            tab_label=ty_label, year=year,
        )

        ly_year = year - 1
        ly_s_ts, ly_e_ts, ly_s_yyyymmdd, ly_e_yyyymmdd = month_date_range(ly_year, month_num)
        ly_label = f"{month_name} {ly_year}"
        ly_spec = PeriodSpec(
            brand=brand, mode="monthly",
            start_ts=ly_s_ts, end_ts=ly_e_ts,
            start_yyyymmdd=ly_s_yyyymmdd, end_yyyymmdd=ly_e_yyyymmdd,
            tab_label=ly_label, year=ly_year,
        )

        default_filename = f"{file_prefix}_Lifecycle_Report_{month_name}_{year}.xlsx"
        print_header = f"{brand} Lifecycle Report  |  {month_name} {year}"

    if args.output:
        output_path = Path(args.output)
    else:
        reports_dir.mkdir(parents=True, exist_ok=True)
        output_path = reports_dir / default_filename

    print(f"{print_header}  |  -> {output_path}\n")

    print("Loading canvas lookups from YAML files...")
    step_name_lookup = load_canvas_step_name_lookup()
    canvas_api_id_lookup = load_canvas_api_id_lookup()
    print(f"  {len(step_name_lookup)} step name mappings  |  {len(canvas_api_id_lookup)} canvas IDs\n")

    print(f"=== TY: {ty_spec.tab_label} ===")
    ty_data = run_for_period(ty_spec, step_name_lookup, canvas_api_id_lookup)
    print()

    print(f"=== LY: {ly_spec.tab_label} ===")
    ly_data = run_for_period(ly_spec, step_name_lookup, canvas_api_id_lookup)
    print()

    print("Building Excel workbook...")
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: uv add openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    print(f"  TY tabs ({ty_spec.tab_label})...")
    write_summary_tab(wb, ty_data, ty_spec.tab_label)
    write_bb_tab(wb, ty_data, ty_spec.tab_label)
    write_sms_tab(wb, ty_data, ty_spec.tab_label)
    write_triggers_tab(wb, ty_data, ty_spec.tab_label)

    print(f"  LY tabs ({ly_spec.tab_label})...")
    write_summary_tab(wb, ly_data, ly_spec.tab_label)
    write_bb_tab(wb, ly_data, ly_spec.tab_label)
    write_sms_tab(wb, ly_data, ly_spec.tab_label)
    write_triggers_tab(wb, ly_data, ly_spec.tab_label)

    print("  WBR tab...")
    write_wbr_tab(wb, ty_data, ly_data, ty_spec.tab_label)

    wb.save(output_path)
    print(f"\nSaved: {output_path}\n")

    # Validation printout
    ty_s = ty_data["summary"]
    bb = ty_s["bb"]
    sms = ty_s["sms"]
    trig = ty_s["triggers"]
    print(f"=== TY Validation ({ty_spec.tab_label}) ===")
    print(f"  B&B      : Delivered={bb.get('delivered', 0):>12,}  Orders={bb.get('orders', 0):>5,}  Revenue=${bb.get('revenue', 0):>12,.2f}")
    print(f"  SMS      : Delivered={sms.get('delivered', 0):>12,}  Orders={sms.get('orders', 0):>5,}  Revenue=${sms.get('revenue', 0):>12,.2f}")
    print(f"  Triggers : Delivered={trig.get('delivered', 0):>12,}  SMS Del={trig.get('sms_delivered', 0):>6,}  Orders={trig.get('orders', 0):>5,}  Revenue=${trig.get('revenue', 0):>12,.2f}")
    if brand == "BUR" and ty_spec.mode == "monthly" and ty_spec.tab_label == "February 2026":
        print()
        print("  Reference targets (Feb 2026):")
        print("    B&B      : Delivered=   4,987,921  Orders=   47  Revenue=$    65,595.92")
        print("    SMS      : Delivered=      52,635  Orders=    2  Revenue=$     1,064.97")
        print("    Triggers : Delivered=      30,454  SMS Del= 4,732  Orders=   21  Revenue=$    35,574.61")


if __name__ == "__main__":
    main()

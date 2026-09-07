#!/usr/bin/env python3
"""Build email-link-guide.xlsx from the guide data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parents[2] / "reports" / "email-link-guide.xlsx"

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY      = "1B2E4A"
WHITE     = "FFFFFF"
LIGHTBG   = "F8F9FA"
HDRBG     = "E5E7EB"
ALTROW    = "F3F4F6"
NOTE_BG   = "FEF9C3"   # soft yellow for warning rows
TEXTDRK   = "111827"
TEXTMID   = "374151"
GRAY      = "6B7280"

BRAND_COLORS = {
    "BUR": "DBEAFE",   # light blue
    "CZ":  "D1FAE5",   # light green
    "HAV": "FEE2E2",   # light red/rose
    "ID":  "EDE9FE",   # light purple
    "STF": "FEF3C7",   # light amber
    "TI":  "CCFBF1",   # light teal
    "ALL": "F3F4F6",   # gray
}

# ── Data ──────────────────────────────────────────────────────────────────────
# Columns: brand, section, page_name, url, when_to_use, notes

ROWS = [
    # ── Universal ─────────────────────────────────────────────────────────────
    ("ALL", "Footer", "Burrow",          "https://burrow.com",            "Standard footer element in every email",    ""),
    ("ALL", "Footer", "Interior Define", "https://interiordefine.com",    "Standard footer element in every email",    ""),
    ("ALL", "Footer", "The Citizenry",   "https://the-citizenry.com",     "Standard footer element in every email",    ""),
    ("ALL", "Footer", "St. Frank",       "https://stfrank.com",           "Standard footer element in every email",    ""),
    ("ALL", "Footer", "The Inside",      "https://theinside.com",         "Standard footer element in every email",    ""),
    ("ALL", "Footer", "Havenly",         "https://havenly.com",           "Standard footer element in every email",    ""),

    # ── BUR — Category / Department ───────────────────────────────────────────
    ("BUR", "Category",   "Homepage",       "https://burrow.com",                             "Default CTA when no specific category applies; general brand awareness", ""),
    ("BUR", "Category",   "Seating",        "https://burrow.com/seating",                     "Seating highlight emails, general promotional", ""),
    ("BUR", "Category",   "Dining",         "https://burrow.com/dining",                      "Dining-focused highlights", ""),
    ("BUR", "Category",   "Storage",        "https://burrow.com/storage",                     "Storage/shelving highlights", ""),
    ("BUR", "Category",   "Bedroom",        "https://burrow.com/bedroom",                     "Bedroom-focused emails", ""),
    ("BUR", "Category",   "Rugs & Decor",   "https://burrow.com/rugs-decor",                  "Rug/decor spotlights", ""),
    ("BUR", "Category",   "Outdoor",        "https://burrow.com/outdoor",                     "Outdoor/seasonal emails", ""),
    ("BUR", "Category",   "Living",         "https://burrow.com/living",                      "Living room roundups", ""),
    ("BUR", "Category",   "Sofas",          "https://burrow.com/sofas",                       "Sofa roundup / general", ""),
    ("BUR", "Category",   "Sectionals",     "https://burrow.com/sectionals",                  "Sectional category spotlight", ""),
    ("BUR", "Category",   "Leather Seating","https://burrow.com/leather-seating",             "Leather-specific callout", ""),

    # ── BUR — Collection ──────────────────────────────────────────────────────
    ("BUR", "Collection", "Clearance",             "https://burrow.com/collections/clearance",           "Sale emails, clearance CTAs", ""),
    ("BUR", "Collection", "Best Sellers",           "https://burrow.com/collections/best-sellers",        "'Best of' roundups, re-engagement", ""),
    ("BUR", "Collection", "Ready to Ship",          "https://burrow.com/ready-to-ship",                   "Urgency/availability messaging", ""),
    ("BUR", "Collection", "Sleeper Sofas",          "https://burrow.com/collections/sleeper-sofas",       "Sleeper-specific spotlight", ""),
    ("BUR", "Collection", "Modular Furniture",      "https://burrow.com/collections/modular-furniture",   "Modular system feature emails", ""),
    ("BUR", "Collection", "Pro Collection",         "https://burrow.com/collections/pro-collection",      "Pro/Plus tier messaging", ""),
    ("BUR", "Collection", "Pro + Plus Series",      "https://burrow.com/pro-and-plus-series",             "Upgrade/comparison content", ""),
    ("BUR", "Collection", "Range Collection",       "https://burrow.com/collections/range",               "Range feature emails", ""),
    ("BUR", "Collection", "Nomad Collection",       "https://burrow.com/collections/nomad",               "Nomad feature emails", ""),
    ("BUR", "Collection", "Union Collection",       "https://burrow.com/collections/union",               "Union feature emails", ""),
    ("BUR", "Collection", "Span Collection",        "https://burrow.com/collections/span",                "Span/sleeper storage feature", ""),
    ("BUR", "Collection", "Accent Chairs",          "https://burrow.com/collections/accent-chairs",       "Accent chair highlight", ""),
    ("BUR", "Collection", "Dining Tables",          "https://burrow.com/collections/dining-tables",       "Dining table feature", ""),
    ("BUR", "Collection", "Dining Chairs",          "https://burrow.com/collections/dining-chairs",       "Dining chair feature", ""),
    ("BUR", "Collection", "Back to School",         "https://burrow.com/collections/back-to-school",      "Seasonal dorm/student campaign", ""),

    # ── BUR — Products ────────────────────────────────────────────────────────
    ("BUR", "Product",    "Opera Media Console",           "https://burrow.com/products/opera-media-console",           "Frequently featured product", ""),
    ("BUR", "Product",    "Opera Tall Media Console",      "https://burrow.com/products/opera-tall-media-console",      "Frequently featured product", ""),
    ("BUR", "Product",    "Span Sleeper Sofa",             "https://burrow.com/products/span-sleeper-sofa",             "Frequently featured product", ""),
    ("BUR", "Product",    "Shift Sleeper Sofa",            "https://burrow.com/products/shift-sleeper-sofa",            "Frequently featured product", ""),
    ("BUR", "Product",    "Listo Dining Table",            "https://burrow.com/products/listo-dining-table",            "Frequently featured product", ""),
    ("BUR", "Product",    "Gimlet Chair",                  "https://burrow.com/products/gimlet-chair",                  "Frequently featured product", ""),
    ("BUR", "Product",    "Airmail Chair",                 "https://burrow.com/products/airmail-chair",                 "Frequently featured product", ""),
    ("BUR", "Product",    "Vesper Leather Lounge Chair",   "https://burrow.com/products/vesper-leather-lounge-chair",   "Frequently featured product", ""),
    ("BUR", "Product",    "Rye Recliner",                  "https://burrow.com/products/rye-recliner",                  "Frequently featured product", ""),
    ("BUR", "Product",    "Range Plus 3-Piece Sofa",       "https://burrow.com/products/range-plus-3-piece-sofa",       "Frequently featured product", ""),
    ("BUR", "Product",    "Union Pro 108 Chaise Sectional","https://burrow.com/products/union-pro-108-chaise-sectional","Frequently featured product", ""),

    # ── BUR — Utility ─────────────────────────────────────────────────────────
    ("BUR", "Utility",    "Swatches",            "https://burrow.com/swatches",                       "Free swatch offers, fabric-first messaging", ""),
    ("BUR", "Utility",    "Showrooms",           "https://burrow.com/showrooms",                      "In-person/local event tie-ins", ""),
    ("BUR", "Utility",    "Pet-Friendly",        "https://burrow.com/pet-friendly-furniture",         "Pet owner segment emails", ""),
    ("BUR", "Utility",    "Gift Card",           "https://burrow.com/products/burrow-gift-card",      "Gift guide or holiday emails", ""),
    ("BUR", "Utility",    "Refer a Friend",      "https://burrow.com/refer",                          "Referral/loyalty campaigns", ""),
    ("BUR", "Utility",    "Fall Preview",        "https://burrow.com/fall-preview",                   "Seasonal preview launches", ""),
    ("BUR", "Utility",    "Cart",                "https://burrow.com/cart",                           "Abandoned cart flows only", ""),

    # ── BUR — Third-Party ─────────────────────────────────────────────────────
    ("BUR", "Third-Party","Showroom Booking (Acuity)","https://burrowhouseappointments.as.me/schedule/...","In-person design consult campaigns", ""),

    # ── BUR — PDP Variant Rule ────────────────────────────────────────────────
    ("BUR", "⚠ PDP Variant Rule", "Variant URL Format",
     "https://burrow.com/products/[product-slug]?[Option%20Name]=[Option%20Value]",
     "Always use %20 for spaces in query params — Braze converts + to %2B on save, breaking variant matching",
     "Example: ?Wood%20Finish=Walnut%20-%20Wood&Fabric=Nomad%20Linen%20Natural"),

    # ── CZ — Top-Level Shop ───────────────────────────────────────────────────
    ("CZ", "Category",    "Homepage",                  "https://the-citizenry.com",                                         "Default CTA, general brand", ""),
    ("CZ", "Category",    "Shop All Bedding",           "https://the-citizenry.com/collections/shop-all-bedding-2",          "Bedding-focused campaigns", ""),
    ("CZ", "Category",    "Shop All Rugs",              "https://the-citizenry.com/collections/shop-all-rugs-1",             "Rug feature / rug sale", ""),
    ("CZ", "Category",    "Shop All Furniture",         "https://the-citizenry.com/collections/shop-all-furniture",          "Furniture spotlight", ""),
    ("CZ", "Category",    "Shop All Pillows",           "https://the-citizenry.com/collections/shop-all-pillows",            "Pillow highlight", ""),
    ("CZ", "Category",    "All Accents",                "https://the-citizenry.com/collections/all-accents",                 "Décor-focused sends", ""),
    ("CZ", "Category",    "Shop All Bath",              "https://the-citizenry.com/collections/shop-all-bath",               "Bath product spotlights", ""),
    ("CZ", "Category",    "All Baskets",                "https://the-citizenry.com/collections/all-baskets",                 "Basket/storage feature", ""),
    ("CZ", "Category",    "All Bed Bundles",            "https://the-citizenry.com/collections/all-bed-bundles",             "Bundle / value messaging", ""),
    ("CZ", "Category",    "All Best Sellers",           "https://the-citizenry.com/collections/all-best-sellers",            "Re-engagement, best-of roundups", ""),
    ("CZ", "Category",    "All New Arrivals",           "https://the-citizenry.com/collections/all-new-arrivals",            "New arrival announcement", ""),
    ("CZ", "Category",    "All Back in Stock",          "https://the-citizenry.com/collections/all-back-in-stock",           "BIS/back-in-stock alerts", ""),
    ("CZ", "Category",    "Ready to Ship",              "https://the-citizenry.com/collections/ready-to-ship",               "Availability / urgency", ""),
    ("CZ", "Category",    "Ready to Ship — Furniture",  "https://the-citizenry.com/collections/ready-to-ship-furniture",     "Furniture-specific urgency", ""),

    # ── CZ — Sale ─────────────────────────────────────────────────────────────
    ("CZ", "Sale / Event","Archive Sale",              "https://the-citizenry.com/collections/archive-sale",                "Clearance/archive sale campaigns (highest use)", ""),
    ("CZ", "Sale / Event","Fresh Foundations Sale",    "https://the-citizenry.com/collections/the-fresh-foundations-sale",  "This specific sale", ""),
    ("CZ", "Sale / Event","Spring Event",              "https://the-citizenry.com/collections/the-spring-event",            "Spring sale/event", ""),
    ("CZ", "Sale / Event","Bedroom Event",             "https://the-citizenry.com/collections/the-bedroom-event",           "Bedroom-focused event", ""),
    ("CZ", "Sale / Event","Summer Retreat Sale",       "https://the-citizenry.com/collections/the-summer-retreat-sale",     "Summer sale", ""),
    ("CZ", "Sale / Event","Sunset Sale",               "https://the-citizenry.com/collections/the-sunset-sale",             "Late-summer sale", ""),
    ("CZ", "Sale / Event","Weekender Sale",            "https://the-citizenry.com/collections/the-weekender-sale",          "Weekender event", ""),
    ("CZ", "Sale / Event","Fall Refresh Event",        "https://the-citizenry.com/collections/the-fall-refresh-event-2025", "Fall event", ""),
    ("CZ", "Sale / Event","Holiday Shop",              "https://the-citizenry.com/pages/2025-holiday-shop",                 "Holiday gift guide", ""),

    # ── CZ — Editorial ────────────────────────────────────────────────────────
    ("CZ", "Editorial",   "The Layered Bed",           "https://the-citizenry.com/pages/the-layered-bed",                          "Bedding editorial / how-to style", ""),
    ("CZ", "Editorial",   "American Craft Collection", "https://the-citizenry.com/pages/the-american-craft-collection",            "Craft/artisan editorial", ""),
    ("CZ", "Editorial",   "Fall Collection 2025",      "https://the-citizenry.com/pages/the-fall-collection-2025",                 "New collection reveal", ""),
    ("CZ", "Editorial",   "Pillow Pairings",           "https://the-citizenry.com/pages/pillow-pairings",                          "Pillow styling guide", ""),
    ("CZ", "Editorial",   "Artisan Index",             "https://the-citizenry.com/pages/artisan-index",                            "Brand story / artisan origin content", ""),
    ("CZ", "Editorial",   "Shop by Country",           "https://the-citizenry.com/pages/shop-by-country",                         "Origin/artisan-focused emails", ""),
    ("CZ", "Editorial",   "Sustainability",            "https://the-citizenry.com/collections/shop-sustainably",                   "Impact / values messaging", ""),
    ("CZ", "Editorial",   "About",                     "https://the-citizenry.com/pages/about",                                    "Brand story, new subscriber onboarding", ""),
    ("CZ", "Editorial",   "Morocco Collection",        "https://the-citizenry.com/collections/the-morocco-collection",             "Collection-specific campaign", ""),
    ("CZ", "Editorial",   "Mexico Collection",         "https://the-citizenry.com/collections/the-mexico-collection",              "Collection-specific campaign", ""),
    ("CZ", "Editorial",   "Portugal Linen Collection", "https://the-citizenry.com/pages/the-portugal-stonewashed-linen-bedding-collection", "Collection-specific", ""),
    ("CZ", "Editorial",   "Oaxaca Collection",         "https://the-citizenry.com/pages/the-oaxaca-collection",                    "Collection-specific", ""),
    ("CZ", "Editorial",   "Greenery Shop",             "https://the-citizenry.com/pages/the-greenery-shop",                        "Seasonal/plant-adjacent", ""),
    ("CZ", "Editorial",   "Blog — Design",             "https://the-citizenry.com/blogs/design/[post-slug]",                       "Trend/style content", "Pattern: /blogs/design/[slug]"),
    ("CZ", "Editorial",   "Blog — Travel",             "https://the-citizenry.com/blogs/travel/[post-slug]",                       "Travel/origin story content", "Pattern: /blogs/travel/[slug]"),

    # ── CZ — Utility ──────────────────────────────────────────────────────────
    ("CZ", "Utility",     "Trade Program",             "https://the-citizenry.com/pages/trade-program-1",                "Trade segment or cross-sell", ""),
    ("CZ", "Utility",     "Dallas Store",              "https://the-citizenry.com/pages/the-citizenry-dallas",           "Local event / store-specific", ""),
    ("CZ", "Utility",     "CZ Flagship",               "https://the-citizenry.com/pages/the-citizenry-flagship",         "Flagship store / in-person", ""),
    ("CZ", "Utility",     "Store Locator",             "https://the-citizenry.com/pages/store-locator",                  "Event-driven / proximity", ""),
    ("CZ", "Utility",     "Rug Size Guide",            "https://the-citizenry.com/pages/rug-size-and-style-guide",       "Educational / nurture", ""),
    ("CZ", "Utility",     "Cart",                      "https://the-citizenry.com/cart",                                 "Abandoned cart flows only", ""),
    ("CZ", "Utility",     "Gift Card",                 "https://the-citizenry.com/products/gift-card",                   "Gift-focused campaigns", ""),
    ("CZ", "Utility",     "Write a Review",            "https://the-citizenry.com/a/review/write",                       "Post-purchase review request", ""),
    ("CZ", "Utility",     "Catalog Opt-In",            "https://the-citizenry.com/pages/catalog-opt-in",                 "Physical catalog campaign", ""),

    # ── CZ — Third-Party ──────────────────────────────────────────────────────
    ("CZ", "Third-Party", "Flagship Styling Booking (Acuity)", "https://citizenryflagshipstyling.as.me/schedule/...", "Flagship styling appointment booking", ""),
    ("CZ", "Third-Party", "Typeform Sweepstakes",              "https://form.typeform.com/to/CKxdXtJj",              "Sweepstakes / feedback form", ""),
    ("CZ", "Third-Party", "Typeform Feedback (shared w/ ID)",  "https://form.typeform.com/to/avQR2W9q",              "Shared feedback survey", ""),

    # ── HAV — Core Service ────────────────────────────────────────────────────
    ("HAV", "Core Service", "Homepage",                "https://havenly.com",                                    "Default CTA", ""),
    ("HAV", "Core Service", "Shop",                    "https://havenly.com/shop",                               "Product shop landing", ""),
    ("HAV", "Core Service", "Pricing",                 "https://havenly.com/pricing",                            "Conversion / nurture for non-customers", ""),
    ("HAV", "Core Service", "Interior Design Services","https://havenly.com/interior-design-services",           "Service pitch", ""),
    ("HAV", "Core Service", "AI Interior Design",      "https://havenly.com/ai-interior-design",                 "AI feature campaigns", ""),
    ("HAV", "Core Service", "In-Person",               "https://havenly.com/in-person",                          "In-person design service promotion", ""),
    ("HAV", "Core Service", "Interior Designers",      "https://havenly.com/interior-designers",                 "Find-a-designer CTAs", ""),
    ("HAV", "Core Service", "Current Promotions",      "https://havenly.com/current-promotions",                 "Sale / promo landing", ""),
    ("HAV", "Core Service", "Shop My Room",            "https://havenly.com/shop-my-room",                       "Post-consult / design reveal follow-up", ""),

    # ── HAV — Category ────────────────────────────────────────────────────────
    ("HAV", "Category",   "Living Room Furniture",     "https://havenly.com/shop/category/living-room-furniture","Room-specific campaigns", ""),
    ("HAV", "Category",   "Bedroom Furniture",         "https://havenly.com/shop/category/bedroom-furniture",    "Bedroom-specific", ""),
    ("HAV", "Category",   "Dining Room Furniture",     "https://havenly.com/shop/category/dining-room-furniture","Dining-specific", ""),
    ("HAV", "Category",   "Decor & Pillows",           "https://havenly.com/shop/category/decor-pillows",        "Decor-focused", ""),
    ("HAV", "Category",   "Rugs",                      "https://havenly.com/shop/category/rugs",                 "Rug highlight", ""),
    ("HAV", "Category",   "Lighting",                  "https://havenly.com/shop/category/lighting",             "Lighting feature", ""),
    ("HAV", "Category",   "Outdoor Furniture",         "https://havenly.com/shop/category/outdoor-furniture",    "Outdoor/seasonal", ""),

    # ── HAV — Collection / Brand ──────────────────────────────────────────────
    ("HAV", "Collection", "The Citizenry",             "https://havenly.com/shop/collection/the-citizenry",            "CZ cross-sell or CZ-focused", ""),
    ("HAV", "Collection", "Burrow",                    "https://havenly.com/shop/collection/burrow",                   "BUR cross-sell", ""),
    ("HAV", "Collection", "Interior Define",           "https://havenly.com/shop/collection/interior-define",          "ID cross-sell", ""),
    ("HAV", "Collection", "St. Frank",                 "https://havenly.com/shop/collection/st-frank",                 "STF cross-sell", ""),
    ("HAV", "Collection", "The Inside by Havenly",     "https://havenly.com/shop/collection/the-inside-by-havenly",    "Private label", ""),
    ("HAV", "Collection", "Sale",                      "https://havenly.com/shop/collection/sale",                     "Sale discovery", ""),
    ("HAV", "Collection", "Sofas & Sectionals",        "https://havenly.com/shop/collection/sofas-sectionals",         "Sofa feature", ""),
    ("HAV", "Collection", "Dining Tables",             "https://havenly.com/shop/collection/dining-tables",            "Dining feature", ""),
    ("HAV", "Collection", "Bedroom Favorites",         "https://havenly.com/shop/collection/bedroom-favorites",        "Bedroom roundup", ""),
    ("HAV", "Collection", "Decor Under $100",          "https://havenly.com/shop/collection/decor-under-100",          "Budget/gifting", ""),

    # ── HAV — Blog ────────────────────────────────────────────────────────────
    ("HAV", "Blog",       "Room ideas",                "https://havenly.com/blog/[post-slug]",                   "Room-specific segments — heavy use for pre-converted nurture", "Pattern: /blog/[slug]"),
    ("HAV", "Blog",       "Trend content",             "https://havenly.com/blog/[post-slug]",                   "Trend newsletters", "Pattern: /blog/[slug]"),
    ("HAV", "Blog",       "Product guides",            "https://havenly.com/blog/[post-slug]",                   "Pre-purchase nurture", "Pattern: /blog/[slug]"),
    ("HAV", "Blog",       "Color content",             "https://havenly.com/blog/[post-slug]",                   "Decorating inspiration", "Pattern: /blog/[slug]"),
    ("HAV", "Blog",       "Before/After",              "https://havenly.com/blog/[post-slug]",                   "Transformation stories", "Pattern: /blog/[slug]"),
    ("HAV", "Blog",       "Design Mistakes",           "https://havenly.com/blog/[post-slug]",                   "Problem-aware messaging", "Pattern: /blog/[slug]"),

    # ── HAV — Inspiration ─────────────────────────────────────────────────────
    ("HAV", "Inspiration","Rooms",                     "https://havenly.com/rooms",                              "Design portfolio / inspiration", ""),
    ("HAV", "Inspiration","Interior Design Ideas",     "https://havenly.com/exp/interior-design-ideas",          "SEO landing / browse", ""),
    ("HAV", "Inspiration","Design Board (specific)",   "https://havenly.com/interior-design-ideas/design-board/[id]", "Show actual design work in emails", ""),
    ("HAV", "Inspiration","Style Quiz",                "https://havenly.com/interior-design-style-quiz",         "Top-of-funnel / re-engagement", ""),

    # ── HAV — Utility ─────────────────────────────────────────────────────────
    ("HAV", "Utility",    "Reviews",                   "https://havenly.com/reviews",                            "Social proof emails", ""),
    ("HAV", "Utility",    "Gift",                      "https://havenly.com/gift",                               "Gift card / gifting campaigns", ""),
    ("HAV", "Utility",    "Cart",                      "https://havenly.com/cart",                               "Abandoned cart flows", ""),

    # ── HAV — Third-Party ─────────────────────────────────────────────────────
    ("HAV", "Third-Party","Branch.io App Link",        "https://havenly.app.link/...",                           "App download / deep link CTAs", ""),
    ("HAV", "Third-Party","Apple App Store",           "https://apps.apple.com/us/app/havenly-interior-design/id1149153371", "App download (iOS)", ""),
    ("HAV", "Third-Party","Google Forms Survey",       "https://docs.google.com/forms/...",                      "Survey / feedback", ""),

    # ── ID — Category ─────────────────────────────────────────────────────────
    ("ID", "Category",    "Homepage",                  "https://interiordefine.com",                                     "Default CTA", ""),
    ("ID", "Category",    "All Custom Sectionals",     "https://interiordefine.com/living/all-custom-sectionals",        "Sectional spotlight", ""),
    ("ID", "Category",    "All Custom Sofas",          "https://interiordefine.com/living/all-custom-sofas",             "Sofa feature", ""),
    ("ID", "Category",    "All Custom Chairs",         "https://interiordefine.com/living/all-custom-chairs",            "Accent chair feature", ""),
    ("ID", "Category",    "Custom Accent Chairs",      "https://interiordefine.com/living/all-custom-chairs/custom-accent-chairs", "Accent chair specific", ""),
    ("ID", "Category",    "Dining",                    "https://interiordefine.com/dining",                              "Dining category", ""),
    ("ID", "Category",    "Bedroom",                   "https://interiordefine.com/bedroom",                             "Bedroom category", ""),
    ("ID", "Category",    "All Beds",                  "https://interiordefine.com/bedroom/all-beds",                    "Bed-specific", ""),
    ("ID", "Category",    "Rugs",                      "https://interiordefine.com/rugs",                                "Rug feature", ""),
    ("ID", "Category",    "Decor",                     "https://interiordefine.com/decor",                               "Décor / accessories", ""),
    ("ID", "Category",    "Lighting",                  "https://interiordefine.com/lighting",                            "Lighting feature", ""),
    ("ID", "Category",    "Outdoor",                   "https://interiordefine.com/outdoor",                             "Outdoor seasonal", ""),
    ("ID", "Category",    "New Arrivals",              "https://interiordefine.com/new-arrivals",                        "Launch / new product", ""),

    # ── ID — Availability ─────────────────────────────────────────────────────
    ("ID", "Availability","In Stock",                  "https://interiordefine.com/in-stock",                    "Inventory-based urgency messaging", ""),
    ("ID", "Availability","Quick Ship",                "https://interiordefine.com/quick-ship",                  "Lead time / shipping-focused", ""),
    ("ID", "Availability","Quick Ship Collections",    "https://interiordefine.com/quick-ship-collections",      "Urgency-driven promotions", ""),

    # ── ID — Collection ───────────────────────────────────────────────────────
    ("ID", "Collection",  "Sloan Collection",          "https://interiordefine.com/sloan-collection",            "Sloan feature email", ""),
    ("ID", "Collection",  "James Collection",          "https://interiordefine.com/james-collection",            "James feature email", ""),
    ("ID", "Collection",  "Maxwell Collection",        "https://interiordefine.com/maxwell-collection",          "Maxwell feature email", ""),
    ("ID", "Collection",  "Tatum Collection",          "https://interiordefine.com/tatum-collection",            "Tatum feature email", ""),
    ("ID", "Collection",  "Jasper Collection",         "https://interiordefine.com/jasper-collection",           "Jasper feature", ""),
    ("ID", "Collection",  "Lee Collection",            "https://interiordefine.com/lee-collection",              "Lee feature", ""),
    ("ID", "Collection",  "Saylor Collection",         "https://interiordefine.com/saylor-collection",           "Saylor feature", ""),

    # ── ID — Content ──────────────────────────────────────────────────────────
    ("ID", "Content",     "Design Services",           "https://interiordefine.com/design-services",             "Free design consultation offer", ""),
    ("ID", "Content",     "Sectional Buying Guide",    "https://interiordefine.com/sectional-buying-guide",      "Educational / nurture", ""),
    ("ID", "Content",     "Rug Buying Guide",          "https://interiordefine.com/rug-buying-guide",            "Rug-specific nurture", ""),
    ("ID", "Content",     "Performance Fabrics Guide", "https://interiordefine.com/performance-fabrics-guide",   "Fabric education", ""),
    ("ID", "Content",     "Comfort Guide",             "https://interiordefine.com/comfort-guide",               "Consideration-stage nurture", ""),
    ("ID", "Content",     "Shop the Catalog",          "https://interiordefine.com/shop-the-catalog",            "Full catalog discovery", ""),
    ("ID", "Content",     "Best Sellers",              "https://interiordefine.com/best-sellers",                "Re-engagement", ""),
    ("ID", "Content",     "Contract Grade",            "https://interiordefine.com/contract-grade",              "Trade / commercial segment", ""),
    ("ID", "Content",     "Book (Consult)",            "https://interiordefine.com/book",                        "Design consultation conversion", ""),

    # ── ID — Location ─────────────────────────────────────────────────────────
    ("ID", "Location",    "All Locations",             "https://interiordefine.com/locations",                   "Event-driven / store visits", ""),
    ("ID", "Location",    "Dallas",                    "https://interiordefine.com/locations/dallas",            "Geo-targeted campaigns", ""),
    ("ID", "Location",    "Boston",                    "https://interiordefine.com/locations/boston",            "Geo-targeted campaigns", ""),
    ("ID", "Location",    "Baltimore",                 "https://interiordefine.com/locations/baltimore",         "Geo-targeted campaigns", ""),

    # ── ID — Utility ──────────────────────────────────────────────────────────
    ("ID", "Utility",     "Gift Card",                 "https://interiordefine.com/gift-card",                   "Holiday / gifting", ""),
    ("ID", "Utility",     "Sale",                      "https://interiordefine.com/sale",                        "Sale-specific campaigns", ""),
    ("ID", "Utility",     "Cart",                      "https://interiordefine.com/cart",                        "Abandoned cart flows only", ""),
    ("ID", "Utility",     "Contact Us",                "https://interiordefine.com/contact-us",                  "CS-driven or support follow-up", ""),
    ("ID", "Utility",     "Fall Preview",              "https://interiordefine.com/fall-preview",                "Seasonal preview", ""),
    ("ID", "Utility",     "Spring Edit",               "https://interiordefine.com/the-spring-edit",             "Seasonal preview", ""),
    ("ID", "Utility",     "Swatches (subdomain)",      "https://swatches.interiordefine.com",                    "Free swatch request CTAs", ""),
    ("ID", "Utility",     "Trade (subdomain)",         "https://trade.interiordefine.com",                       "Trade program / B2B emails", ""),

    # ── ID — Third-Party ──────────────────────────────────────────────────────
    ("ID", "Third-Party", "Typeform Feedback (shared w/ CZ)", "https://form.typeform.com/to/avQR2W9q", "Shared feedback survey", ""),
    ("ID", "Third-Party", "Google Forms VIP Survey",          "https://docs.google.com/forms/...",     "Customer feedback / VIP surveys", ""),

    # ── STF — Category ────────────────────────────────────────────────────────
    ("STF", "Category",   "Homepage",                  "https://stfrank.com",                                    "Default CTA", ""),
    ("STF", "Category",   "Pillows",                   "https://stfrank.com/collections/pillows",                "Pillow-focused campaigns (most used)", ""),
    ("STF", "Category",   "Bedding",                   "https://stfrank.com/collections/bedding",                "Bedding spotlight", ""),
    ("STF", "Category",   "Wallpaper",                 "https://stfrank.com/collections/wallpaper",              "Wallpaper feature", ""),
    ("STF", "Category",   "Window Treatments",         "https://stfrank.com/collections/window-treatments",      "Curtain/drapery feature", ""),
    ("STF", "Category",   "Fabric by the Yard",        "https://stfrank.com/collections/fabric-by-the-yard",     "Fabric yardage purchase", ""),
    ("STF", "Category",   "Fabric Custom",             "https://stfrank.com/collections/fabric-custom",          "Custom order CTAs", ""),
    ("STF", "Category",   "New Release",               "https://stfrank.com/collections/new-release",            "New pattern launches", ""),
    ("STF", "Category",   "Outdoor Fabric",            "https://stfrank.com/collections/outdoor-fabric",         "Outdoor/summer feature", ""),
    ("STF", "Category",   "Outdoor Pillows",           "https://stfrank.com/collections/outdoor-pillows",        "Outdoor pillow highlight", ""),
    ("STF", "Category",   "All Outdoor",               "https://stfrank.com/collections/all-outdoor",            "Outdoor category page", ""),
    ("STF", "Category",   "Shop All Furniture",        "https://stfrank.com/collections/shop-all-furniture",     "Furniture cross-sell", ""),
    ("STF", "Category",   "Furniture",                 "https://stfrank.com/collections/furniture",              "Furniture category (broader)", ""),
    ("STF", "Category",   "Rugs",                      "https://stfrank.com/collections/rugs",                   "Rug feature emails", ""),
    ("STF", "Category",   "Decor",                     "https://stfrank.com/collections/decor",                  "General décor/accessories", ""),
    ("STF", "Category",   "Swatches",                  "https://stfrank.com/collections/swatches",               "Free swatch offer", ""),
    ("STF", "Category",   "Best Sellers",              "https://stfrank.com/collections/best-seller",            "Re-engagement / top picks", ""),
    ("STF", "Category",   "Back in Stock",             "https://stfrank.com/collections/back-in-stock",          "BIS / restocked product alerts", ""),
    ("STF", "Category",   "Quick Ship",                "https://stfrank.com/collections/quick-ship",             "Lead time / availability messaging", ""),
    ("STF", "Category",   "Art & Curiosities",         "https://stfrank.com/collections/art-curiosities",        "Wall art / decor feature", ""),
    ("STF", "Category",   "Sale",                      "https://stfrank.com/collections/sale",                   "General sale landing page", ""),

    # ── STF — Sale ────────────────────────────────────────────────────────────
    ("STF", "Sale",       "Studio Sale",               "https://stfrank.com/collections/the-studio-sale",          "Clearance / archive sale", ""),
    ("STF", "Sale",       "Spring Event",              "https://stfrank.com/collections/the-spring-event",         "Seasonal sale", ""),
    ("STF", "Sale",       "Winter Refresh Event",      "https://stfrank.com/collections/the-winter-refresh-event", "Winter sale event", ""),
    ("STF", "Sale",       "Outdoor Flash Sale",        "https://stfrank.com/collections/the-outdoor-flash-sale",   "Flash sale / urgency", ""),
    ("STF", "Sale",       "Black Friday Sale",         "https://stfrank.com/collections/black-friday-sale",        "BFCM campaigns", ""),
    ("STF", "Sale",       "Sample Sale",               "https://stfrank.com/collections/sample-sale",              "Clearance / one-off availability", ""),
    ("STF", "Sale",       "Yardage Sale",              "https://stfrank.com/collections/yardage-sale",             "Fabric-by-the-yard discount", ""),
    ("STF", "Sale",       "Pillow Sale",               "https://stfrank.com/collections/pillow-sale",              "Pillow-specific markdown", ""),
    ("STF", "Sale",       "Collaboration Sale",        "https://stfrank.com/collections/collaboration-sale",       "End-of-collaboration clearance", ""),

    # ── STF — Pattern / Collection ────────────────────────────────────────────
    ("STF", "Pattern",    "Gary Linden × St. Frank",        "https://stfrank.com/collections/gary-linden-x-st-frank",          "Collaboration launch", ""),
    ("STF", "Pattern",    "Forsyth × St. Frank",             "https://stfrank.com/collections/forsyth-x-st-frank",               "Collaboration feature", ""),
    ("STF", "Pattern",    "Forsyth × St. Frank Pillows",     "https://stfrank.com/collections/forsyth-x-st-frank-pillows",       "Collaboration pillow spotlight", ""),
    ("STF", "Pattern",    "Forsyth × St. Frank Rugs",        "https://stfrank.com/collections/forsyth-x-st-frank-rugs",          "Collaboration rug spotlight", ""),
    ("STF", "Pattern",    "Etkie × St. Frank",               "https://stfrank.com/collections/st-frank-x-etkie",                 "Collaboration feature", ""),
    ("STF", "Pattern",    "Sally King Benedict × St. Frank", "https://stfrank.com/collections/sally-king-benedict-x-st-frank",   "Collaboration feature", ""),
    ("STF", "Pattern",    "The Foggy Dog × St. Frank",       "https://stfrank.com/collections/the-foggy-dog-x-st-frank",         "Pet-themed collaboration", ""),
    ("STF", "Pattern",    "Mexico City in Photographs",      "https://stfrank.com/collections/robert-malmberg-x-st-frank",       "Photography / artisan collaboration", ""),
    ("STF", "Pattern",    "Green Lattice & Indigo Daisy",    "https://stfrank.com/collections/green-lattice-baule-indigo-daisy-suzani", "Pattern pairing spotlight", ""),
    ("STF", "Pattern",    "Suzani",                          "https://stfrank.com/collections/suzani",                           "Suzani pattern feature", ""),
    ("STF", "Pattern",    "Coastal Cool",                    "https://stfrank.com/collections/coastal-cool",                     "Coastal / summer feature", ""),
    ("STF", "Pattern",    "Fuchsia Daisy Suzani",      "https://stfrank.com/collections/fuchsia-daisy-suzani",           "Pattern spotlight", ""),
    ("STF", "Pattern",    "Espresso Checkerboard",     "https://stfrank.com/collections/espresso-checkerboard-suzani",   "Pattern spotlight", ""),
    ("STF", "Pattern",    "Teal Vines Suzani",         "https://stfrank.com/collections/teal-vines-suzani",              "Pattern spotlight", ""),
    ("STF", "Pattern",    "Black Daisy Suzani",        "https://stfrank.com/collections/black-daisy-suzani",             "Pattern spotlight", ""),
    ("STF", "Pattern",    "Chambray Lattice Baule",    "https://stfrank.com/collections/chambray-lattice-baule",         "Pattern spotlight", ""),
    ("STF", "Pattern",    "Sage Ribbon Suzani Bedding","https://stfrank.com/collections/sage-ribbon-suzani-bedding",     "Bedding + pattern bundle", ""),
    ("STF", "Pattern",    "Shell Daisy Suzani",        "https://stfrank.com/collections/shell-daisy-suzani",             "Pattern spotlight", ""),
    ("STF", "Pattern",    "Perfect Pairings",          "https://stfrank.com/collections/perfect-pairings",               "Cross-sell / how-to-style", ""),
    ("STF", "Pattern",    "French Pleat Curtains",     "https://stfrank.com/collections/french-pleat-curtains",          "Curtain-focused CTAs", ""),

    # ── STF — Bedding Sub ─────────────────────────────────────────────────────
    ("STF", "Bedding",    "Sheet Sets",                "https://stfrank.com/collections/sheet-sets",             "Bedding detail / separation", ""),
    ("STF", "Bedding",    "Duvets",                    "https://stfrank.com/collections/duvets",                 "Duvet-specific", ""),
    ("STF", "Bedding",    "Quilts / Coverlets",        "https://stfrank.com/collections/quilts-coverlets",       "Seasonal warm/layer", ""),
    ("STF", "Bedding",    "Tabletop Linens",           "https://stfrank.com/collections/tabletop-linens",        "Entertaining / dining", ""),
    ("STF", "Bedding",    "Bedding Bundles",           "https://stfrank.com/collections/bedding-bundles",        "Value / gifting", ""),

    # ── STF — Destination ─────────────────────────────────────────────────────
    ("STF", "Destination","Destination: Paris",        "https://stfrank.com/collections/destination-paris",        "Travel/editorial themed", ""),
    ("STF", "Destination","Destination: Nantucket",    "https://stfrank.com/collections/destination-nantucket",    "Coastal/seasonal", ""),
    ("STF", "Destination","Destination: Lake Como",    "https://stfrank.com/collections/destination-lake-como",    "Aspirational lifestyle", ""),
    ("STF", "Destination","Destination: Tuscany",      "https://stfrank.com/collections/destination-tuscany",      "Aspirational lifestyle", ""),
    ("STF", "Destination","Destination: Venice",        "https://stfrank.com/collections/destination-venice",       "Aspirational lifestyle", ""),
    ("STF", "Destination","Italian Getaway",           "https://stfrank.com/collections/your-italian-getaway",     "Travel/editorial", ""),

    # ── STF — Content ─────────────────────────────────────────────────────────
    ("STF", "Content",    "Press",                     "https://stfrank.com/press",                              "Brand credibility; linked in virtually every email as footer nav", ""),
    ("STF", "Content",    "FAQ",                       "https://stfrank.com/pages/faq",                          "Objection-handling / post-purchase", ""),
    ("STF", "Content",    "Shop the Look: Dining",     "https://stfrank.com/pages/shop-the-look-dining-rooms",   "Styled room inspiration", ""),
    ("STF", "Content",    "Shop the Look: Living",     "https://stfrank.com/pages/shop-the-look-living-rooms",   "Styled room inspiration", ""),
    ("STF", "Content",    "Shop the Look: Bedrooms",   "https://stfrank.com/pages/shop-the-look-bedrooms",       "Styled room inspiration", ""),
    ("STF", "Content",    "Style Guide",               "https://stfrank.com/pages/style-guide",                  "How-to-style educational", ""),
    ("STF", "Content",    "Trade",                     "https://stfrank.com/pages/trade",                        "Trade program", ""),

    # ── STF — Utility ─────────────────────────────────────────────────────────
    ("STF", "Utility",    "Gift Card",                 "https://stfrank.com/products/gift-card-1",               "Holiday / gifting", ""),

    # ── STF: Seasonal Edits ───────────────────────────────────────────────────
    ("STF", "Seasonal",   "The Fall Edit",             "https://stfrank.com/collections/the-fall-edit",          "Fall seasonal email", ""),
    ("STF", "Seasonal",   "The Winter Edit",           "https://stfrank.com/collections/the-winter-edit",        "Winter seasonal email", ""),
    ("STF", "Seasonal",   "The Spring Edit",           "https://stfrank.com/collections/the-spring-edit",        "Spring seasonal email", ""),
    ("STF", "Seasonal",   "Summer Essentials",         "https://stfrank.com/collections/summer-essentials",      "Summer seasonal email", ""),
    ("STF", "Seasonal",   "The Atelier Collection",    "https://stfrank.com/collections/the-atelier",            "Elevated / curated capsule", ""),

    # ── STF: Tabletop & Dining ────────────────────────────────────────────────
    ("STF", "Tabletop",   "Tabletop",                  "https://stfrank.com/collections/tabletop",               "General tabletop / entertaining", ""),
    ("STF", "Tabletop",   "Entertaining",              "https://stfrank.com/collections/entertaining",           "Entertaining-focused emails", ""),
    ("STF", "Tabletop",   "Dinnerware",                "https://stfrank.com/collections/dinnerware",             "Dinnerware feature", ""),
    ("STF", "Tabletop",   "Serveware",                 "https://stfrank.com/collections/serveware",              "Serving pieces spotlight", ""),
    ("STF", "Tabletop",   "Glassware",                 "https://stfrank.com/collections/glassware",              "Glassware feature", ""),
    ("STF", "Tabletop",   "Table Linens",              "https://stfrank.com/collections/table-linens",           "Table linens (overlaps with tabletop linens)", ""),
    ("STF", "Tabletop",   "Tabletop Linens + Decor",   "https://stfrank.com/collections/tabletop-linens-decor",  "Table styling roundup", ""),

    # ── STF: Art & Framed ─────────────────────────────────────────────────────
    ("STF", "Art",        "Framed Art",                "https://stfrank.com/collections/framed-textiles-and-prints", "Framed art spotlight", ""),
    ("STF", "Art",        "Framed Textiles",           "https://stfrank.com/collections/framed-textiles",            "Framed textile feature", ""),
    ("STF", "Art",        "Framed Prints",             "https://stfrank.com/collections/prints",                     "Print collection", ""),
    ("STF", "Art",        "Vintage Art",               "https://stfrank.com/collections/vintage-art",                "Vintage / one-of-a-kind art", ""),
    ("STF", "Art",        "Photography",               "https://stfrank.com/collections/photography",                "Photography art feature", ""),

    # ── STF: Specialty Rugs & Textiles ────────────────────────────────────────
    ("STF", "Textiles",   "Cactus Silk Rugs",          "https://stfrank.com/collections/cactus-silk-rugs",       "Cactus silk rug feature", ""),
    ("STF", "Textiles",   "Boujaad Rugs",              "https://stfrank.com/collections/boujaad-rugs",           "Moroccan rug spotlight", ""),
    ("STF", "Textiles",   "Kilim Collection",          "https://stfrank.com/collections/kilim-collection",       "Kilim rug feature", ""),
    ("STF", "Textiles",   "Oaxacan Embroidery",        "https://stfrank.com/collections/oaxacan-embroidery",     "Artisan / origin-story content", ""),
    ("STF", "Textiles",   "Huipil Collection",         "https://stfrank.com/collections/huipil-pillows-art",     "Artisan / cultural origin content", ""),
    ("STF", "Textiles",   "Linen Bedding",             "https://stfrank.com/collections/linen-bedding",          "Linen-specific bedding feature", ""),
    ("STF", "Textiles",   "Cotton Percale Bedding",    "https://stfrank.com/collections/cotton-percale-bedding", "Percale-specific bedding feature", ""),

    # ── STF: Gift Collections ─────────────────────────────────────────────────
    ("STF", "Gift",       "Gifts",                     "https://stfrank.com/collections/gifts",                  "Gift-focused campaigns", ""),
    ("STF", "Gift",       "Gift Guide",                "https://stfrank.com/collections/gift-guide",             "Holiday gift guide CTA", ""),
    ("STF", "Gift",       "Gift Sets",                 "https://stfrank.com/collections/gift-bundles",           "Bundled gift sets", ""),
    ("STF", "Gift",       "For Her",                   "https://stfrank.com/collections/for-her",                "Gifting for her segment", ""),
    ("STF", "Gift",       "For Him",                   "https://stfrank.com/collections/for-him",                "Gifting for him segment", ""),
    ("STF", "Gift",       "For the Host",              "https://stfrank.com/collections/gifts-for-the-host",     "Entertaining / host gift CTA", ""),
    ("STF", "Gift",       "Mother's Day",              "https://stfrank.com/collections/mothers-day",            "Mother's Day campaign", ""),
    ("STF", "Gift",       "Valentine's Day",           "https://stfrank.com/collections/valentines-day",         "Valentine's Day campaign", ""),

    # ── STF: Holiday ──────────────────────────────────────────────────────────
    ("STF", "Holiday",    "Holiday",                   "https://stfrank.com/collections/holiday",                "General holiday landing", ""),
    ("STF", "Holiday",    "Holiday Decor",             "https://stfrank.com/collections/holiday-decor",          "Holiday décor spotlight", ""),
    ("STF", "Holiday",    "Stocking Stuffers",         "https://stfrank.com/collections/stocking-stuffers",      "Stocking-stuffer gift guide", ""),
    ("STF", "Holiday",    "For the Tree",              "https://stfrank.com/collections/for-the-tree",           "Tree ornaments / holiday decor", ""),
    ("STF", "Holiday",    "Stockings & Tree Skirts",   "https://stfrank.com/collections/stockings-tree-skirts",  "Holiday accessories", ""),

    # ── STF: Nursery & Kids ───────────────────────────────────────────────────
    ("STF", "Kids",       "Nursery + Kids",            "https://stfrank.com/collections/nursery-kids",           "Kids / nursery segment", ""),
    ("STF", "Kids",       "Nursery + Kids Decor",      "https://stfrank.com/collections/nursery-kids-decor",     "Nursery décor feature", ""),

    # ── The Inside (TI) ───────────────────────────────────────────────────────
    ("TI",  "Category",   "Homepage",                  "https://theinside.com",                                            "Default CTA", ""),
    ("TI",  "Category",   "Living Room Edit",          "https://www.theinside.com/collections/living-room-edit",           "Living room / sofa feature", ""),
    ("TI",  "Category",   "The Bedroom Edit",          "https://www.theinside.com/collections/the-bedroom-edit",           "Bedroom-focused campaign", ""),
    ("TI",  "Category",   "Benches & Ottomans",        "https://www.theinside.com/collections/benchesandottomans",         "Accent furniture feature", ""),
    ("TI",  "Category",   "Kids' Furniture",           "https://www.theinside.com/collections/kids-furniture",             "Kids' room segment", ""),

    ("TI",  "Fabric",     "Fabric Swatches",           "https://www.theinside.com/fabric-swatches",                        "Free swatch offer; fabric-first messaging", ""),
    ("TI",  "Fabric",     "Decide Fabric Later",       "https://www.theinside.com/collections/decide-fabric-later",        "Low-commitment entry CTA", ""),

    ("TI",  "Seasonal",   "Spring Into Style",         "https://www.theinside.com/collections/spring-2025-trends",         "Spring seasonal send", ""),
    ("TI",  "Seasonal",   "Garden Party Edit",         "https://www.theinside.com/collections/garden-party-edit",          "Spring / outdoor entertaining", ""),
    ("TI",  "Seasonal",   "French Riviera Era",        "https://www.theinside.com/collections/frenchrivieraera",           "Aspirational summer editorial", ""),
    ("TI",  "Seasonal",   "Country Club Edit",         "https://www.theinside.com/collections/country-club-edit",          "Preppy / summer lifestyle", ""),
    ("TI",  "Seasonal",   "Hosting Edit",              "https://www.theinside.com/collections/summer-dining-edit",         "Dining / entertaining feature", ""),
    ("TI",  "Seasonal",   "Hosting Essentials",        "https://www.theinside.com/collections/hosting-essentials",         "Dining / hosting roundup", ""),

    ("TI",  "Pattern",    "Marigold Delphine",         "https://www.theinside.com/collections/marigolddelphine",           "Pattern spotlight", ""),
    ("TI",  "Pattern",    "Delphine",                  "https://www.theinside.com/collections/delphine",                   "Pattern spotlight", ""),
    ("TI",  "Pattern",    "Florals",                   "https://www.theinside.com/collections/floral",                     "Floral print feature", ""),
    ("TI",  "Pattern",    "Stripes",                   "https://www.theinside.com/collections/striped-furniture",          "Stripe pattern feature", ""),
    ("TI",  "Pattern",    "Tigresse",                  "https://www.theinside.com/collections/tigresse",                   "Pattern spotlight", ""),
    ("TI",  "Pattern",    "Coastal Cool",              "https://www.theinside.com/collections/coastalcool",                "Coastal / casual palette", ""),
    ("TI",  "Pattern",    "Coastal Fisherman",         "https://www.theinside.com/collections/trending-coastal-fisherman", "Trending coastal look", ""),
    ("TI",  "Pattern",    "Cherry Blossom",            "https://www.theinside.com/collections/cherry-blossom",             "Spring / floral pattern", ""),
    ("TI",  "Pattern",    "Citrus Season",             "https://www.theinside.com/collections/citrusseason",               "Summer / bright palette", ""),
    ("TI",  "Pattern",    "Citrine Cabana Stripe",     "https://www.theinside.com/collections/citrine-cabana-stripe",      "Summer stripe feature", ""),
    ("TI",  "Pattern",    "Summer Blues",              "https://www.theinside.com/collections/blue-furniture",             "Blue colorway / summer", ""),
    ("TI",  "Pattern",    "It Was All Yellow",         "https://www.theinside.com/collections/yellow-furniture",           "Yellow colorway feature", ""),
    ("TI",  "Pattern",    "Central Park Toile",        "https://www.theinside.com/collections/central-park-toile",         "Toile / NYC-themed pattern", ""),
    ("TI",  "Pattern",    "Animal Prints",             "https://www.theinside.com/collections/animal-prints",              "Animal print feature", ""),

    ("TI",  "Destination","Hudson Valley",             "https://www.theinside.com/collections/destination-hudson-valley",  "Destination / lifestyle editorial", ""),
    ("TI",  "Destination","New England",               "https://www.theinside.com/collections/new-england-summer",         "Coastal / preppy summer", ""),
    ("TI",  "Destination","Italy Travel Edit",         "https://www.theinside.com/collections/italy-travel-edit",          "Aspirational travel editorial", ""),

    ("TI",  "Collab",     "CW Stockwell",              "https://www.theinside.com/collaborators/cw-stockwell",             "Collaboration launch / feature", ""),
]


def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def build():
    wb = Workbook()

    # ── Sheet 1: All Links ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Links"

    headers = ["Brand", "Section", "Page / Link Name", "URL", "When to Use", "Notes"]
    col_widths = [8, 20, 32, 68, 55, 45]

    # Header row
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = make_fill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Freeze header
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data rows
    for row_idx, row_data in enumerate(ROWS, 2):
        brand = row_data[0]
        is_warning = row_data[1].startswith("⚠")
        bg = NOTE_BG if is_warning else BRAND_COLORS.get(brand, LIGHTBG)
        if row_idx % 2 == 0 and not is_warning:
            # Slightly darker alternate for even rows within same brand color
            pass

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=9,
                             bold=(col_idx == 1 or is_warning),
                             color="B45309" if is_warning else TEXTMID)
            cell.fill = make_fill(bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

        ws.row_dimensions[row_idx].height = 28

    # ── Sheet 2: PDP Variant Rules ────────────────────────────────────────────
    ws2 = wb.create_sheet("PDP Variant Rules")
    ws2.sheet_properties.tabColor = "E5B800"

    rules_content = [
        ("Rule", "Details"),
        ("General rule",
         "When an email features a specific product in a specific color/fabric/finish, always link "
         "directly to that variant — not the base PDP. Showing a blue sofa in the email but landing "
         "on the default beige creates friction and increases bounce."),
        ("URL format (Shopify)",
         "https://[brand].com/products/[product-slug]?[Option Name]=[Option Value]\n"
         "Parameter names and values must match the option names configured in the product exactly, "
         "including capitalization and spacing."),
        ("How to get the correct URL",
         "Navigate to the PDP on the live site, select the desired variant, and copy the URL from "
         "the browser address bar — it will have the correct parameter names and encoded values."),
        ("⚠ Burrow only — Braze encoding bug",
         "All spaces in Burrow PDP query parameters MUST be encoded as %20, never as +.\n\n"
         "Braze re-encodes + to %2B on save. This means:\n"
         "• If you enter Wood+Finish=Walnut+-+Wood, Braze saves it as Wood%2BFinish=Walnut%20-%20Wood\n"
         "• The parameter key becomes 'Wood+Finish' (literal plus) instead of 'Wood Finish', so the "
         "PDP variant-matching script can't find a match and falls back to the default variant.\n\n"
         "Always verify the final saved URL in Braze before launching to confirm no %2B appears "
         "where a space is intended."),
        ("Burrow example (correct)",
         "https://burrow.com/products/nomad-plus-sofa?Wood%20Finish=Walnut%20-%20Wood&Fabric=Nomad%20Linen%20Natural"),
    ]

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 90

    for r_idx, (label, detail) in enumerate(rules_content, 1):
        is_hdr = r_idx == 1
        is_warning = label.startswith("⚠")
        bg = NAVY if is_hdr else (NOTE_BG if is_warning else (HDRBG if r_idx % 2 == 0 else WHITE))
        txt_color = WHITE if is_hdr else ("B45309" if is_warning else TEXTDRK)

        for c_idx, val in enumerate([label, detail], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", bold=(is_hdr or c_idx == 1), size=9 if not is_hdr else 10,
                             color=txt_color)
            cell.fill = make_fill(bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws2.row_dimensions[r_idx].height = 80 if is_warning else (20 if is_hdr else 45)

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Notes ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Notes")
    ws3.sheet_properties.tabColor = "6B7280"

    notes = [
        ("Item", "Note"),
        ("CZ /cart",             "Appears only in triggered/abandonment flows, not batch sends"),
        ("ID /cart",             "Appears only in triggered/abandonment flows, not batch sends"),
        ("HAV /cart",            "Appears only in triggered/abandonment flows, not batch sends"),
        ("STF /press",           "Linked in virtually every email as a footer navigation item, not a CTA"),
        ("BUR sister-brand links","ID, CZ, HAV, STF, TI homepages appear 29x each — single template variant with sister-brand footer links on every send"),
        ("ID clicks.interiordefine.com","Legacy click-tracking URLs from an older email platform — should be phased out in favor of Braze native tracking"),
        ("HAV app links",        "HAV is the only brand with an iOS app — reflected in Apple App Store links"),
        ("BUR PDP variant links","Always use %20 for spaces in query parameters. Braze converts + → %2B on save, breaking variant selection. See PDP Variant Rules sheet for full details."),
        ("Data basis",           "Analysis based on 1,556 email HTML files sent in the last year across HAV, CZ, ID, BUR (Aug 2025+), and STF"),
    ]

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 90

    for r_idx, (label, note) in enumerate(notes, 1):
        is_hdr = r_idx == 1
        bg = NAVY if is_hdr else (ALTROW if r_idx % 2 == 0 else WHITE)
        txt_color = WHITE if is_hdr else TEXTDRK

        for c_idx, val in enumerate([label, note], 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", bold=(is_hdr or c_idx == 1), size=9 if not is_hdr else 10,
                             color=txt_color)
            cell.fill = make_fill(bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws3.row_dimensions[r_idx].height = 40 if not is_hdr else 22

    ws3.freeze_panes = "A2"

    # ── Sheets 4-8: Per-brand ─────────────────────────────────────────────────
    BRAND_NAMES = {
        "BUR": "Burrow",
        "CZ":  "The Citizenry",
        "HAV": "Havenly",
        "ID":  "Interior Define",
        "STF": "St. Frank",
        "TI":  "The Inside",
    }
    BRAND_TAB_COLORS = {
        "BUR": "3B82F6",
        "CZ":  "10B981",
        "HAV": "EF4444",
        "ID":  "8B5CF6",
        "STF": "F59E0B",
        "TI":  "0F766E",
    }

    brand_headers = ["Section", "Page / Link Name", "URL", "When to Use", "Notes"]
    brand_col_widths = [20, 32, 68, 55, 45]

    for brand_code, brand_full in BRAND_NAMES.items():
        brand_rows = [r for r in ROWS if r[0] == brand_code]
        bg_color = BRAND_COLORS[brand_code]

        wsb = wb.create_sheet(brand_full)
        wsb.sheet_properties.tabColor = BRAND_TAB_COLORS[brand_code]

        # Header
        for col_idx, (hdr, width) in enumerate(zip(brand_headers, brand_col_widths), 1):
            cell = wsb.cell(row=1, column=col_idx, value=hdr)
            cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
            cell.fill = make_fill(NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
            wsb.column_dimensions[get_column_letter(col_idx)].width = width
        wsb.row_dimensions[1].height = 22
        wsb.freeze_panes = "A2"
        wsb.auto_filter.ref = f"A1:{get_column_letter(len(brand_headers))}1"

        # Data (strip brand column — cols 1-5 of original = cols 0-4, skip col 0)
        for row_idx, row_data in enumerate(brand_rows, 2):
            is_warning = row_data[1].startswith("⚠")
            row_bg = NOTE_BG if is_warning else bg_color

            for col_idx, value in enumerate(row_data[1:], 1):  # skip brand column
                cell = wsb.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=9,
                                 bold=(col_idx == 1 or is_warning),
                                 color="B45309" if is_warning else TEXTMID)
                cell.fill = make_fill(row_bg)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border()
            wsb.row_dimensions[row_idx].height = 28

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Rows in All Links sheet: {len(ROWS)}")


if __name__ == "__main__":
    build()

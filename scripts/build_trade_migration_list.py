#!/usr/bin/env python3
"""
Build the ID Trade -> The Expert migration list.

This produces a sequencing spreadsheet for BACKEND ACCOUNT CREATION, not a Klaviyo
import. Accounts get created batch by batch; Klaviyo lists are only built afterwards.
That is why contacts carrying a contact_brand_origin value are included here even
though scripts/sync_trade_approved_to_klaviyo.py deliberately excludes them -- that
guard exists to stop *automatic* Klaviyo creation, which this flow avoids by design.

Five sheets:
  1. Migration batches      -- every trade contact EXCEPT batches 16-19 (see tabs 4/5),
                               one batch number each
  2. Already in The Expert -- trade contacts whose email is already in TE HubSpot
  3. Trade approved gaps    -- ID HubSpot trade-approved but not getting trade mail
  4. Hold - data gaps       -- batch 19 (incomplete name and/or no ID HubSpot record)
  5. Do not migrate         -- batches 16-18: chronic-bouncer / unsubscribed /
                               never-opened-in-12mo+ cohorts

Tabs 4 and 5 are EXTRACTED out of tab 1, not just flagged there -- neither the
data-quality hold nor the do-not-migrate cohorts appear on tab 1 at all. This isn't
a hand-picked exclusion list -- those three cohorts are never members of COHORT_ORDER
(the ramp), so every contact in them lands in exactly batch 16/17/18/19 with no other
possibility, meaning excluding those batch numbers from tab 1 excludes exactly this
population and nothing else. Tabs 4 and 5 can still overlap each other (e.g. a chronic
bouncer who's also missing a name appears on both, sitting in batch 19) since they
answer different questions: "do we have enough info to migrate this person" vs
"should we migrate this person at all."

Tabs 2 and 3 never share a row (tab 2 is a subset of tab 1's trade list; tab 3 is
explicitly contacts NOT on it) but both independently answer "already in TE?" for
their own population -- tab 3 carries its own in_te_hubspot column for that.

Subscription status comes from Braze /users/export/ids, which accepts only ONE email
per request (~4 req/s, so ~90 min for the full list). Results are cached to JSON so
the pass is resumable and reruns are free. Run the status pass on its own first:

    uv run python scripts/build_trade_migration_list.py --status-only

then build the workbook:

    uv run python scripts/build_trade_migration_list.py --out exports/trade_migration.xlsx

Other usage:
    uv run python scripts/build_trade_migration_list.py --dry-run
    uv run python scripts/build_trade_migration_list.py --skip-braze-status
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

REPO = Path(__file__).parent.parent
load_dotenv(REPO / ".env")
sys.path.insert(0, str(Path(__file__).parent))

from snowflake_client import get_snowflake_client  # noqa: E402
from utils.braze_datashare import get_app_group_id, get_datashare_location  # noqa: E402

BRAND = "ID"
BRAZE_HOST = os.environ.get("BRAZE_BASE_URL", "https://rest.iad-07.braze.com").rstrip("/")
# The plain BRAZE_API_KEY_ID lacks users.export.ids scope (403); the _DELETE key has it.
BRAZE_STATUS_KEY_ENV = "BRAZE_API_KEY_ID_DELETE"
STATUS_CACHE = REPO / "data" / "trade_migration_braze_status.json"
DEFAULT_OUT = REPO / "exports" / "trade_migration_list.xlsx"

# Cohort release order. Batches 2-13 draw from the first six in this order.
COHORT_ORDER = [
    "clicked_90d",
    "opened_30d",
    "opened_31_90d",
    "opened_3_6mo",
    "opened_6_12mo",
    "never_opened_joined_12mo",
]

# Batch 1 is hand-picked, not drawn from the ramp order: a small internal-only test
# batch, Havenly-family contacts confirmed to be (a) on the ID trade list, (b) in ID
# HubSpot, and (c) NOT already existing in TE HubSpot or TE Klaviyo (checked directly
# against TE's Klaviyo account 2026-08-24 via KlaviyoClient.get_email_marketing_consent,
# since TE Klaviyo profiles don't sync to Snowflake). No external padding -- batch 1 is
# just these three, chosen by Jordan from the qualifying-candidate list (Robyn
# Pleggenkuhle also qualified but was dropped from the final pick).
TEAM_SEED_EMAILS = [
    "jordan.rubenstein+tradetest1@havenly.com",
    "carlos.gartner+swatches@interiordefine.com",
    "hcrockett@the-citizenry.com",
]
TEAM_SEED_EXTERNAL_COUNT = 0
TEAM_SEED_WEEK = "1"
TEAM_SEED_NOTE_TEAM = "Internal test batch - Havenly team member"
TEAM_SEED_NOTE_EXTERNAL = "Internal test batch - first live contacts"

# Batches 2-13: two per week, matching the 2 sends/week cadence. Batch 1 (above) is
# also week 1, so week 1 ends up with three batches instead of two.
BATCH_SIZES = [500, 500, 750, 750, 1250, 1250, 1750, 1750, 2250, 2250, 2379, 2379]
BATCH_WEEKS = [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6]
BATCH_START_NO = 2  # batch 1 is the hand-picked seed batch above, assigned separately

# Tail cohorts, each its own batch. (batch_no, cohort, week_label, marketing_eligible, note)
TAIL_BATCHES = [
    (14, "opened_12_18mo", "8+ (optional)", "Y", "Optional extension - 0.5% open rate on ID"),
    (15, "opened_18_24mo", "8+ (optional)", "Y", "Optional extension - 0.9% open rate on ID"),
    (16, "never_opened_joined_over_12mo", "after ramp", "N",
     "Account only - 0.4% open rate, excluded from marketing"),
    (17, "unsubscribed", "after ramp", "N",
     "Account + transactional SendGrid only - unsubscribed in Braze"),
    (18, "chronic_bouncer", "after ramp", "N",
     "Account only - undeliverable, 3+ bounces in 90d, do not mail"),
]

# Contacts failing either data-quality check -- an incomplete name (missing a first
# name, a last name, or both), or no ID HubSpot record at all -- are held back into
# this single bucket, pulled out of the ramp/tail pool entirely BEFORE batches 2-18
# are sized -- so batch 2 is still exactly 500 contacts, etc.; these contacts never
# displace someone with a usable name/record out of a batch. The two checks are
# independent: a contact with no HubSpot record can still have a complete name (Braze
# fills in FIRST_NAME/LAST_NAME on its own), and a contact with a real HubSpot record
# can still be missing a name. `hold_reason` records which check(s) failed; `cohort`
# records the engagement tier (chronic bouncer, unengaged, etc.) so nothing about
# standing is lost if the gap is resolved later and the contact moves back into the ramp.
NO_NAME_BATCH = 19
NO_NAME_WEEK = "hold - data gaps"
NO_NAME_NOTE = ("Held back pending a decision on whether to migrate -- see hold_reason for "
                "why (missing name and/or no ID HubSpot record) and cohort for the "
                "engagement tier this contact would otherwise sit in.")

# Cohorts that should never be migrated into Klaviyo as marketing contacts, regardless
# of which batch they currently sit in (a contact can be both a chronic bouncer AND
# a data-quality hold, and will show up on both the "Hold - data gaps" and "Do not
# migrate" lenses).
DO_NOT_MIGRATE_COHORTS = {
    "chronic_bouncer": "Chronic bouncer -- 3+ soft bounces on trade sends in the last 90 "
                       "days, likely a dead address",
    "never_opened_joined_over_12mo": "Never opened or clicked a trade send, and it's been "
                                     "over 12 months since their first one -- dormant",
    "unsubscribed": "Unsubscribed in Braze (or filed a spam complaint) -- do not remarket",
}

# Contacts with no brand_origin in ID HubSpot default to Interior Define. Deliberately
# the same literal HubSpot already uses for ID-origin contacts, so the column carries one
# label per brand instead of two synonyms the team would have to reconcile; the companion
# contact_brand_origin_source column is what preserves "was this set or defaulted".
DEFAULT_BRAND_ORIGIN = "Interior Define"


# ---------------------------------------------------------------- Snowflake

def _trade_list_sql() -> str:
    db, schema = get_datashare_location(BRAND)
    bz = f"{db}.{schema}"
    app = get_app_group_id(BRAND)
    return f"""
WITH aud AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MIN(TO_TIMESTAMP(TIME)) first_trade_send
  FROM {bz}.USERS_MESSAGES_EMAIL_SEND_SHARED
  WHERE APP_GROUP_ID='{app}' AND UPPER(CAMPAIGN_NAME) LIKE '%TRADE%'
    AND TO_TIMESTAMP(TIME) >= DATEADD('day',-90,CURRENT_TIMESTAMP())
    AND EMAIL_ADDRESS IS NOT NULL
  GROUP BY 1
),
first_send AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MIN(TO_TIMESTAMP(TIME)) f
  FROM {bz}.USERS_MESSAGES_EMAIL_SEND_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL
  GROUP BY 1
),
-- all opens, machine included, to match how Klaviyo reports
opens AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MAX(TO_TIMESTAMP(TIME)) last_open
  FROM {bz}.USERS_MESSAGES_EMAIL_OPEN_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL
  GROUP BY 1
),
clicks AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MAX(TO_TIMESTAMP(TIME)) last_click
  FROM {bz}.USERS_MESSAGES_EMAIL_CLICK_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL
    AND (IS_SUSPECTED_BOT_CLICK IS NULL OR IS_SUSPECTED_BOT_CLICK='false')
  GROUP BY 1
),
bouncers AS (
  SELECT LOWER(EMAIL_ADDRESS) em, COUNT(*) bounce_events
  FROM {bz}.USERS_MESSAGES_EMAIL_SOFTBOUNCE_SHARED
  WHERE APP_GROUP_ID='{app}' AND UPPER(CAMPAIGN_NAME) LIKE '%TRADE%'
    AND TO_TIMESTAMP(TIME) >= DATEADD('day',-90,CURRENT_TIMESTAMP())
  GROUP BY 1 HAVING COUNT(*) >= 3
),
optouts AS (
  SELECT em, MAX(t) last_optout FROM (
    SELECT LOWER(EMAIL_ADDRESS) em, TO_TIMESTAMP(TIME) t
      FROM {bz}.USERS_MESSAGES_EMAIL_UNSUBSCRIBE_SHARED WHERE APP_GROUP_ID='{app}'
    UNION ALL
    SELECT LOWER(EMAIL_ADDRESS), TO_TIMESTAMP(TIME)
      FROM {bz}.USERS_MESSAGES_EMAIL_MARKASSPAM_SHARED WHERE APP_GROUP_ID='{app}'
  ) WHERE em IS NOT NULL GROUP BY 1
),
braze_attrs AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MAX(FIRST_NAME) bz_first, MAX(LAST_NAME) bz_last,
         MAX(HOME_CITY) bz_city
  FROM {bz}.USER_DEFAULT_ATTRIBUTES_VIEW_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL
  GROUP BY 1
),
hs AS (
  SELECT LOWER(CONTACT_EMAIL) em,
         MAX(HUBSPOT_CONTACT_FULL_NAME) hs_full_name,
         MAX(COMPANY_NAME)              hs_company_name,
         MAX(STUDIO_NAME)               hs_studio_name,
         MAX(BRAND_ORIGIN)              brand_origin,
         MIN(TRADE_APPROVAL_AT)         trade_approval_at,
         MAX(HUBSPOT_CITY)              hs_city,
         MAX(HUBSPOT_STATE)             hs_state
  FROM PROD.ID_WAREHOUSE.STG_CONTACTS
  WHERE CONTACT_EMAIL IS NOT NULL
  GROUP BY 1
),
te_hs AS (
  SELECT LOWER(CONTACT_EMAIL_ADDRESS) em, MIN(CONTACT_CREATED_AT) te_created_at
  FROM PROD.ANALYTICS_THEEXPERT.HUBSPOT_CONTACTS
  WHERE CONTACT_EMAIL_ADDRESS IS NOT NULL
  GROUP BY 1
),
te_acct AS (
  SELECT DISTINCT LOWER(CLIENT_LOGIN_EMAIL) em
  FROM PROD.ANALYTICS_THEEXPERT.CUSTOMERS
  WHERE CLIENT_LOGIN_EMAIL IS NOT NULL
)
SELECT
  aud.em                                    AS email,
  hs.hs_full_name, hs.hs_company_name, hs.hs_studio_name,
  hs.brand_origin, hs.trade_approval_at, hs.hs_city, hs.hs_state,
  ba.bz_first, ba.bz_last, ba.bz_city,
  fs.f                                      AS first_send_ever,
  o.last_open, k.last_click,
  CASE WHEN b.em IS NOT NULL THEN 1 ELSE 0 END       AS is_bouncer,
  b.bounce_events,
  CASE WHEN oo.em IS NOT NULL THEN 1 ELSE 0 END      AS is_optout_datashare,
  oo.last_optout,
  CASE WHEN te.em IS NOT NULL THEN 1 ELSE 0 END      AS in_te_hubspot,
  te.te_created_at,
  CASE WHEN ta.em IS NOT NULL THEN 1 ELSE 0 END      AS has_te_platform_account
FROM aud
LEFT JOIN hs         ON hs.em = aud.em
LEFT JOIN braze_attrs ba ON ba.em = aud.em
LEFT JOIN first_send fs  ON fs.em = aud.em
LEFT JOIN opens o    ON o.em  = aud.em
LEFT JOIN clicks k   ON k.em  = aud.em
LEFT JOIN bouncers b ON b.em  = aud.em
LEFT JOIN optouts oo ON oo.em = aud.em
LEFT JOIN te_hs te   ON te.em = aud.em
LEFT JOIN te_acct ta ON ta.em = aud.em
"""


def _sheet3_sql() -> str:
    db, schema = get_datashare_location(BRAND)
    bz = f"{db}.{schema}"
    app = get_app_group_id(BRAND)
    return f"""
WITH trade_list AS (
  SELECT DISTINCT LOWER(EMAIL_ADDRESS) em
  FROM {bz}.USERS_MESSAGES_EMAIL_SEND_SHARED
  WHERE APP_GROUP_ID='{app}' AND UPPER(CAMPAIGN_NAME) LIKE '%TRADE%'
    AND TO_TIMESTAMP(TIME) >= DATEADD('day',-90,CURRENT_TIMESTAMP())
    AND EMAIL_ADDRESS IS NOT NULL
),
in_braze AS (
  SELECT DISTINCT LOWER(EMAIL_ADDRESS) em
  FROM {bz}.USERS_MESSAGES_EMAIL_SEND_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL
),
braze_attrs AS (
  SELECT LOWER(EMAIL_ADDRESS) em, MAX(FIRST_NAME) bz_first, MAX(LAST_NAME) bz_last,
         MAX(HOME_CITY) bz_city
  FROM {bz}.USER_DEFAULT_ATTRIBUTES_VIEW_SHARED
  WHERE APP_GROUP_ID='{app}' AND EMAIL_ADDRESS IS NOT NULL GROUP BY 1
),
approved AS (
  SELECT LOWER(CONTACT_EMAIL) em,
         MAX(HUBSPOT_CONTACT_FULL_NAME) hs_full_name,
         MAX(COMPANY_NAME) hs_company_name,
         MAX(STUDIO_NAME)  hs_studio_name,
         MAX(BRAND_ORIGIN) brand_origin,
         MIN(TRADE_APPROVAL_AT) trade_approval_at,
         MIN(TRADE_APPLICATION_RECEIVED_AT) trade_application_received_at,
         MAX(HUBSPOT_CITY) hs_city,
         MAX(HUBSPOT_STATE) hs_state
  FROM PROD.ID_WAREHOUSE.STG_CONTACTS
  WHERE CONTACT_EMAIL IS NOT NULL AND TRADE_APPROVAL_AT IS NOT NULL
  GROUP BY 1
),
te_hs AS (
  SELECT DISTINCT LOWER(CONTACT_EMAIL_ADDRESS) em
  FROM PROD.ANALYTICS_THEEXPERT.HUBSPOT_CONTACTS WHERE CONTACT_EMAIL_ADDRESS IS NOT NULL
)
SELECT a.em AS email, a.hs_full_name, a.hs_company_name, a.hs_studio_name,
  a.brand_origin, a.trade_approval_at, a.trade_application_received_at,
  a.hs_city, a.hs_state,
  ba.bz_first, ba.bz_last, ba.bz_city,
  CASE WHEN ib.em IS NULL THEN 'not_in_braze' ELSE 'in_braze_not_on_trade_list' END AS gap_type,
  CASE WHEN te.em IS NOT NULL THEN 1 ELSE 0 END AS in_te_hubspot
FROM approved a
LEFT JOIN in_braze ib ON ib.em = a.em
LEFT JOIN braze_attrs ba ON ba.em = a.em
LEFT JOIN te_hs te ON te.em = a.em
WHERE a.em NOT IN (SELECT em FROM trade_list)
"""


def fetch(sql: str) -> list[dict]:
    db, schema = get_datashare_location(BRAND)
    client = get_snowflake_client(schema=schema, database=db)
    return client.execute_query(sql)


# ---------------------------------------------------------------- Braze status

def fetch_braze_status(emails: list[str], limit: int | None = None) -> dict[str, str]:
    """
    Resolve email_subscribe for each address via /users/export/ids.

    The endpoint takes a single email per request (a batched email_addresses array
    returns 400), so this is inherently slow. Results are cached to JSON after every
    chunk, making the pass resumable across runs and interruptions.
    """
    key = os.environ.get(BRAZE_STATUS_KEY_ENV)
    if not key:
        raise SystemExit(f"{BRAZE_STATUS_KEY_ENV} not set in .env")

    cache: dict[str, str] = {}
    if STATUS_CACHE.exists():
        cache = json.loads(STATUS_CACHE.read_text())

    todo = [e for e in emails if e not in cache]
    if limit:
        todo = todo[:limit]
    print(f"  braze status: {len(cache):,} cached, {len(todo):,} to fetch")
    if not todo:
        return cache

    url = f"{BRAZE_HOST}/users/export/ids"
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    started = time.time()

    for i, email in enumerate(todo, 1):
        body = json.dumps({
            "email_address": email,
            "fields_to_export": ["email", "email_subscribe"],
        }).encode()
        for attempt in range(4):
            try:
                req = urllib.request.Request(url, data=body, headers=headers)
                with urllib.request.urlopen(req, timeout=30) as resp:
                    users = json.load(resp).get("users", [])
                # Multiple profiles can share an address; unsubscribed wins so we never
                # mail someone because a duplicate profile looked subscribed.
                states = {u.get("email_subscribe") for u in users if u.get("email_subscribe")}
                if not states:
                    cache[email] = "no_profile"
                elif "unsubscribed" in states:
                    cache[email] = "unsubscribed"
                elif "subscribed" in states:
                    cache[email] = "subscribed"
                else:
                    cache[email] = sorted(states)[0]
                break
            except urllib.error.HTTPError as exc:
                if exc.code == 429:
                    time.sleep(int(exc.headers.get("Retry-After", 5)))
                    continue
                if exc.code >= 500 and attempt < 3:
                    time.sleep(2 * (attempt + 1))
                    continue
                cache[email] = f"error_{exc.code}"
                break
            except Exception:
                if attempt < 3:
                    time.sleep(2 * (attempt + 1))
                    continue
                cache[email] = "error"
                break

        if i % 250 == 0 or i == len(todo):
            STATUS_CACHE.parent.mkdir(parents=True, exist_ok=True)
            STATUS_CACHE.write_text(json.dumps(cache))
            rate = i / max(time.time() - started, 0.01)
            eta = (len(todo) - i) / max(rate, 0.01) / 60
            print(f"    {i:,}/{len(todo):,}  {rate:.1f}/s  eta {eta:.0f}m", flush=True)

    STATUS_CACHE.write_text(json.dumps(cache))
    return cache


# ---------------------------------------------------------------- derivation

def _days_ago(ts, now) -> float | None:
    if ts is None:
        return None
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=timezone.utc)
    return (now - ts).total_seconds() / 86400.0


def classify(row: dict, now: datetime) -> str:
    """Mutually exclusive cohort. Bouncers and opt-outs take precedence over engagement."""
    if row.get("IS_BOUNCER"):
        return "chronic_bouncer"
    if row.get("IS_OPTOUT_DATASHARE") or row.get("_braze_unsubscribed"):
        return "unsubscribed"

    click = _days_ago(row.get("LAST_CLICK"), now)
    open_ = _days_ago(row.get("LAST_OPEN"), now)
    joined = _days_ago(row.get("FIRST_SEND_EVER"), now)

    if click is not None and click <= 90:
        return "clicked_90d"
    if open_ is None:
        return "never_opened_joined_12mo" if (joined is not None and joined <= 365) \
            else "never_opened_joined_over_12mo"
    if open_ <= 30:
        return "opened_30d"
    if open_ <= 90:
        return "opened_31_90d"
    if open_ <= 180:
        return "opened_3_6mo"
    if open_ <= 365:
        return "opened_6_12mo"
    if open_ <= 548:
        return "opened_12_18mo"
    return "opened_18_24mo"


def derive_company(row: dict) -> tuple[str, str]:
    """
    company_name -> full name (HubSpot) -> full name (Braze fallback). Returns
    (value, source).

    The third tier matters for contacts with no ID HubSpot record at all: derive_names()
    still resolves a name for them from Braze's FIRST_NAME/LAST_NAME defaults, so
    company_name should too rather than sitting blank next to a populated first/last
    name -- e.g. admin@dkorinteriors.com has no HubSpot record but Braze knows her as
    Ivonne Ronderos, so company_name should read "Ivonne Ronderos" as well.

    STUDIO_NAME is deliberately NOT in this chain. Despite the name it holds the ID
    showroom market the contact belongs to -- 14 distinct values, all cities (New York,
    Boston, Denver...) -- so using it as a company fallback writes "New York" into a
    company column for ~8.7k contacts. STUDIO_NAME is still selected in SQL in case it's
    needed for debugging, but it is no longer surfaced as a workbook column.
    """
    for field, label in (("HS_COMPANY_NAME", "company_name"),
                         ("HS_FULL_NAME", "full_name")):
        val = (row.get(field) or "").strip()
        if val:
            return val, label
    first, last = derive_names(row)
    full = " ".join(p for p in (first, last) if p)
    if full:
        return full, "braze_name"
    return "", "none"


def derive_names(row: dict) -> tuple[str, str]:
    """Split HubSpot full name; fall back to Braze first/last."""
    full = (row.get("HS_FULL_NAME") or "").strip()
    if full:
        parts = full.split()
        if len(parts) == 1:
            return parts[0], ""
        return " ".join(parts[:-1]), parts[-1]
    return (row.get("BZ_FIRST") or "").strip(), (row.get("BZ_LAST") or "").strip()


def derive_location(row: dict) -> tuple[str, str]:
    """
    city -> ID HubSpot's HUBSPOT_CITY, falling back to Braze's HOME_CITY when HubSpot
    has none. state -> ID HubSpot's HUBSPOT_STATE only; Braze has no state-level
    attribute to fall back to, so it's blank when HubSpot doesn't have it.
    """
    city = (row.get("HS_CITY") or "").strip() or (row.get("BZ_CITY") or "").strip()
    state = (row.get("HS_STATE") or "").strip()
    return city, state


def _has_complete_name(row: dict) -> bool:
    first, last = derive_names(row)
    return bool(first) and bool(last)


def _in_id_hubspot(row: dict) -> bool:
    """
    True if this contact has ANY ID HubSpot record at all -- not just a usable name.
    A contact can pass _has_complete_name() purely off Braze's FIRST_NAME/LAST_NAME
    defaults (e.g. admin@dkorinteriors.com) while having zero record in ID HubSpot,
    which means no company, no brand_origin, no trade approval -- nothing establishing
    they went through a real trade application at all.
    """
    return bool(row.get("HS_FULL_NAME") or row.get("HS_COMPANY_NAME")
                or row.get("BRAND_ORIGIN") or row.get("TRADE_APPROVAL_AT"))


def _ramp_sort_key(r: dict):
    return (r.get("LAST_CLICK") or r.get("LAST_OPEN") or r.get("FIRST_SEND_EVER")
            or datetime(1970, 1, 1, tzinfo=timezone.utc))


def _assign_ramp_and_tail(rows: list[dict], start_batch_no: int = BATCH_START_NO) -> None:
    """
    Run the ramp (batches start_batch_no..start_batch_no+11) + tail (14-18) assignment
    over exactly this set of rows, in place. Resets _batch first so a second call over
    a different subset (e.g. name-complete contacts only, or contacts left over after
    the batch-1 seed is pulled out) sizes the ramp batches off THIS pool, not whatever
    was left over from a prior call over a different set of rows.
    """
    for r in rows:
        r["_batch"] = None

    tail_lookup = {c: (n, wk, elig, note) for n, c, wk, elig, note in TAIL_BATCHES}

    # Ramp cohorts, ordered by most recent engagement first so the strongest go earliest.
    ramp: list[dict] = []
    for cohort in COHORT_ORDER:
        members = [r for r in rows if r["_cohort"] == cohort]
        members.sort(key=_ramp_sort_key, reverse=True)
        ramp.extend(members)

    idx = 0
    last_batch_no = start_batch_no - 1
    for offset, (size, week) in enumerate(zip(BATCH_SIZES, BATCH_WEEKS)):
        batch_no = start_batch_no + offset
        for r in ramp[idx:idx + size]:
            r["_batch"] = batch_no
            r["_week"] = str(week)
            r["_eligible"] = "Y"
            r["_note"] = ""
        idx += size
        last_batch_no = batch_no
    # Any ramp overflow from cohort drift joins the last ramp batch rather than vanishing.
    for r in ramp[idx:]:
        r["_batch"] = last_batch_no
        r["_week"] = str(BATCH_WEEKS[-1])
        r["_eligible"] = "Y"
        r["_note"] = "overflow - cohort counts shifted since planning"

    for r in rows:
        if r.get("_batch"):
            continue
        batch_no, week, elig, note = tail_lookup[r["_cohort"]]
        r["_batch"] = batch_no
        r["_week"] = week
        r["_eligible"] = elig
        r["_note"] = note


def assign_batches(rows: list[dict], now: datetime) -> None:
    """Attach cohort, batch, week and eligibility to every row, in place."""
    for r in rows:
        r["_cohort"] = classify(r, now)

    # Data-quality holdback, decided on two independent checks (not cohort) and pulled
    # out of the pool before the real assignment runs: an incomplete name, and/or no ID
    # HubSpot record at all (which a complete Braze-derived name can mask -- see
    # _in_id_hubspot's docstring).
    for r in rows:
        r["_hold_reason"] = ""
    holdback = [r for r in rows if not _has_complete_name(r) or not _in_id_hubspot(r)]
    eligible = [r for r in rows if _has_complete_name(r) and _in_id_hubspot(r)]
    for r in holdback:
        r["_batch"] = NO_NAME_BATCH
        r["_week"] = NO_NAME_WEEK
        r["_eligible"] = "N"
        reasons = []
        if not _has_complete_name(r):
            reasons.append("missing_name")
        if not _in_id_hubspot(r):
            reasons.append("no_id_hubspot_record")
        r["_hold_reason"] = " + ".join(reasons)
        r["_note"] = NO_NAME_NOTE

    # Batch 1: hand-picked team seed + the next TEAM_SEED_EXTERNAL_COUNT highest-priority
    # external contacts (same recency-first ordering the ramp itself uses), pulled out of
    # the pool before batches 2-13 are sized so they aren't thinned out by this carve-out.
    by_email = {r["EMAIL"].lower(): r for r in eligible}
    team_rows = []
    for email in TEAM_SEED_EMAILS:
        r = by_email.get(email.lower())
        if r is not None and r not in team_rows:
            team_rows.append(r)
    team_emails = {r["EMAIL"].lower() for r in team_rows}

    ramp_order: list[dict] = []
    for cohort in COHORT_ORDER:
        members = [r for r in eligible
                   if r["_cohort"] == cohort and r["EMAIL"].lower() not in team_emails]
        members.sort(key=_ramp_sort_key, reverse=True)
        ramp_order.extend(members)
    external_seed = ramp_order[:TEAM_SEED_EXTERNAL_COUNT]

    for r in team_rows:
        r["_batch"] = 1
        r["_week"] = TEAM_SEED_WEEK
        r["_eligible"] = "Y"
        r["_note"] = TEAM_SEED_NOTE_TEAM
    for r in external_seed:
        r["_batch"] = 1
        r["_week"] = TEAM_SEED_WEEK
        r["_eligible"] = "Y"
        r["_note"] = TEAM_SEED_NOTE_EXTERNAL

    seed_emails = team_emails | {r["EMAIL"].lower() for r in external_seed}
    remaining_eligible = [r for r in eligible if r["EMAIL"].lower() not in seed_emails]

    # The real assignment for everyone else, over name-complete contacts only, so
    # batches 2-13 come out at their target sizes (batch 2 is still 500, etc.) rather
    # than being thinned out by whoever this run happens to be missing a name, or by
    # the batch-1 seed carve-out above.
    _assign_ramp_and_tail(remaining_eligible)


# ---------------------------------------------------------------- output

def _fmt_ts(ts) -> str:
    return ts.strftime("%Y-%m-%d") if ts else ""


SHEET2_HEADERS = [
    "email", "first_name", "last_name", "city", "state", "company_name",
    "hs_company_name", "company_name_source", "contact_brand_origin",
    "contact_brand_origin_source", "braze_subscription_status",
    "marketing_eligible", "cohort", "migration_batch", "batch_week",
    "hold_reason", "notes",
]
SHEET1_HEADERS = [
    "email", "first_name", "last_name", "city", "state", "company_name",
    "contact_brand_origin", "braze_subscription_status", "in_te_hubspot",
    "has_te_platform_account", "te_contact_created_at", "migration_batch",
]
SHEET3_HEADERS = [
    "email", "first_name", "last_name", "city", "state", "company_name",
    "hs_company_name", "contact_brand_origin", "gap_type", "trade_approval_at",
    "trade_application_received_at", "in_te_hubspot",
]
SHEET4_HEADERS = [
    "email", "first_name", "last_name", "city", "state", "company_name",
    "hs_company_name", "company_name_source", "contact_brand_origin",
    "contact_brand_origin_source", "braze_subscription_status", "cohort",
    "hold_reason", "notes",
]
SHEET5_HEADERS = [
    "email", "first_name", "last_name", "city", "state", "company_name",
    "contact_brand_origin", "braze_subscription_status", "cohort", "reason",
    "migration_batch", "batch_week", "notes",
]


def build_rows(rows: list[dict], status: dict[str, str]) -> list[dict]:
    out = []
    for r in rows:
        first, last = derive_names(r)
        company, source = derive_company(r)
        city, state = derive_location(r)
        out.append({
            "email": r["EMAIL"],
            "first_name": first,
            "last_name": last,
            "city": city,
            "state": state,
            "company_name": company,
            "hs_company_name": r.get("HS_COMPANY_NAME") or "",
            "company_name_source": source,
            "contact_brand_origin": (r.get("BRAND_ORIGIN") or "").strip() or DEFAULT_BRAND_ORIGIN,
            "contact_brand_origin_source": "hubspot" if (r.get("BRAND_ORIGIN") or "").strip()
                                           else "defaulted_to_id",
            # Braze is authoritative when we have it; otherwise fall back to the
            # datashare-derived opt-out flag so the column is never empty, suffixed
            # so a reader can always tell which source a value came from.
            "braze_subscription_status": status.get(
                r["EMAIL"],
                "unsubscribed (derived)" if r.get("IS_OPTOUT_DATASHARE") else "subscribed (derived)",
            ),
            "marketing_eligible": r["_eligible"],
            "cohort": r["_cohort"],
            "migration_batch": r["_batch"],
            "batch_week": r["_week"],
            "hold_reason": r.get("_hold_reason", ""),
            "notes": r["_note"],
            "_in_te_hubspot": r.get("IN_TE_HUBSPOT"),
            "_has_te_platform": r.get("HAS_TE_PLATFORM_ACCOUNT"),
            "_te_created": r.get("TE_CREATED_AT"),
            "_optout_datashare": r.get("IS_OPTOUT_DATASHARE"),
        })
    return out


def write_xlsx(sheet2: list[dict], sheet3_raw: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="176B6B")
    note_font = Font(italic=True, color="666666")

    def add_sheet(ws, headers, records, note: str | None = None):
        header_row = 1
        if note:
            ws.append([note])
            ws.merge_cells(start_row=1, start_column=1, end_row=1,
                            end_column=max(len(headers), 1))
            ws["A1"].font = note_font
            ws["A1"].alignment = Alignment(vertical="center")
            header_row = 2
        ws.append(headers)
        for cell in ws[header_row]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for rec in records:
            ws.append([rec.get(h, "") for h in headers])
        ws.freeze_panes = f"A{header_row + 1}"
        for i, h in enumerate(headers, start=1):
            width = max(len(h) + 2, min(38, max((len(str(r.get(h, ""))) for r in records),
                                                default=10) + 2))
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

    wb = Workbook()

    # Batches that don't belong on tab 1 at all: batch 19 (data-quality hold -- lives
    # only on tab 4) and the "do not migrate" tail batches, whichever batch numbers
    # TAIL_BATCHES currently assigns to a DO_NOT_MIGRATE_COHORTS cohort (chronic
    # bouncer, unsubscribed, never-opened-in-12mo+ -- currently 16/17/18) -- those live
    # only on tab 5. Derived from the cohort mapping rather than hardcoded batch
    # numbers so it can't silently drift if TAIL_BATCHES is renumbered later.
    _TAB1_EXCLUDED_BATCHES = {NO_NAME_BATCH} | {
        batch_no for batch_no, cohort, _week, _elig, _note in TAIL_BATCHES
        if cohort in DO_NOT_MIGRATE_COHORTS
    }

    # Tab 1: Migration batches -- the main working sheet, so it's what opens first.
    # Excluded batches above aren't part of the migration plan (data-quality hold) or
    # shouldn't be marketed to at all (do-not-migrate cohorts) -- they live only on
    # tabs 4 and 5 respectively, not here too.
    ws1 = wb.active
    ws1.title = "1 - Migration batches"
    ordered = sorted((r for r in sheet2 if r["migration_batch"] not in _TAB1_EXCLUDED_BATCHES),
                      key=lambda r: (r["migration_batch"], r["email"]))
    add_sheet(ws1, SHEET2_HEADERS, ordered)

    # Tab 2: Already in The Expert -- scoped explicitly to the trade-list population.
    # (Sheet 3 has its own in_te_hubspot column answering the same question for its own,
    # disjoint population -- the note on sheet 3 below points back here.)
    ws2 = wb.create_sheet("2 - Already in TE (trade list)")
    s1 = [{
        "email": r["email"], "first_name": r["first_name"], "last_name": r["last_name"],
        "city": r["city"], "state": r["state"],
        "company_name": r["company_name"], "contact_brand_origin": r["contact_brand_origin"],
        "braze_subscription_status": r["braze_subscription_status"],
        "in_te_hubspot": "Y",
        "has_te_platform_account": "Y" if r["_has_te_platform"] else "N",
        "te_contact_created_at": _fmt_ts(r["_te_created"]),
        "migration_batch": r["migration_batch"],
    } for r in sheet2 if r["_in_te_hubspot"]]
    s1.sort(key=lambda r: (r["migration_batch"], r["email"]))
    add_sheet(ws2, SHEET1_HEADERS, s1,
              note="Scope: contacts on the ID TRADE LIST (tab 1) who are also already in "
                   "TE HubSpot. For contacts NOT on the trade list, see the in_te_hubspot "
                   "column on tab 3 instead -- these two tabs never share a row.")

    ws3 = wb.create_sheet("3 - Trade approved gaps")
    s3_note = ("Scope: ID HubSpot trade-approved contacts who are NOT on the trade list "
               "(tab 1), so they never appear on tab 2 either. Check this sheet's own "
               "in_te_hubspot column for whether any of these are already in TE.")
    s3 = []
    for r in sheet3_raw:
        full = (r.get("HS_FULL_NAME") or "").strip()
        if full:
            parts = full.split()
            first, last = (parts[0], "") if len(parts) == 1 else (" ".join(parts[:-1]), parts[-1])
        else:
            first, last = (r.get("BZ_FIRST") or ""), (r.get("BZ_LAST") or "")
        company, _ = derive_company(r)
        city, state = derive_location(r)
        s3.append({
            "email": r["EMAIL"], "first_name": first, "last_name": last,
            "city": city, "state": state,
            "company_name": company,
            "hs_company_name": r.get("HS_COMPANY_NAME") or "",
            "contact_brand_origin": (r.get("BRAND_ORIGIN") or DEFAULT_BRAND_ORIGIN).strip()
                                    or DEFAULT_BRAND_ORIGIN,
            "gap_type": r["GAP_TYPE"],
            "trade_approval_at": _fmt_ts(r.get("TRADE_APPROVAL_AT")),
            "trade_application_received_at": _fmt_ts(r.get("TRADE_APPLICATION_RECEIVED_AT")),
            "in_te_hubspot": "Y" if r.get("IN_TE_HUBSPOT") else "N",
        })
    s3.sort(key=lambda r: (r["gap_type"], r["email"]))
    add_sheet(ws3, SHEET3_HEADERS, s3, note=s3_note)

    # Tab 4: Hold - data gaps -- lens on tab 1's batch 19. Every row here also appears
    # on tab 1; hold_reason says which check failed, cohort shows engagement standing
    # independent of either gap.
    ws4 = wb.create_sheet("4 - Hold - data gaps")
    s4 = [r for r in sheet2 if r["migration_batch"] == NO_NAME_BATCH]
    s4.sort(key=lambda r: (r["hold_reason"], r["cohort"], r["email"]))
    add_sheet(ws4, SHEET4_HEADERS, s4,
              note="Scope: contacts on the trade list (tab 1, batch 19) failing one or "
                   "both data-quality checks -- missing a first name, a last name, or "
                   "both (hold_reason: missing_name), and/or no ID HubSpot record at all "
                   "(hold_reason: no_id_hubspot_record) -- a gap a complete Braze-derived "
                   "name can otherwise mask. Held out of the migration ramp entirely -- "
                   "pulled before batches 2-18 were sized, so those batches aren't "
                   "thinned out by this population -- pending a decision on whether to "
                   "send them in as-is. See the cohort column for their engagement "
                   "standing (chronic bouncer, unengaged, etc.) independent of either gap.")

    # Tab 5: Do not migrate -- lens on cohort, not batch, so it also catches contacts in
    # these cohorts who are currently sitting in batch 19 for lack of a name.
    ws5 = wb.create_sheet("5 - Do not migrate")
    s5 = [{
        "email": r["email"], "first_name": r["first_name"], "last_name": r["last_name"],
        "city": r["city"], "state": r["state"],
        "company_name": r["company_name"], "contact_brand_origin": r["contact_brand_origin"],
        "braze_subscription_status": r["braze_subscription_status"],
        "cohort": r["cohort"], "reason": DO_NOT_MIGRATE_COHORTS[r["cohort"]],
        "migration_batch": r["migration_batch"], "batch_week": r["batch_week"],
        "notes": r["notes"],
    } for r in sheet2 if r["cohort"] in DO_NOT_MIGRATE_COHORTS]
    s5.sort(key=lambda r: (r["cohort"], r["email"]))
    add_sheet(ws5, SHEET5_HEADERS, s5,
              note="Scope: chronic bouncers, Braze-unsubscribed contacts, and contacts who "
                   "have never opened or clicked a trade send in over 12 months since their "
                   "first one -- across the whole trade list, regardless of what batch "
                   "they're currently in. These should not be migrated into Klaviyo as "
                   "marketing contacts. Some also appear on tab 4 (data gaps) -- the two tabs "
                   "answer different questions and are allowed to overlap.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(ordered), len(s1), len(s3), len(s4), len(s5)


# ---------------------------------------------------------------- main

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="print distributions, write nothing")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--skip-braze-status", action="store_true",
                    help="leave subscription status as the datashare-derived value")
    ap.add_argument("--status-only", action="store_true",
                    help="only populate the Braze status cache, then exit")
    ap.add_argument("--status-limit", type=int, default=None,
                    help="cap how many Braze lookups this run performs")
    args = ap.parse_args()

    now = datetime.now(timezone.utc)

    print("Querying trade list...")
    rows = fetch(_trade_list_sql())
    print(f"  {len(rows):,} contacts on the ID trade list")

    if args.status_only:
        fetch_braze_status([r["EMAIL"] for r in rows], limit=args.status_limit)
        print("Status cache updated.")
        return

    status: dict[str, str] = {}
    if not args.skip_braze_status:
        print("Resolving Braze subscription status...")
        status = fetch_braze_status([r["EMAIL"] for r in rows], limit=args.status_limit)

    # Braze is authoritative: an unsubscribed profile lands in the transactional-only batch
    # even when the datashare shows no opt-out event.
    for r in rows:
        r["_braze_unsubscribed"] = status.get(r["EMAIL"]) == "unsubscribed"

    assign_batches(rows, now)
    built = build_rows(rows, status)

    print("\nCohort distribution:")
    for cohort, n in sorted(Counter(r["cohort"] for r in built).items()):
        print(f"  {cohort:<32}{n:>7,}")
    print("\nBatch distribution:")
    for batch, n in sorted(Counter(r["migration_batch"] for r in built).items()):
        sample = next(r for r in built if r["migration_batch"] == batch)
        print(f"  batch {batch:>2}  wk {sample['batch_week']:<14}{n:>7,}  "
              f"marketing={sample['marketing_eligible']}")
    print("\nField coverage:")
    total = len(built)
    for field in ("first_name", "last_name", "company_name", "contact_brand_origin"):
        filled = sum(1 for r in built if r[field])
        print(f"  {field:<22}{filled:>7,}  ({100*filled/total:.0f}%)")
    print("\ncontact_brand_origin:")
    for k, n in sorted(Counter(
            f"{r['contact_brand_origin']} [{r['contact_brand_origin_source']}]"
            for r in built).items()):
        print(f"  {k:<40}{n:>7,}")

    if status:
        print("\nBraze subscription status:")
        for k, n in sorted(Counter(r["braze_subscription_status"] for r in built).items()):
            print(f"  {k:<22}{n:>7,}")
        mismatch = sum(1 for r in built
                       if bool(r["_optout_datashare"]) != (r["braze_subscription_status"] == "unsubscribed"))
        print(f"  datashare/Braze opt-out disagreement: {mismatch:,}")

    print("\nQuerying sheet 3 (trade approved, not on trade list)...")
    s3 = fetch(_sheet3_sql())
    print("  " + "  ".join(f"{k}={v:,}" for k, v in
                           sorted(Counter(r["GAP_TYPE"] for r in s3).items())))

    if args.dry_run:
        print("\n--dry-run: no file written")
        return

    n_batches, n_te, n_gaps, n_noname, n_donotmigrate = write_xlsx(built, s3, args.out)
    print(f"\nWrote {args.out}")
    print(f"  tab 1 (migration batches):   {n_batches:,} rows")
    print(f"  tab 2 (already in TE):       {n_te:,} rows")
    print(f"  tab 3 (trade approved gaps): {n_gaps:,} rows")
    print(f"  tab 4 (no name - hold):      {n_noname:,} rows")
    print(f"  tab 5 (do not migrate):      {n_donotmigrate:,} rows")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Combine Braze campaign/canvas data with GA4 data into a formatted Excel lifecycle report.

Merges GA4 (Session campaign, Event count, Total revenue) with Braze (campaign/canvas analytics),
classifies into Batch & Blast, SMS, and Triggers, excludes TRADE, and outputs formatted Excel
with Summary, Batch & Blast, SMS, and Triggers sheets.

Accepts DataFrames or file paths for ga4_path, braze_path, canvas_path.
Used by generate_lifecycle_report.py for automated reporting.
"""

import re
import sys
import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Optional, List, Union, Tuple
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DESIRED_COLS = [
    "Delivered", "Confirmed Deliveries (SMS)", "Total Opens", "Total Open Rate", "Unique Opens", "Unique Open Rate",
    "Total Clicks", "Total CTR", "Total Clicks (SMS)", "Total CTR (SMS)", "Unique Clicks", "Unique CTR", "Unique CTO",
    "Orders", "Revenue", "AOV", "Conversion Rate", "$/M",
]
RAW_COLS = ["Delivered", "Confirmed Deliveries (SMS)", "Total Opens", "Unique Opens", "Total Clicks",
            "Total Clicks (SMS)", "Unique Clicks", "Orders", "Revenue"]


def pct(a, b):
    a = pd.to_numeric(a, errors="coerce")
    b = pd.to_numeric(b, errors="coerce")
    return pd.Series(np.where((b != 0) & (~pd.isna(b)), a / b, np.nan))


def extract_date_from_name(name):
    if pd.isna(name):
        return pd.NaT
    m = re.search(r'_(20\d{2})_(\d{1,2})_(\d{1,2})', str(name))
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return pd.Timestamp(year=y, month=mo, day=d)
        except ValueError:
            return pd.NaT
    return pd.NaT


def extract_t_number(step):
    if pd.isna(step):
        return 0
    m = re.search(r'_T(\d+)', str(step))
    return int(m.group(1)) if m else 0


def safe_col(df, name, fill=np.nan):
    return df[name] if name in df.columns else pd.Series([fill] * len(df), index=df.index)


def reorder(df, id_cols):
    cols = id_cols + [c for c in DESIRED_COLS if c in df.columns]
    seen = set()
    out = []
    for c in cols:
        if c not in seen:
            out.append(c)
            seen.add(c)
    return df[out]


def agg_totals(df, label_col, label_value):
    df2 = df.reindex(columns=df.columns.tolist() + [c for c in RAW_COLS if c not in df.columns])
    sums = df2.reindex(columns=RAW_COLS).apply(pd.to_numeric, errors="coerce").sum(numeric_only=True)
    out = {label_col: label_value}
    for k in RAW_COLS:
        out[k] = float(sums.get(k, np.nan)) if pd.notna(sums.get(k, np.nan)) else np.nan
    D = out.get("Delivered")
    UO = out.get("Unique Opens")
    TO = out.get("Total Opens")
    TC = out.get("Total Clicks")
    UC = out.get("Unique Clicks")
    O = out.get("Orders")
    R = out.get("Revenue")
    CD_SMS = out.get("Confirmed Deliveries (SMS)")
    TC_SMS = out.get("Total Clicks (SMS)")

    def sd(a, b):
        return (a / b) if (a is not None and b not in (None, 0, np.nan)) else np.nan

    out["Total Open Rate"] = sd(TO, D)
    out["Unique Open Rate"] = sd(UO, D)
    out["Total CTR"] = sd(TC, D)
    out["Unique CTR"] = sd(UC, D)
    out["Unique CTO"] = sd(UC, UO)
    out["Total CTR (SMS)"] = sd(TC_SMS, CD_SMS)
    out["AOV"] = sd(R, O)
    out["Conversion Rate"] = sd(O, UC)
    out["$/M"] = sd(R, D) * 1000 if sd(R, D) == sd(R, D) else np.nan
    return out


def style_sheet(ws):
    percent_cols = {"Total Open Rate", "Unique Open Rate", "Total CTR", "Unique CTR", "Unique CTO", "Total CTR (SMS)", "Conversion Rate"}
    currency_cols = {"Revenue", "AOV", "$/M"}
    integer_cols = {"Delivered", "Total Opens", "Unique Opens", "Total Clicks", "Unique Clicks", "Confirmed Deliveries (SMS)", "Total Clicks (SMS)", "Orders"}
    name_cols = {"Email Name", "SMS Name", "Canvas Name", "Step Name", "Section"}
    ws.freeze_panes = "B2"
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for c, h in enumerate(headers, start=1):
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            if h in percent_cols:
                cell.number_format = "0.00%"
            elif h in currency_cols:
                cell.number_format = "$#,##0"
            elif h in integer_cols:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(wrap_text=False, vertical="center")
    for c, h in enumerate(headers, start=1):
        if h in name_cols:
            width = 38 if h != "Section" else 16
        elif h in currency_cols:
            width = 14
        elif h in percent_cols:
            width = 10
        elif h in integer_cols:
            width = 12
        else:
            width = 12
        ws.column_dimensions[get_column_letter(c)].width = width
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and (v.startswith("Subtotal —") or v == "Total"):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).font = Font(bold=True)


def build_report(ga4_path: Union[Path, pd.DataFrame], braze_path: Union[Path, pd.DataFrame],
                 canvas_path: Optional[Union[Path, pd.DataFrame]], out_path: Path, skiprows_ga4: int = 9):
    """Build lifecycle report Excel from GA4, Braze campaign, and Braze canvas data."""
    # Load
    if isinstance(ga4_path, pd.DataFrame):
        ga4 = ga4_path.copy()
    else:
        ga4 = pd.read_csv(ga4_path, skiprows=skiprows_ga4)

    if isinstance(braze_path, pd.DataFrame):
        braze = braze_path.copy()
    else:
        braze = pd.read_csv(braze_path)

    if canvas_path is None:
        canvas = pd.DataFrame()
    elif isinstance(canvas_path, pd.DataFrame):
        canvas = canvas_path.copy()
    else:
        canvas = pd.read_csv(canvas_path)

    # Merge & exclude
    merged = pd.merge(ga4, braze, left_on="Session campaign", right_on="Campaign Name", how="left").copy()
    merged["Email Name"] = merged["Campaign Name"].fillna(merged["Session campaign"])
    merged = merged[~merged["Email Name"].astype(str).str.contains("TRADE", case=False, na=False)].copy()
    if not canvas.empty and "Canvas Name" in canvas.columns:
        canvas = canvas[~canvas["Canvas Name"].astype(str).str.contains("TRADE", case=False, na=False)].copy()

    name = merged["Email Name"].astype(str)
    is_p_em = name.str.contains(r"\bP_EM_", na=False)
    is_p_sms = name.str.contains(r"\bP_SMS_", na=False) | name.str.endswith("_SMS")
    is_p_year = name.str.contains(r"\bP_\d{4}", na=False)
    is_p_year_no_em = is_p_year & ~name.str.contains(r"\bP_EM_", na=False)

    # Batch & Blast
    bb_src = merged[(is_p_em) | (is_p_year_no_em & ~is_p_sms)].copy()
    bb_src["__date"] = bb_src["Email Name"].apply(extract_date_from_name)
    bb_src["__force_long_tail"] = (is_p_year_no_em & ~is_p_sms).reindex(bb_src.index).fillna(False)
    bb_base = pd.DataFrame({
        "Email Name": bb_src["Email Name"],
        "Delivered": safe_col(bb_src, "Deliveries (Email)"),
        "Confirmed Deliveries (SMS)": np.nan,
        "Total Opens": safe_col(bb_src, "Total Opens (Email)"),
        "Unique Opens": safe_col(bb_src, "Unique Opens (Email)"),
        "Total Clicks": safe_col(bb_src, "Total Clicks (Email)"),
        "Total Clicks (SMS)": np.nan,
        "Unique Clicks": safe_col(bb_src, "Unique Clicks (Email)"),
        "Orders": safe_col(bb_src, "Event count"),
        "Revenue": safe_col(bb_src, "Total revenue"),
        "__date": bb_src["__date"],
        "__force_long_tail": bb_src["__force_long_tail"]
    }).sort_values(["__force_long_tail", "__date", "Email Name"], kind="mergesort").reset_index(drop=True)
    bb_base["Total Open Rate"] = pct(bb_base["Total Opens"], bb_base["Delivered"])
    bb_base["Unique Open Rate"] = pct(bb_base["Unique Opens"], bb_base["Delivered"])
    bb_base["Total CTR"] = pct(bb_base["Total Clicks"], bb_base["Delivered"])
    bb_base["Unique CTR"] = pct(bb_base["Unique Clicks"], bb_base["Delivered"])
    bb_base["Unique CTO"] = pct(bb_base["Unique Clicks"], bb_base["Unique Opens"])
    bb_base["AOV"] = pct(bb_base["Revenue"], bb_base["Orders"])
    bb_base["Conversion Rate"] = pct(bb_base["Orders"], bb_base["Unique Clicks"])
    bb_base["$/M"] = pct(bb_base["Revenue"], bb_base["Delivered"]) * 1000
    is_long_tail = bb_base["Delivered"].isna() | (bb_base["Delivered"] == 0) | (bb_base["__force_long_tail"] == True)
    bb_long = bb_base[is_long_tail].drop(columns=["__date", "__force_long_tail"]).copy()
    bb_main = bb_base[~is_long_tail].drop(columns=["__date", "__force_long_tail"]).copy()
    bb_long_out = reorder(bb_long, ["Email Name"])
    bb_main_out = reorder(bb_main, ["Email Name"])
    bb_long_sub = pd.DataFrame([agg_totals(bb_long_out, "Email Name", "Subtotal — Long Tail")])
    bb_main_sub = pd.DataFrame([agg_totals(bb_main_out, "Email Name", "Subtotal — Batch & Blast")])
    bb_total = pd.DataFrame([agg_totals(pd.concat([bb_long_out, bb_main_out], ignore_index=True), "Email Name", "Total")])
    blank_bb = pd.DataFrame([{c: np.nan for c in bb_long_out.columns}])
    bb_sheet = pd.concat([
        bb_long_out, blank_bb, bb_long_sub, blank_bb,
        bb_main_out, blank_bb, bb_main_sub, blank_bb, bb_total
    ], ignore_index=True)

    # SMS
    sms_src = merged[(is_p_sms)].copy()
    sms_src["__date"] = sms_src["Email Name"].apply(extract_date_from_name)
    sms_base = pd.DataFrame({
        "SMS Name": sms_src["Email Name"],
        "Delivered": safe_col(sms_src, "Confirmed Deliveries (SMS)"),
        "Total Clicks": safe_col(sms_src, "Total Clicks (SMS)"),
        "Total Clicks (SMS)": safe_col(sms_src, "Total Clicks (SMS)"),
        "Confirmed Deliveries (SMS)": safe_col(sms_src, "Confirmed Deliveries (SMS)"),
        "Orders": safe_col(sms_src, "Event count"),
        "Revenue": safe_col(sms_src, "Total revenue"),
        "__date": sms_src["__date"],
    }).sort_values("__date", kind="mergesort").reset_index(drop=True)
    sms_base["Total CTR"] = pct(sms_base["Total Clicks"], sms_base["Delivered"])
    sms_base["Total CTR (SMS)"] = pct(sms_base["Total Clicks (SMS)"], sms_base["Confirmed Deliveries (SMS)"])
    sms_base["AOV"] = pct(sms_base["Revenue"], sms_base["Orders"])
    sms_base["Conversion Rate"] = pct(sms_base["Orders"], sms_base["Total Clicks"])
    sms_base["$/M"] = pct(sms_base["Revenue"], sms_base["Delivered"]) * 1000
    sms_out = reorder(sms_base.drop(columns=["__date"]), ["SMS Name"])
    sms_total = pd.DataFrame([agg_totals(sms_out.rename(columns={"SMS Name": "Email Name"}), "SMS Name", "Total")])
    sms_total = sms_total.reindex(columns=sms_out.columns)
    sms_sheet = pd.concat([sms_out, pd.DataFrame([{c: np.nan for c in sms_out.columns}]), sms_total], ignore_index=True)

    # Triggers: Start from canvas (Braze) and left-merge GA4 so we always show all canvas steps.
    # Previous logic started from GA4 TRG rows; if GA4 Session campaign names didn't exactly match
    # Braze Step Name, no rows matched. Now canvas is the source of truth.
    if canvas.empty or "Step Name" not in canvas.columns:
        triggers_sheet = pd.DataFrame(columns=["Canvas Name", "Step Name"] + DESIRED_COLS)
    else:
        ga4_trg = ga4[ga4["Session campaign"].astype(str).str.contains("TRG", na=False)].copy()
        # Left merge: canvas as base, join GA4 on Step Name = Session campaign
        trig = pd.merge(
            canvas, ga4_trg, left_on="Step Name", right_on="Session campaign", how="left"
        ).copy()
        # Keep only canvas rows with non-empty Canvas Name (drop any stray GA4-only rows)
        trig = trig[trig["Canvas Name"].notna() & (trig["Canvas Name"].astype(str).str.strip() != "")].copy()
        trig["Canvas Name Clean"] = trig["Canvas Name"].astype(str).str.strip().str.lower()
        trig["T_num"] = trig["Step Name"].apply(extract_t_number)
        trig = trig.sort_values(["Canvas Name Clean", "T_num", "Step Name"], kind="mergesort")
        trigger_frames = []
        for cname, grp in trig.groupby("Canvas Name Clean", sort=False):
            g = grp.copy()
            g["Canvas Name"] = g["Canvas Name"].astype(str).str.strip()
            detail = pd.DataFrame({
                "Canvas Name": g["Canvas Name"],
                "Step Name": g["Step Name"],
                "Delivered": safe_col(g, "Deliveries (Email)"),
                "Confirmed Deliveries (SMS)": safe_col(g, "Confirmed Deliveries (SMS)"),
                "Total Opens": safe_col(g, "Total Opens (Email)").fillna(safe_col(g, "Unique Opens (Email)")),
                "Unique Opens": safe_col(g, "Unique Opens (Email)"),
                "Total Clicks": safe_col(g, "Total Clicks (Email)"),
                "Total Clicks (SMS)": safe_col(g, "Total Clicks (SMS)"),
                "Unique Clicks": safe_col(g, "Unique Clicks (Email)"),
                "Orders": safe_col(g, "Event count"),
                "Revenue": safe_col(g, "Total revenue"),
            })
            detail["Total Open Rate"] = pct(detail["Total Opens"], detail["Delivered"])
            detail["Unique Open Rate"] = pct(detail["Unique Opens"], detail["Delivered"])
            detail["Total CTR"] = pct(detail["Total Clicks"], detail["Delivered"])
            detail["Unique CTR"] = pct(detail["Unique Clicks"], detail["Delivered"])
            detail["Unique CTO"] = pct(detail["Unique Clicks"], detail["Unique Opens"])
            detail["Total CTR (SMS)"] = pct(detail["Total Clicks (SMS)"], detail["Confirmed Deliveries (SMS)"])
            detail["AOV"] = pct(detail["Revenue"], detail["Orders"])
            detail["Conversion Rate"] = pct(detail["Orders"], detail["Unique Clicks"])
            detail["$/M"] = pct(detail["Revenue"], detail["Delivered"]) * 1000
            detail = reorder(detail, ["Canvas Name", "Step Name"])
            trigger_frames.append(detail)
            sub = pd.DataFrame([agg_totals(detail.rename(columns={"Canvas Name": "Email Name"}), "Canvas Name", f"Subtotal — {detail['Canvas Name'].iloc[0]}")])
            sub = sub.reindex(columns=detail.columns)
            trigger_frames.append(sub)
            trigger_frames.append(pd.DataFrame([{c: np.nan for c in detail.columns}]))
        triggers_out = pd.concat(trigger_frames, ignore_index=True) if trigger_frames else pd.DataFrame(columns=["Canvas Name", "Step Name"] + DESIRED_COLS)
        detail_only = triggers_out[~triggers_out["Canvas Name"].astype(str).str.startswith(("Subtotal —", "Total"), na=False)].dropna(how="all")
        trg_total = pd.DataFrame([agg_totals(detail_only.rename(columns={"Canvas Name": "Email Name"}), "Canvas Name", "Total")])
        trg_total = trg_total.reindex(columns=triggers_out.columns)
        triggers_sheet = pd.concat([triggers_out, trg_total], ignore_index=True)

    # Write Excel
    out_path = Path(out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        bb_sheet.to_excel(writer, index=False, sheet_name="Batch & Blast")
        sms_sheet.to_excel(writer, index=False, sheet_name="SMS")
        triggers_sheet.to_excel(writer, index=False, sheet_name="Triggers")

    # Formatting + Summary
    wb = load_workbook(out_path)
    for s in ["Batch & Blast", "SMS", "Triggers"]:
        if s in wb.sheetnames:
            style_sheet(wb[s])

    def hdr_map(ws):
        return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    def find_total_row(ws):
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=1).value == "Total":
                return r
        return None

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary", 0)
    for j, nm in enumerate(["Section"] + DESIRED_COLS, start=1):
        cell = ws_sum.cell(row=1, column=j, value=nm)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30
    ws_sum.freeze_panes = "B2"
    ws_bb = wb["Batch & Blast"]
    ws_sms = wb["SMS"]
    ws_trg = wb["Triggers"]
    h_bb = hdr_map(ws_bb)
    h_sms = hdr_map(ws_sms)
    h_trg = hdr_map(ws_trg)
    r_bb = find_total_row(ws_bb)
    r_sms = find_total_row(ws_sms)
    r_trg = find_total_row(ws_trg)

    def pull(ws_src, hmap, r, colname):
        if r is None:
            return None
        if colname not in hmap:
            return None
        return ws_src.cell(row=r, column=hmap[colname]).value

    rows = [("Batch & Blast", ws_bb, h_bb, r_bb), ("SMS", ws_sms, h_sms, r_sms), ("Triggers", ws_trg, h_trg, r_trg)]
    for i, (label, ws_src, hmap, r) in enumerate(rows, start=2):
        ws_sum.cell(row=i, column=1, value=label)
        for j, colname in enumerate(DESIRED_COLS, start=2):
            ws_sum.cell(row=i, column=j, value=pull(ws_src, hmap, r, colname))

    def find_col(colname):
        hdrs = ["Section"] + DESIRED_COLS
        return hdrs.index(colname) + 1 if colname in hdrs else None

    r_gt = 5
    ws_sum.cell(row=r_gt, column=1, value="Grand Total").font = Font(bold=True)
    raw_cols = ["Delivered", "Confirmed Deliveries (SMS)", "Total Opens", "Unique Opens", "Total Clicks", "Total Clicks (SMS)", "Unique Clicks", "Orders", "Revenue"]
    for j, nm in enumerate(DESIRED_COLS, start=2):
        if nm in raw_cols:
            total = 0.0
            for rr in [2, 3, 4]:
                v = ws_sum.cell(row=rr, column=j).value
                if isinstance(v, (int, float)):
                    total += v
            ws_sum.cell(row=r_gt, column=j, value=total).font = Font(bold=True)

    def safe_div(a, b):
        try:
            return (a / b) if (a is not None and b not in (None, 0)) else None
        except Exception:
            return None

    tot = {nm: ws_sum.cell(row=r_gt, column=find_col(nm)).value for nm in raw_cols}
    ws_sum.cell(row=r_gt, column=find_col("Total Open Rate"), value=safe_div(tot["Total Opens"], tot["Delivered"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Unique Open Rate"), value=safe_div(tot["Unique Opens"], tot["Delivered"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Total CTR"), value=safe_div(tot["Total Clicks"], tot["Delivered"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Unique CTR"), value=safe_div(tot["Unique Clicks"], tot["Delivered"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Unique CTO"), value=safe_div(tot["Unique Clicks"], tot["Unique Opens"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Total CTR (SMS)"), value=safe_div(tot["Total Clicks (SMS)"], tot["Confirmed Deliveries (SMS)"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("AOV"), value=safe_div(tot["Revenue"], tot["Orders"])).font = Font(bold=True)
    ws_sum.cell(row=r_gt, column=find_col("Conversion Rate"), value=safe_div(tot["Orders"], tot["Unique Clicks"])).font = Font(bold=True)
    den = tot["Delivered"] if tot["Delivered"] not in (None, 0) else None
    ws_sum.cell(row=r_gt, column=find_col("$/M"), value=(safe_div(tot["Revenue"], den) * 1000) if den else None).font = Font(bold=True)

    percent_cols = {"Total Open Rate", "Unique Open Rate", "Total CTR", "Unique CTR", "Unique CTO", "Total CTR (SMS)", "Conversion Rate"}
    currency_cols = {"Revenue", "AOV", "$/M"}
    integer_cols = {"Delivered", "Total Opens", "Unique Opens", "Total Clicks", "Unique Clicks", "Confirmed Deliveries (SMS)", "Total Clicks (SMS)", "Orders"}
    name_cols = {"Section"}
    for c in range(1, ws_sum.max_column + 1):
        header = ws_sum.cell(row=1, column=c).value
        if header in name_cols:
            width = 16
        elif header in currency_cols:
            width = 14
        elif header in percent_cols:
            width = 10
        elif header in integer_cols:
            width = 12
        else:
            width = 12
        ws_sum.column_dimensions[get_column_letter(c)].width = width
        for r in [2, 3, 4, 5]:
            cell = ws_sum.cell(row=r, column=c)
            if header in percent_cols:
                cell.number_format = "0.00%"
            elif header in currency_cols:
                cell.number_format = "$#,##0"
            elif header in integer_cols:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(wrap_text=False, vertical="center")
        ws_sum.cell(row=r_gt, column=c).font = Font(bold=True)
    wb.save(out_path)


if __name__ == "__main__":
    # Standalone usage: combine_braze_ga4.py ga4.csv braze.csv [canvas.csv] [--out path]
    ap = argparse.ArgumentParser(description="Combine GA4 + Braze + Canvas into formatted Excel report.")
    ap.add_argument("ga4_csv", type=str, help="GA4 CSV path")
    ap.add_argument("braze_csv", type=str, help="Braze CSV path")
    ap.add_argument("canvas_csv", type=str, nargs="?", default=None, help="Braze Canvas CSV path (optional)")
    ap.add_argument("--out", type=str, default=None, help="Output .xlsx path")
    ap.add_argument("--skiprows-ga4", type=int, default=9, help="Rows to skip for GA4 (default 9)")
    args = ap.parse_args()

    ga4_path = Path(args.ga4_csv).expanduser().resolve()
    if not ga4_path.exists():
        sys.exit(f"GA4 CSV not found: {ga4_path}")

    braze_path = Path(args.braze_csv).expanduser().resolve()
    if not braze_path.exists():
        sys.exit(f"Braze CSV not found: {braze_path}")

    canvas_path = None
    if args.canvas_csv:
        canvas_path = Path(args.canvas_csv).expanduser().resolve()
        if not canvas_path.exists():
            sys.exit(f"Canvas CSV not found: {canvas_path}")

    out_path = Path(args.out) if args.out else ga4_path.parent / "Combined_Braze_GA4_Report_FINAL.xlsx"
    build_report(ga4_path, braze_path, canvas_path, out_path, skiprows_ga4=args.skiprows_ga4)
    print(f"Report created: {out_path}")
